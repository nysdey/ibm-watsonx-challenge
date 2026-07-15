"""Load/validate Step 1's output, and write Step 2's output following the
dated-file + latest.xlsx convention in ../SCHEMA_CONTRACT.md.

Schema is validated on load and fails loudly (raises SchemaError) rather than
silently coercing or skipping — per the pipeline's global rules.
"""
import logging
import shutil
from datetime import date
from pathlib import Path

import openpyxl

log = logging.getLogger("account_tiering.schema_io")

REQUIRED_STEP1_COLUMNS = ["Account Name", "Industry"]


class SchemaError(RuntimeError):
    pass


def _load_accounts_sheet(path, sheet_names, required_cols):
    """Load (header, rows) from the first of `sheet_names` present in the workbook.
    Rows are dicts; blank-Account-Name rows are skipped. Raises SchemaError with a
    clear message on any structural problem (missing file/sheet/columns)."""
    path = Path(path)
    if not path.exists():
        raise SchemaError(f"Input not found at {path}.")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in sheet_names if s in wb.sheetnames), None)
    if sheet is None:
        wb.close()
        raise SchemaError(
            f"{path} has none of the expected sheet(s) {sheet_names} (found: {wb.sheetnames})."
        )
    ws = wb[sheet]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        wb.close()
        raise SchemaError(f"{path} sheet {sheet!r} is empty (no header row).")
    missing = [c for c in required_cols if c not in header]
    if missing:
        wb.close()
        raise SchemaError(
            f"{path} sheet {sheet!r} is missing required column(s) {missing}. "
            f"Found columns: {header}. See ../SCHEMA_CONTRACT.md."
        )
    rows = []
    for r in rows_iter:
        row = dict(zip(header, r))
        if row.get("Account Name"):
            rows.append(row)
    wb.close()
    return header, rows


def load_step1_accounts(path):
    """Returns (header, rows) for Tiering's base accounts.

    Prefers Account Segmentation's SEGMENTED_ACCOUNTS (config.SEGMENTED_LATEST,
    sheet 'Segmented Accounts') so the IBM install intel carries through the rest
    of the pipeline; falls back to raw DEDUPED (`path`, sheet 'Company Rollup')
    when Segmentation hasn't run. `path` stays the argument for backward compat.
    """
    import config
    seg = Path(config.SEGMENTED_LATEST)
    if seg.exists():
        try:
            header, rows = _load_accounts_sheet(
                seg, ["Segmented Accounts"], REQUIRED_STEP1_COLUMNS)
            if rows:
                log.info("Tiering base = SEGMENTED_ACCOUNTS (%s): %d accounts, %d cols "
                         "(IBM install intel carried through)", seg, len(rows), len(header))
                return header, rows
        except SchemaError as e:
            log.warning("Segmentation output present but unusable (%s) — falling back to DEDUPED.", e)
    header, rows = _load_accounts_sheet(path, ["Company Rollup"], REQUIRED_STEP1_COLUMNS)
    log.info("Tiering base = DEDUPED_ACCOUNTS (%s): %d accounts (Segmentation not run — "
             "no install intel to carry through)", path, len(rows))
    return header, rows


def write_dated_and_latest(rows, header, output_dir, prefix, sheet_name):
    """Writes {prefix}_{YYYYMMDD}.xlsx then overwrites latest.xlsx with a copy,
    only after the dated file is fully written — so a crash mid-write never
    leaves latest.xlsx pointing at a half-written file."""
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)
    dated_name = output_dir / f"{prefix}_{date.today().strftime('%Y%m%d')}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(header)
    for row in rows:
        ws.append([_sanitize_cell(row.get(col)) for col in header])
    wb.save(dated_name)

    latest = output_dir / "latest.xlsx"
    shutil.copyfile(dated_name, latest)
    return dated_name, latest


def _sanitize_cell(value):
    """Prevent Excel/CSV formula injection (CWE-1236) — same fix as
    ISC_Scraper_App/_internal/http_scraper.py's _sanitize_cell, since Step 2's
    input ultimately traces back to externally-scraped data too."""
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value
