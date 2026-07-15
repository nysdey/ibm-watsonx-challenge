"""Load Step 3's output (filtered to one day), write Step 4's contact-export
review file — see ../SCHEMA_CONTRACT.md."""
import datetime as _dt
from pathlib import Path

import openpyxl

REQUIRED_STEP3_COLUMNS = ["Account Name", "Planned_Call_Date", "Planned_Tier"]


class SchemaError(RuntimeError):
    pass


def _norm_date(value):
    """Canonical 'YYYY-MM-DD' for a plan date regardless of how it's stored.
    Call Planning writes an ISO *string*, but a human opening/re-saving the xlsx
    in Excel can retype it as a real date/datetime — so compare on the normalized
    day, not raw ==, or a re-saved plan silently matches 0 accounts."""
    if value is None:
        return ""
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()[:10]


def load_accounts_for_date(path, target_date_str):
    path = Path(path)
    if not path.exists():
        raise SchemaError(
            f"Step 3 output not found at {path}. Run Call_Planning/run.py first — "
            f"it writes output/latest.xlsx on every successful run."
        )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Call Plan" not in wb.sheetnames:
        raise SchemaError(
            f"{path} has no 'Call Plan' sheet (found: {wb.sheetnames}). "
            f"This isn't a Step 3 export — see ../SCHEMA_CONTRACT.md."
        )
    ws = wb["Call Plan"]
    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))

    missing = [c for c in REQUIRED_STEP3_COLUMNS if c not in header]
    if missing:
        raise SchemaError(
            f"{path} is missing required column(s) {missing}. "
            f"Found columns: {header}. See ../SCHEMA_CONTRACT.md for the Step 3 contract."
        )

    want = _norm_date(target_date_str)
    rows = []
    for r in rows_iter:
        row = dict(zip(header, r))
        if _norm_date(row.get("Planned_Call_Date")) == want:
            rows.append(row)
    wb.close()
    return rows


def load_all_step2_accounts(path):
    """Loads every account from Step 2's 'Tiered Accounts' sheet, unfiltered —
    used by manual mode (see run.py) so the user can pick accounts directly
    instead of going through Step 3's date-based plan."""
    path = Path(path)
    if not path.exists():
        raise SchemaError(
            f"Step 2 output not found at {path}. Run Account_Tiering/run.py first."
        )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Tiered Accounts" not in wb.sheetnames:
        raise SchemaError(
            f"{path} has no 'Tiered Accounts' sheet (found: {wb.sheetnames}). "
            f"This isn't a Step 2 export — see ../SCHEMA_CONTRACT.md."
        )
    ws = wb["Tiered Accounts"]
    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))
    rows = [dict(zip(header, r)) for r in rows_iter if r and r[header.index("Account Name")]]
    wb.close()
    return header, rows


def write_contact_export(rows, header, output_dir, target_date_str, sheet_name):
    """Always writes to a dated file (never overwrites a prior day's or a prior
    run's review for the same day, since these feed a human approval step and a
    silently-overwritten review file would undermine that checkpoint)."""
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)
    from datetime import datetime
    stamp = datetime.now().strftime("%H%M%S")
    path = output_dir / f"contacts_{target_date_str.replace('-', '')}_{stamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(header)
    for row in rows:
        ws.append([_sanitize_cell(row.get(col)) for col in header])
    wb.save(path)
    return path


def _sanitize_cell(value):
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value
