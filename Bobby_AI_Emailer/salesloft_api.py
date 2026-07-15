"""Salesloft REST client for Bobby — MOCKED for the WatsonX Clone.

The real client opened the saved Salesloft browser session headless, lifted the app's
bearer JWT, and called ``api.salesloft.com/v2`` read-only. Here the cadences, steps, and
enrolled people are synthesized deterministically from the shared fake-data pool — no
browser, no token, no Salesloft. The public surface Bobby depends on is preserved:
``SessionExpired`` and a ``SalesloftClient`` with ``from_session`` / ``me`` /
``list_cadences`` / ``find_cadence`` / ``people_at_email_steps`` / ``person``.
"""
import logging
import sys

import config

sys.path.insert(0, str(config.REPO_ROOT))
import fake_data  # noqa: E402

logger = logging.getLogger("salesloft_api")


class SessionExpired(Exception):
    """Kept for parity with the real client; the mock never raises it."""


def _account_names_from_isc():
    """Seed cadence people with the current demo's account companies when the ISC step
    has run, so Bobby's drafts reference the same companies as the rest of the demo."""
    path = config.REPO_ROOT / "ISC_Scraper_App" / "output" / "latest.xlsx"
    if not path.exists():
        return None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Company Rollup"] if "Company Rollup" in wb.sheetnames else wb.active
        rows = ws.iter_rows(values_only=True)
        header = list(next(rows))
        idx = header.index("Account Name") if "Account Name" in header else 2
        names = [r[idx] for r in rows if r and r[idx]]
        wb.close()
        return names or None
    except Exception:
        return None


class SalesloftClient:
    def __init__(self):
        self._people_by_id = {}   # populated by people_at_email_steps, read by person()

    @classmethod
    def from_session(cls):
        return cls()

    def me(self):
        return {"name": "Demo Seller", "email": "demo.seller@ibm.com"}

    def list_cadences(self):
        return [{"id": i + 1, "name": name} for i, name in enumerate(fake_data.SALESLOFT_CADENCES)]

    def find_cadence(self, name, cadences):
        want = (name or "").strip().lower()
        for c in cadences:
            if (c.get("name") or "").strip().lower() == want:
                return c
        for c in cadences:
            if want and want in (c.get("name") or "").strip().lower():
                return c
        return None

    def _cadence_name(self, cadence_id):
        try:
            return fake_data.SALESLOFT_CADENCES[int(cadence_id) - 1]
        except Exception:
            return fake_data.SALESLOFT_CADENCES[0]

    def people_at_email_steps(self, cadence_id):
        """Returns (email_steps, by_step). email_steps: step dicts with display_name/day/
        step_number/name. by_step: {step_id: [membership dicts]} where each membership is
        {id, person:{id, first_name, last_name}} — the shape Bobby expects."""
        cadence_name = self._cadence_name(cadence_id)
        steps = fake_data.salesloft_cadence_steps(cadence_name)
        email_steps = []
        for s in steps:
            if s["type"] != "email":
                continue
            email_steps.append({
                "id": s["id"], "day": s["day"], "step_number": s["step_number"],
                "name": s["name"], "display_name": f"Day {s['day']} · {s['name']}",
            })

        people = fake_data.salesloft_people_for_cadence(cadence_name, _account_names_from_isc())
        by_step = {s["id"]: [] for s in email_steps}
        for p in people:
            if p["step_id"] not in by_step:
                continue
            self._people_by_id[p["id"]] = {
                "id": p["id"], "first_name": p["first_name"], "last_name": p["last_name"],
                "title": p["title"], "email_address": p["email"], "person_company_name": p["company"],
            }
            by_step[p["step_id"]].append({
                "id": p["membership_id"],
                "person": {"id": p["id"], "first_name": p["first_name"], "last_name": p["last_name"]},
            })
        return email_steps, by_step

    def person(self, person_id):
        return self._people_by_id.get(person_id, {
            "first_name": "", "last_name": "", "title": "", "email_address": "",
            "person_company_name": "",
        })
