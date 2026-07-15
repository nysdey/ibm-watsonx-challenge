"""
Company-level market-intelligence rollup of the scraped Excel file.

TWO-STAGE PIPELINE (see CONTEXT.md for the full investigation history):

Stage 1 — exact-duplicate removal (location grain). Groups rows by CMR
Number (Salesforce's own Customer Master Record identifier), falling back
to a strict normalized Name+Address+City+State match when CMR is blank.
This exists purely to catch the same location's row having been fetched
twice by an overlapping territory scrape (verified: in a real 7,788-row
export, the only 20 repeated CMR Numbers were byte-for-byte identical
duplicate rows). Quantifiable fields take MAX across an exact-duplicate
group, never SUM — summing exact clones would silently double the same
real number. This stage never merges two genuinely different real
locations; a company with 24 real branches still has 24 rows after Stage 1.

Stage 2 — company-level rollup (the actual deliverable). Groups the
Stage-1 output by Account Name (falling back to Name when Account Name is
blank) — this is the real identity that matters for market intelligence
and account planning, per explicit direction: internal identifiers (CMR
Number, D&B numbers, Location ID, Company ID, Account Number) and
location-specific qualitative fields (Address, City, State/Province) are
dropped entirely from the output, since they describe one branch, not the
company, and aren't useful for account-level research. Quantifiable fields
are SUMMED across all of a company's distinct locations (Contact Count,
Employee Count, Location Annual Revenue, Total IT Spend, Cloud Spend, all
4 IBM Spend columns) to produce a true company-wide total — EXCEPT Global
Annual Revenue, which already claims to represent an aggregate/global
figure on each row (not a per-location number), so summing it across
locations would compound an already-aggregated figure; it takes MAX
instead. Technology Client Status is consolidated by priority (the most
"engaged" status found across any location wins — an existing relationship
anywhere in the company is a fact worth surfacing, not averaging away).
Coverage ID is consolidated into the distinct set of territories the
company's locations appeared under. Headquarters becomes "Yes" if any
location was flagged HQ, else "No" if any location says "No", else
"Unknown".

BACKTRACE: every Stage-2 output row carries "Distinct Locations" (how many
real, distinct locations from Stage 1 were rolled into this company row)
and "Merged From Row(s)" (the exact 1-based row numbers in the ORIGINAL
input file, flattened across both stages) — so any company row can be
traced back precisely to every source row that contributed to it, with a
hard invariant check that aborts rather than ship an unverifiable export.

Adds intel tabs with breakdowns by industry, computed from the final
company-level rows.
"""

import argparse
import re
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

# --- Stage 1: quantifiable columns consolidated via MAX (exact-duplicate
# location rows only -- see module docstring). ---
STAGE1_QUANTIFIABLE_COLUMNS = [
    "Contact Count",
    "Employee Count",
    "Location Annual Revenue",
    "Global Annual Revenue",
    "Total IT Spend",
    "Cloud Spend",
    "IBM Spend Current Year",
    "IBM Spend Prior Year",
    "IBM Spend Prior Year - 1",
    "IBM Spend Prior Year - 2",
]

# --- Stage 2: company-level rollup column roles ---

# Internal identifiers / location-specific fields that don't generalize to
# "the company" -- dropped from the final output entirely.
#
# "Account Number" is deliberately NOT dropped: it is the IBM client/buying-group
# hierarchy code (e.g. DC46JLHF / DB500LNP), constant per Account Name, and it is
# the deterministic join key the Account Segmentation step uses to map IBM
# install intel onto each account (names differ across IBM's source systems, but
# these codes are identical). See account_crosswalk.json below.
STAGE2_DROP_COLUMNS = [
    "CMR Number", "Global D&B", "Domestic D&B",
    "Location ID", "Company ID", "Address", "City", "State/Province",
]

# Per-location quantifiable metrics -- summed across all of a company's
# distinct locations to get a true company-wide total.
STAGE2_SUM_COLUMNS = [
    "Contact Count", "Employee Count", "Location Annual Revenue",
    "Total IT Spend", "Cloud Spend", "IBM Spend Current Year",
    "IBM Spend Prior Year", "IBM Spend Prior Year - 1", "IBM Spend Prior Year - 2",
]

# Already-aggregate figures -- MAX, not SUM, to avoid compounding an
# already-global number across a company's locations.
STAGE2_MAX_COLUMNS = ["Global Annual Revenue"]

# Distinct values across a company's locations are joined into one list --
# useful context (e.g. "this company spans these territories"), not a raw
# internal code.
STAGE2_UNIQUE_LIST_COLUMNS = ["Coverage ID"]

# Most "engaged" status wins -- an existing relationship at ANY location is
# a fact worth surfacing at the company level, not averaging away.
TECH_STATUS_PRIORITY = [
    "existing (continued)", "existing", "existing (py new client)",
    "existing (py-1 new client)", "existing (py-2 new client)", "existing (py-3 new client)",
    "new (active)", "new (pending)", "new (whitespace)", "new (dormant)",
]

HEADQUARTERS_COLUMN = "Headquarters"

# Everything else (Name, Account Name, Country, Industry, Sub Industry,
# Headquarters Country, LinkedIn URL, ...) is retained from the chosen
# representative row, coalescing blanks from other rows in the group.


def normalize(val):
    """Normalize a value for identity-key comparison. Critically, this
    treats 123, 123.0, and "123" as equal — openpyxl (or a human opening
    and resaving the file in real Excel) can silently retype a purely
    numeric string like a CMR Number as a float, and without this the
    identity key would split one real account into two "different" groups
    just because of a cell's underlying Excel type."""
    if isinstance(val, float) and val.is_integer():
        val = int(val)
    s = str(val if val is not None else "").strip().lower()
    s = re.sub(r'\s+', ' ', s)
    return s


def parse_number(val):
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("$", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def is_blank(val):
    return val is None or (isinstance(val, str) and val.strip() == "")


def find_col(headers, name):
    for i, h in enumerate(headers):
        if h and h.strip().lower() == name.lower():
            return i
    return None


def write_header_row(ws, headers):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def auto_width(ws, headers, row_count):
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        for row_idx in range(2, min(52, row_count + 2)):
            cell_len = len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)


def _blank_count(row):
    return sum(1 for v in row if is_blank(v))


# ============================== STAGE 1 ==============================

def _stage1_identity_key(row, row_num, cmr_col, name_col, addr_col, city_col, state_col):
    """Same exact-duplicate identity key used for location-grain dedup.
    See module docstring. An all-blank fallback key is never shared between
    rows (each gets a unique key), so two genuinely blank/incomplete rows
    are never accidentally merged with each other."""
    cmr = row[cmr_col] if cmr_col is not None and cmr_col < len(row) else None
    if not is_blank(cmr):
        return ("cmr", normalize(cmr))
    fallback_parts = (
        normalize(row[name_col]) if name_col is not None and name_col < len(row) else "",
        normalize(row[addr_col]) if addr_col is not None and addr_col < len(row) else "",
        normalize(row[city_col]) if city_col is not None and city_col < len(row) else "",
        normalize(row[state_col]) if state_col is not None and state_col < len(row) else "",
    )
    if all(p == "" for p in fallback_parts):
        return ("unique", row_num)
    return ("fallback",) + fallback_parts


def _pick_parent_index(group_rows, hq_col):
    """Prefer Headquarters == "Yes" > fewest blank fields > first row
    encountered (stable/deterministic — never arbitrary)."""
    if hq_col is not None:
        for i, row in enumerate(group_rows):
            if hq_col < len(row) and normalize(row[hq_col]) == "yes":
                return i
    best_i, best_blanks = 0, _blank_count(group_rows[0])
    for i, row in enumerate(group_rows[1:], 1):
        b = _blank_count(row)
        if b < best_blanks:
            best_i, best_blanks = i, b
    return best_i


def _stage1_merge_group(group_rows, row_numbers, headers, quant_cols, hq_col):
    """Merge a group of exact-duplicate location rows into one, MAX-ing
    quantifiable fields and coalescing blanks from every other field."""
    parent_i = _pick_parent_index(group_rows, hq_col)
    merged = list(group_rows[parent_i])

    for col_idx in range(len(headers)):
        if col_idx in quant_cols:
            continue
        if col_idx < len(merged) and is_blank(merged[col_idx]):
            for row in group_rows:
                if col_idx < len(row) and not is_blank(row[col_idx]):
                    merged[col_idx] = row[col_idx]
                    break

    for col_idx in quant_cols:
        if col_idx >= len(headers):
            continue
        best = None
        for row in group_rows:
            n = parse_number(row[col_idx] if col_idx < len(row) else None)
            if n is not None and (best is None or n > best):
                best = n
        if col_idx < len(merged):
            merged[col_idx] = best if best is not None else ""

    return merged, row_numbers


def dedup_exact_duplicates(rows_with_rownum, headers):
    """Stage 1. Returns list of (merged_row, [source_row_numbers]) in
    first-seen order -- one entry per real, distinct location."""
    cmr_col = find_col(headers, "CMR Number")
    name_col = find_col(headers, "Name")
    addr_col = find_col(headers, "Address")
    city_col = find_col(headers, "City")
    state_col = find_col(headers, "State/Province")
    hq_col = find_col(headers, HEADQUARTERS_COLUMN)

    quant_cols = {i for i, h in enumerate(headers)
                  if h and normalize(h) in {normalize(q) for q in STAGE1_QUANTIFIABLE_COLUMNS}}

    groups = {}
    order = []
    for row_num, row in rows_with_rownum:
        key = _stage1_identity_key(row, row_num, cmr_col, name_col, addr_col, city_col, state_col)
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append((row_num, row))

    results = []
    for key in order:
        entries = groups[key]
        row_nums = [e[0] for e in entries]
        group_rows = [e[1] for e in entries]
        merged, src_rows = _stage1_merge_group(group_rows, row_nums, headers, quant_cols, hq_col)
        results.append((merged, src_rows))
    return results


# ============================== STAGE 2 ==============================

def _stage2_identity_key(row, name_col, acct_col):
    """Company identity: Account Name, falling back to Name when blank.
    This is deliberately the ONLY key at this stage -- market intelligence
    and account planning care about the company, not any internal code."""
    acct = row[acct_col] if acct_col is not None and acct_col < len(row) else None
    if not is_blank(acct):
        return normalize(acct)
    name = row[name_col] if name_col is not None and name_col < len(row) else None
    return normalize(name)


def _consolidate_status(values):
    present = [v for v in values if not is_blank(v)]
    if not present:
        return ""
    def rank(v):
        n = normalize(v)
        return TECH_STATUS_PRIORITY.index(n) if n in TECH_STATUS_PRIORITY else len(TECH_STATUS_PRIORITY)
    return sorted(present, key=rank)[0]


def _consolidate_headquarters(values):
    norm = [normalize(v) for v in values]
    if "yes" in norm:
        return "Yes"
    if "no" in norm:
        return "No"
    return "Unknown"


def rollup_by_company(stage1_rows, headers):
    """Stage 2. stage1_rows: list of (row, [source_row_numbers]).
    Returns (out_headers, out_rows) where out_rows are fully-built final
    rows (already includes Distinct Locations + Merged From Row(s))."""
    name_col = find_col(headers, "Name")
    acct_col = find_col(headers, "Account Name")

    drop_idx = {find_col(headers, c) for c in STAGE2_DROP_COLUMNS if find_col(headers, c) is not None}
    sum_idx = {find_col(headers, c) for c in STAGE2_SUM_COLUMNS if find_col(headers, c) is not None}
    max_idx = {find_col(headers, c) for c in STAGE2_MAX_COLUMNS if find_col(headers, c) is not None}
    unique_list_idx = {find_col(headers, c) for c in STAGE2_UNIQUE_LIST_COLUMNS if find_col(headers, c) is not None}
    status_idx = find_col(headers, "Technology Client Status")
    hq_idx = find_col(headers, HEADQUARTERS_COLUMN)

    groups = {}
    order = []
    for row, src_rows in stage1_rows:
        key = _stage2_identity_key(row, name_col, acct_col)
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append((row, src_rows))

    keep_idx = [i for i in range(len(headers)) if i not in drop_idx]
    out_headers = [headers[i] for i in keep_idx] + ["Distinct Locations", "Merged From Row(s)"]

    out_rows = []
    for key in order:
        entries = groups[key]
        group_rows = [e[0] for e in entries]
        all_src_rows = [r for e in entries for r in e[1]]

        parent_i = _pick_parent_index(group_rows, hq_idx)
        parent = group_rows[parent_i]

        built = []
        for i in keep_idx:
            if i in sum_idx:
                total = sum(n for n in (parse_number(r[i] if i < len(r) else None) for r in group_rows) if n is not None)
                any_present = any(not is_blank(r[i] if i < len(r) else None) for r in group_rows)
                built.append(total if any_present else "")
            elif i in max_idx:
                vals = [parse_number(r[i] if i < len(r) else None) for r in group_rows]
                vals = [v for v in vals if v is not None]
                built.append(max(vals) if vals else "")
            elif i in unique_list_idx:
                seen = []
                for r in group_rows:
                    v = r[i] if i < len(r) else None
                    if not is_blank(v) and v not in seen:
                        seen.append(v)
                built.append(";".join(str(v) for v in seen))
            elif i == status_idx:
                built.append(_consolidate_status([r[i] if i < len(r) else None for r in group_rows]))
            elif i == hq_idx:
                built.append(_consolidate_headquarters([r[i] if i < len(r) else None for r in group_rows]))
            else:
                val = parent[i] if i < len(parent) else None
                if is_blank(val):
                    for r in group_rows:
                        if i < len(r) and not is_blank(r[i]):
                            val = r[i]
                            break
                built.append(val)

        built.append(len(group_rows))
        built.append(";".join(str(r) for r in sorted(set(all_src_rows))))
        out_rows.append((built, all_src_rows))

    return out_headers, out_rows


def _norm_cust_number(val):
    """Normalize an IBM customer / CMR number to a comparable canonical form:
    keep the numeric part before any country-code suffix, drop leading zeros.
    e.g. '0139800' -> '139800', '5788184-897' -> '5788184'. Returns '' if it
    isn't a usable (>=5 digit) number."""
    s = str(val if val is not None else "").strip().split("-")[0]
    s = re.sub(r"\D", "", s).lstrip("0")
    return s if len(s) >= 5 else ""


def build_crosswalk(stage1_rows, headers):
    """Deterministic identifier -> account-key crosswalk for Segmentation.

    Each stage-1 row is one location carrying its own CMR (IBM customer) number
    and its Account Number (the account/buying-group hierarchy code). One account
    has many customer numbers, so this location-grain map is the ONLY link the
    company-level rollup can't reproduce -- and it's exactly what lets the Cloud
    and ISM/Competitive install files (which expose an IBM customer number but a
    *different* name spelling) map back onto the right account with 100%
    certainty instead of by fuzzy name. Account Number itself and the account
    name are on the rollup already, so Segmentation derives those directly.
    """
    acct_col = find_col(headers, "Account Number")
    cmr_col = find_col(headers, "CMR Number")
    cust_to_key, keys = {}, set()
    for row, _src in stage1_rows:
        key = str(row[acct_col]).strip() if acct_col is not None and acct_col < len(row) and not is_blank(row[acct_col]) else ""
        if not key:
            continue
        keys.add(key)
        cust = _norm_cust_number(row[cmr_col]) if cmr_col is not None and cmr_col < len(row) else ""
        if cust:
            # First writer wins; collisions (same customer no under two accounts)
            # are vanishingly rare and would only affect that one number.
            cust_to_key.setdefault(cust, key)
    return {"keys": sorted(keys), "cust_to_key": cust_to_key}


def run_dedup(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    n_cols = len(headers)

    # Read all rows, tagging each with its real Excel row number for
    # backtrace, and normalize length so a short/ragged row can never
    # shift values into the wrong column downstream.
    all_rows = []
    for excel_row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = list(row)
        if len(row) < n_cols:
            row = row + [None] * (n_cols - len(row))
        elif len(row) > n_cols:
            row = row[:n_cols]
        all_rows.append((excel_row_idx, row))

    print(f"  Original: {len(all_rows)} rows")

    stage1_rows = dedup_exact_duplicates(all_rows, headers)
    print(f"  Stage 1 (exact-duplicate locations removed): {len(stage1_rows)} distinct locations ({len(all_rows) - len(stage1_rows)} exact duplicates removed)")

    out_headers, out_rows = rollup_by_company(stage1_rows, headers)
    print(f"  Stage 2 (rolled up to company level): {len(out_rows)} unique companies")

    # Backtrace invariant check: every original source row must be
    # accounted for exactly once across the final output. If this ever
    # fails, something in the two-stage grouping is dropping or duplicating
    # rows -- surface it loudly rather than silently ship a corrupted export.
    seen = set()
    for _, src_rows in out_rows:
        for r in src_rows:
            if r in seen:
                raise AssertionError(f"Dedup backtrace mismatch: source row {r} appears in more than one company group.")
            seen.add(r)
    expected = {r for r, _ in all_rows}
    if seen != expected:
        missing = expected - seen
        raise AssertionError(f"Dedup backtrace mismatch: source row(s) {sorted(missing)[:10]} never appeared in any company group.")

    # === Write main rolled-up sheet ===
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Company Rollup"

    write_header_row(ws_out, out_headers)
    for row_idx, (built, _) in enumerate(out_rows, 2):
        for col_idx, value in enumerate(built, 1):
            ws_out.cell(row=row_idx, column=col_idx, value=value)

    ws_out.auto_filter.ref = ws_out.dimensions
    ws_out.freeze_panes = "A2"
    auto_width(ws_out, out_headers, len(out_rows))

    # === Intel Tab: Accounts by Industry (company-level now) ===
    out_industry_col = find_col(out_headers, "Industry")
    if out_industry_col is not None:
        ind_counts = defaultdict(int)
        for built, _ in out_rows:
            industry = str(built[out_industry_col] or "").strip() or "Unknown"
            ind_counts[industry] += 1

        ws_ind = wb_out.create_sheet("Companies by Industry")
        ind_headers = ["Industry", "Company Count"]
        write_header_row(ws_ind, ind_headers)
        for row_idx, (ind, count) in enumerate(sorted(ind_counts.items(), key=lambda x: -x[1]), 2):
            ws_ind.cell(row=row_idx, column=1, value=ind)
            ws_ind.cell(row=row_idx, column=2, value=count)
        ws_ind.auto_filter.ref = ws_ind.dimensions
        ws_ind.freeze_panes = "A2"
        auto_width(ws_ind, ind_headers, len(ind_counts))
        print(f"  Companies by Industry: {len(ind_counts)} industries")

    wb_out.save(output_path)
    print(f"  Saved to {output_path}")

    # === Deterministic account crosswalk for the Segmentation step ===
    # customer-number -> account-key, built from the location-grain stage-1 rows
    # (the rollup can't carry this, one account has many customer numbers).
    try:
        import json
        crosswalk = build_crosswalk(stage1_rows, headers)
        xw_path = Path(output_path).parent / "account_crosswalk.json"
        xw_path.write_text(json.dumps(crosswalk))
        print(f"  Account crosswalk: {len(crosswalk['keys'])} account keys, "
              f"{len(crosswalk['cust_to_key'])} customer-number links -> {xw_path.name}")
    except Exception as e:
        print(f"  (warning) could not write account_crosswalk.json: {e}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    run_dedup(args.input, args.output)


if __name__ == "__main__":
    main()
