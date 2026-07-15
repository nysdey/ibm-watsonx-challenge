"""Deterministic fake-data generator for the WatsonX Clone demo.

Everything the real dashboard would have pulled over the network — IBM ISC
(Salesforce) accounts, the five IBM install-base files, ZoomInfo revenue/employee
enrichment, Google-News buying signals, and Salesloft cadences/people — is
synthesized here instead. No network, no browser, no credentials.

Design contract that keeps the pipeline honest:

  * Everything is seeded (hashlib, NOT the salted built-in ``hash()``), so a given
    set of CovIDs always produces the same accounts, and a given account name
    always produces the same identity (hierarchy code, customer number, revenue,
    install rows). That stability is what lets Account Segmentation deterministically
    JOIN the install files back onto the ISC accounts by exact IBM key — exactly as
    it does with real data.
  * This module returns plain Python rows/dicts. A few convenience writers
    (``generate_isc_output``) do workbook I/O because more than one caller needs the
    identical bytes; the per-install-file writers live in each sub-scraper.

The account pool is realistic-but-synthetic: procedural company names across IBM's
real vertical taxonomy (so Account Tiering's Core/Adjacent/Baseline vertical-fit and
Technology-Client-Status scoring spread accounts across all three tiers).
"""
import hashlib
import json
import math
import random
from datetime import date, datetime, timedelta
from pathlib import Path

# ── Deterministic RNG ─────────────────────────────────────────────────────────
# Built-in hash() is randomized per process (PYTHONHASHSEED), which would make the
# "same CovIDs -> same accounts" contract silently break across restarts. md5 of a
# stable string is reproducible forever.


def _rng(*parts):
    key = "|".join(str(p) for p in parts)
    digest = hashlib.md5(key.encode("utf-8")).hexdigest()
    return random.Random(int(digest, 16))


# ── Word banks for procedural, believable company names ───────────────────────
_PREFIXES = [
    "Meridian", "Cascade", "Summit", "Pinnacle", "Vanguard", "Keystone", "Horizon",
    "Atlas", "Beacon", "Sterling", "Granite", "Vertex", "Cardinal", "Sentinel",
    "Apex", "Redwood", "Northwind", "Bluepeak", "Ironwood", "Copperline", "Silverline",
    "Evergreen", "Highland", "Riverstone", "Fairview", "Brightwater", "Clearfield",
    "Oakmont", "Westbridge", "Kingsford", "Lakeshore", "Stonegate", "Ridgeline",
    "Harborview", "Crestwood", "Monarch", "Liberty", "Frontier", "Pacific", "Atlantic",
    "Continental", "Premier", "Allied", "Sunrise", "Trilliant", "Amberton", "Foxglove",
    "Grandview", "Templeton", "Wexford",
]

# (industry, weight, [sub-industries], [industry nouns]). Industry strings match the
# exact values Account_Tiering keys its vertical-fit scoring on.
_INDUSTRIES = [
    ("Healthcare", 3, ["Hospital Systems", "Providers", "Payers"],
     ["Health", "Healthcare", "Medical", "Care", "Wellness"]),
    ("Banking", 3, ["Retail Banking", "Commercial Banking"],
     ["Bank", "Financial", "Bancorp", "Savings", "Trust"]),
    ("Insurance", 2, ["Property & Casualty", "Life", "Health Plans"],
     ["Insurance", "Mutual", "Assurance", "Casualty"]),
    ("Financial Markets", 2, ["Asset Management", "Capital Markets"],
     ["Capital", "Securities", "Asset Management", "Investments", "Advisors"]),
    ("Government, State/Provincial/Local", 2, ["State Agency", "Municipal"],
     ["County", "State", "Municipal", "Regional Authority", "Water District"]),
    ("Government, Central/Federal", 1, ["Federal Agency"],
     ["Federal Services", "National Bureau"]),
    ("Telecommunications", 2, ["Wireless", "Broadband"],
     ["Telecom", "Communications", "Networks", "Wireless", "Broadband"]),
    ("Life Sciences", 2, ["Pharmaceuticals", "Biotech"],
     ["Biosciences", "Pharma", "Therapeutics", "Labs", "Biotech"]),
    ("Retail", 2, ["Specialty Retail", "Grocery"],
     ["Retail", "Markets", "Stores", "Brands"]),
    ("Manufacturing", 3, ["Industrial", "Discrete"],
     ["Manufacturing", "Industries", "Fabrication", "Works"]),
    ("Energy & Utilities", 2, ["Utilities", "Oil & Gas"],
     ["Energy", "Power", "Utilities", "Resources"]),
    ("Transportation", 1, ["Logistics", "Airlines"],
     ["Logistics", "Freight", "Transport", "Cargo"]),
    ("Consumer Products", 1, ["CPG"],
     ["Consumer", "Products", "Goods"]),
    ("Media & Entertainment", 1, ["Broadcast", "Publishing"],
     ["Media", "Broadcasting", "Entertainment", "Publishing"]),
    ("Automotive", 1, ["OEM", "Parts"],
     ["Motors", "Automotive", "Mobility"]),
    ("Education", 1, ["Higher Education"],
     ["University", "College", "Academy"]),
]

_SUFFIXES = ["Holdings", "Group", "Corporation", "Inc.", "Systems", "Partners",
             "LLC", "Company", "Industries", "Enterprises", "Services", "", ""]

# Technology Client Status weights — the strings Account_Tiering scores (relationship
# base 100 down to 15). A whitespace-heavy mix keeps the tier split realistic.
_TECH_STATUS = [
    ("Existing (Continued)", 22), ("Existing", 14), ("Existing (PY New Client)", 8),
    ("New (Active)", 12), ("New (Pending)", 10), ("New (Whitespace)", 24),
    ("New (Dormant)", 10),
]

_STATES = ["CA", "TX", "NY", "IL", "FL", "WA", "MA", "GA", "NC", "OH", "PA", "CO",
           "VA", "NJ", "MN", "AZ", "MI", "OR", "TN", "MD"]
_CITIES = ["Springfield", "Riverside", "Franklin", "Clinton", "Georgetown", "Salem",
           "Fairfield", "Madison", "Arlington", "Ashland", "Dover", "Auburn", "Bristol",
           "Manchester", "Newport", "Oakland", "Kingston", "Milford", "Winchester"]


def _weighted_choice(rng, pairs):
    total = sum(w for _, w in pairs)
    x = rng.uniform(0, total)
    upto = 0
    for item, w in pairs:
        upto += w
        if x <= upto:
            return item
    return pairs[-1][0]


def _hierarchy_code(name):
    """IBM client/buying-group hierarchy code: 2 letters + 6 base36 chars, matching
    Account_Segmentation's ``^[A-Z]{2}[0-9A-Z]{6}$``. Stable per account name — this
    is the primary key Segmentation joins install files on."""
    r = _rng("code", name)
    alphabet = "0123456789ABCDEFGHIJKLMNPQRSTUVWXYZ"
    return r.choice(["GC", "DC", "GB", "DB"]) + "".join(r.choice(alphabet) for _ in range(6))


def _customer_number(name):
    """IBM customer / CMR number (numeric, >=5 digits). Stable per account name."""
    return str(_rng("cust", name).randint(100000, 9999999))


def _employees(r):
    return int(10 ** r.uniform(1.8, 5.3))


def _revenue(r, employees):
    if employees:
        return float(round(employees * r.uniform(90_000, 460_000), -3))
    return float(round(10 ** r.uniform(6.6, 10.4), -3))


def _domain(name):
    core = "".join(ch for ch in name.lower() if ch.isalnum())[:18]
    return f"{core}.com" if core else ""


def _make_account(name, industry, sub_industry):
    """Full, stable account identity derived purely from the name (+ its industry,
    fixed at first sight). Everything a real ISC 'Company Rollup' row carries."""
    r = _rng("account", name)
    status = _weighted_choice(r, _TECH_STATUS)
    is_existing = status.startswith("Existing")
    employees = None if r.random() < 0.16 else _employees(r)
    location_rev = None if r.random() < 0.16 else _revenue(r, employees or _employees(_rng("emp2", name)))
    base_rev = location_rev or _revenue(r, employees or 500)
    global_rev = float(round(base_rev * r.uniform(1.0, 2.6), -3))
    total_it = float(round(base_rev * r.uniform(0.02, 0.09), -3))
    cloud_spend = float(round(total_it * r.uniform(0.1, 0.5), -3))

    # IBM spend: existing clients carry real spend with a trend; whitespace/dormant ~0.
    if status.startswith("New (Whitespace)") or status.startswith("New (Dormant)"):
        cur = 0.0
        prior = prior1 = prior2 = 0.0
    else:
        cur = float(round((0.0006 if is_existing else 0.0002) * base_rev * r.uniform(0.4, 2.4), -2))
        trend = r.choice([1.25, 1.1, 1.0, 0.85, 0.7])   # up / flat / down
        prior = float(round(cur / trend, -2)) if cur else 0.0
        prior1 = float(round(prior * r.uniform(0.75, 1.15), -2))
        prior2 = float(round(prior1 * r.uniform(0.75, 1.15), -2))

    st = r.choice(_STATES)
    return {
        "name": name,
        "account_name": name,
        "account_number": _hierarchy_code(name),
        "customer_number": _customer_number(name),
        "industry": industry,
        "sub_industry": sub_industry,
        "country": "United States",
        "state": st,
        "city": r.choice(_CITIES),
        "headquarters": r.choice(["Yes", "Yes", "No", "Unknown"]),
        "headquarters_country": "United States",
        "employees": employees,
        "location_revenue": location_rev,
        "global_revenue": global_rev,
        "total_it_spend": total_it,
        "cloud_spend": cloud_spend,
        "contact_count": r.randint(1, 42),
        "tech_client_status": status,
        "ibm_spend_current": cur,
        "ibm_spend_prior": prior,
        "ibm_spend_prior1": prior1,
        "ibm_spend_prior2": prior2,
        "linkedin": f"https://www.linkedin.com/company/{_domain(name).split('.')[0]}" if name else "",
        "distinct_locations": r.randint(1, 7),
        "coverage_ids": [],   # filled by accounts_for_covids
    }


def _gen_name(rng):
    """Pick a company (name, industry, sub_industry) using the covid-seeded rng."""
    industry, _w, subs, nouns = _weighted_choice(
        rng, [((ind, w, subs, nouns), w) for ind, w, subs, nouns in _INDUSTRIES])
    prefix = rng.choice(_PREFIXES)
    noun = rng.choice(nouns)
    suffix = rng.choice(_SUFFIXES)
    name = " ".join(p for p in (prefix, noun, suffix) if p).strip()
    return name, industry, rng.choice(subs)


# ── The account pool ──────────────────────────────────────────────────────────

def demo_covids_for_email(email):
    """A stable demo territory (set of CovIDs) for any signed-in email, so the demo
    never dead-ends when the email isn't a real rep in Name Match.xlsx. Deterministic
    per email — the same person always gets the same territory."""
    r = _rng("demo_covids", (email or "demo").strip().lower())
    covids = set()
    while len(covids) < r.randint(11, 17):
        covids.add(f"T{r.randint(10000, 99999):05d}")
    return sorted(covids)


def accounts_for_covids(covids):
    """Deterministic list of account dicts for a set of CovIDs. A company that falls
    in more than one CovID's territory appears once, with every CovID accumulated in
    ``coverage_ids`` (exactly like the real ISC company rollup)."""
    seen = []
    covids = list(dict.fromkeys(str(c).strip() for c in covids if str(c).strip()))
    by_name = {}
    order = []
    for covid in covids:
        rng = _rng("covid", covid)
        for _ in range(rng.randint(9, 26)):
            name, industry, sub = _gen_name(rng)
            if name in by_name:
                if covid not in by_name[name]["coverage_ids"]:
                    by_name[name]["coverage_ids"].append(covid)
                continue
            acct = _make_account(name, industry, sub)
            acct["coverage_ids"] = [covid]
            by_name[name] = acct
            order.append(name)
    return [by_name[n] for n in order]


# ── ISC "Company Rollup" output (Step 1 handoff) ──────────────────────────────
# 23 columns, exact order/casing from the real dedup.rollup_by_company — including
# the trailing space on "Global Annual Revenue " (a real quirk downstream tolerates
# but we reproduce verbatim so the handoff is byte-faithful).
COMPANY_ROLLUP_HEADERS = [
    "Name", "Country", "Account Name", "Account Number", "Coverage ID",
    "Technology Client Status", "Contact Count", "Industry", "Sub Industry",
    "Employee Count", "Location Annual Revenue", "Global Annual Revenue ",
    "Total IT Spend", "Cloud Spend", "Headquarters", "Headquarters Country",
    "LinkedIn URL", "IBM Spend Current Year", "IBM Spend Prior Year",
    "IBM Spend Prior Year - 1", "IBM Spend Prior Year - 2",
    "Distinct Locations", "Merged From Row(s)",
]


def _rollup_row(acct, row_no):
    r = _rng("rollup", acct["name"])
    n_loc = acct["distinct_locations"]
    merged = ", ".join(str(row_no * 3 + i) for i in range(n_loc))
    return [
        acct["name"], acct["country"], acct["account_name"], acct["account_number"],
        ";".join(acct["coverage_ids"]), acct["tech_client_status"], acct["contact_count"],
        acct["industry"], acct["sub_industry"], acct["employees"], acct["location_revenue"],
        acct["global_revenue"], acct["total_it_spend"], acct["cloud_spend"],
        acct["headquarters"], acct["headquarters_country"], acct["linkedin"],
        acct["ibm_spend_current"], acct["ibm_spend_prior"], acct["ibm_spend_prior1"],
        acct["ibm_spend_prior2"], n_loc, merged,
    ]


def company_rollup(accounts):
    return COMPANY_ROLLUP_HEADERS, [_rollup_row(a, i + 1) for i, a in enumerate(accounts)]


def industry_counts(accounts):
    counts = {}
    for a in accounts:
        counts[a["industry"]] = counts.get(a["industry"], 0) + 1
    rows = sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))
    return ["Industry", "Company Count"], [[ind, n] for ind, n in rows]


def crosswalk(accounts):
    """account_crosswalk.json shape: keys = hierarchy codes; cust_to_key maps
    normalized customer number -> hierarchy code."""
    keys = sorted({a["account_number"] for a in accounts})
    cust_to_key = {}
    for a in accounts:
        cust_to_key.setdefault(a["customer_number"], a["account_number"])
    return {"keys": keys, "cust_to_key": cust_to_key}


# ── Workbook writer for the ISC step (used by the dashboard's Get My Accounts) ──

def _write_sheet(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append(row)


def write_workbook(path, sheet, headers, rows):
    """Write a single-sheet .xlsx. Shared by the IBM install sub-scrapers."""
    import openpyxl
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    _write_sheet(ws, headers, rows)
    wb.save(path)


def generate_isc_output(covids, isc_output_dir, run_id=None):
    """Write the exact ISC handoff artifacts a real scrape would: latest.xlsx
    (Company Rollup + Companies by Industry), selected_covids.json, and
    account_crosswalk.json. Returns the account list (so the caller can report a
    count). Pure local — this stands in for the whole Aura/Playwright scrape."""
    import openpyxl
    isc_output_dir = Path(isc_output_dir)
    isc_output_dir.mkdir(parents=True, exist_ok=True)
    accounts = accounts_for_covids(covids)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Company Rollup"
    _write_sheet(ws, *company_rollup(accounts))
    ws2 = wb.create_sheet("Companies by Industry")
    _write_sheet(ws2, *industry_counts(accounts))
    wb.save(isc_output_dir / "latest.xlsx")

    run_id = run_id or datetime.now().strftime("%Y%m%d_%H%M%S")
    (isc_output_dir / "selected_covids.json").write_text(json.dumps({
        "run_id": run_id,
        "selected_at": datetime.now().isoformat(timespec="seconds"),
        "covids": list(dict.fromkeys(str(c).strip() for c in covids if str(c).strip())),
    }, indent=2))
    (isc_output_dir / "account_crosswalk.json").write_text(json.dumps(crosswalk(accounts), indent=2))
    return accounts


# ── IBM install-base rows (Step 2) ────────────────────────────────────────────
# Each install file carries join keys (hierarchy code / customer number / exact
# name) for a deterministic SUBSET of the accounts, so Account Segmentation's exact
# join actually attaches installs (and tier diversity comes out right). Coverage of
# the pool per file: cloud ~45%, power ~40%, storage ~30%, non-infra ~35%, comp ~22%.


def _subset(accounts, tag, fraction):
    out = []
    for a in accounts:
        if _rng("subset", tag, a["name"]).random() < fraction:
            out.append(a)
    return out


CLOUD_HEADERS = ["Coverage ID", "Account Name", "Unique Account Key",
                 "IBM Customer Number with Country Code", "Global Buying Group ID",
                 "Seller Name", "Ledger Market Name", "Trailing 12M Cloud Spend"]

_SELLERS = ["A. Rivera", "J. Chen", "M. Okafor", "S. Patel", "L. Nguyen", "D. Brooks"]
_LEDGERS = ["US Commercial", "US Enterprise", "US Public", "US Financial Services"]


def cloud_install(accounts, covids):
    rows = []
    for a in _subset(accounts, "cloud", 0.45):
        r = _rng("cloudrow", a["name"])
        cov = a["coverage_ids"][0] if a["coverage_ids"] else (covids[0] if covids else "T0000000")
        rows.append([
            cov, a["account_name"], f"{a['account_number']}-{cov}-{r.randint(100, 899)}",
            f"{a['customer_number']}-897", a["account_number"], r.choice(_SELLERS),
            r.choice(_LEDGERS), float(round(r.uniform(20_000, 3_000_000), -2)),
        ])
    return CLOUD_HEADERS, rows


# Power's real export is 128 wide; downstream only reads the join columns + attaches
# the rest, so a believable ~14-col subset with the join keys is faithful enough.
POWER_HEADERS = ["Source Check", "Count", "MACHINE TYPE", "MACHINE MODEL", "MTM",
                 "MACHINE SERIAL", "Primary Customer Name", "domestic client id",
                 "global buying group id", "Customer Number", "SAP CUST Number",
                 "Product Description", "local coverage type id", "Install Country"]
_POWER_MTM = ["9080-HEX", "9043-MRX", "9105-22B", "8286-42A", "9223-42S"]


def power_install(accounts, covids):
    rows = []
    for a in _subset(accounts, "power", 0.40):
        r = _rng("powerrow", a["name"])
        cov = a["coverage_ids"][0] if a["coverage_ids"] else (covids[0] if covids else "T0000000")
        for _ in range(r.randint(1, 3)):
            mtm = r.choice(_POWER_MTM)
            rows.append([
                "OK", 1, mtm.split("-")[0], mtm.split("-")[1], mtm,
                f"{r.randint(10, 99)}-{r.randint(10000, 99999)}", a["account_name"],
                a["account_number"], a["account_number"], a["customer_number"],
                a["customer_number"], "IBM Power System", cov, "United States",
            ])
    return POWER_HEADERS, rows


STORAGE_HEADERS = ["Storage Category", "dom_client_id", "gbl_buy_grp_id", "cust_no",
                   "Account Name", "Machine Type", "Model", "Serial", "Install Country"]
_STORAGE_CATS = ["FlashSystem + SVC", "Tape", "DS8K", "SAN"]
_STORAGE_MT = ["2145-SV3", "3958-DD6", "5341-F04", "2107-996", "4666-24H"]


def storage_install(accounts, covids):
    rows = []
    for a in _subset(accounts, "storage", 0.30):
        r = _rng("storagerow", a["name"])
        for _ in range(r.randint(1, 2)):
            mt = r.choice(_STORAGE_MT)
            rows.append([
                r.choice(_STORAGE_CATS), a["account_number"], a["account_number"],
                a["customer_number"], a["account_name"], mt.split("-")[0], mt.split("-")[1],
                f"{r.randint(10000, 99999)}", "United States",
            ])
    return STORAGE_HEADERS, rows


NON_INFRA_HEADERS = ["L1_ACCOUNT_NAME", "L2_ACCOUNT_NAME", "CUST_NAME",
                     "UTLEVEL10DESCRIPTION", "UTLEVEL15DESCRIPTION", "UTLEVEL17DESCRIPTION",
                     "UTLEVEL20DESCRIPTION", "UTLEVEL30DESCRIPTION", "QTY", "ORDER_DATE",
                     "INSTALL_DATE", "WARRANTY_END_DATE", "SALES_ORDER_NO", "CUST_NO",
                     "ISSUING_COUNTRY_NAME"]
_NONINFRA_PRODUCTS = [
    ("Software", "Data & AI", "Db2", "Db2 Enterprise Server Edition"),
    ("Software", "Automation", "IBM Cloud Pak", "Cloud Pak for Integration"),
    ("Software", "Security", "QRadar", "QRadar SIEM"),
    ("Software", "Data & AI", "watsonx", "watsonx.data"),
    ("Software", "Automation", "Turbonomic", "Turbonomic Application Resource Mgmt"),
]


def ibm_non_infra_install(accounts, covids):
    rows = []
    for a in _subset(accounts, "noninfra", 0.35):
        r = _rng("noninfrarow", a["name"])
        for _ in range(r.randint(1, 4)):
            l10, l15, l17, l20 = r.choice(_NONINFRA_PRODUCTS)
            od = date(2024, 1, 1) + timedelta(days=r.randint(0, 700))
            rows.append([
                a["account_name"], a["account_name"], a["account_name"].upper(),
                l10, l15, l17, l20, l20, r.randint(1, 500), od.isoformat(),
                (od + timedelta(days=r.randint(5, 40))).isoformat(),
                (od + timedelta(days=365 * r.randint(1, 4))).isoformat(),
                f"SO{r.randint(1000000, 9999999)}", a["customer_number"], "United States",
            ])
    return NON_INFRA_HEADERS, rows


COMPETITIVE_HEADERS = ["L1_ACCOUNT_NAME", "L2_ACCOUNT_NAME", "CUST_NAME",
                       "ISSUING_COUNTRY_NAME", "CATEGORY_PARENT", "CATEGORY", "VENDOR",
                       "PRODUCT", "ATTRIBUTES", "DATE_FIRST_VERIFIED", "DATE_LAST_VERIFIED"]
_COMPETITORS = [
    ("Cloud", "IaaS", "Amazon Web Services", "EC2"),
    ("Cloud", "IaaS", "Microsoft Azure", "Azure VMs"),
    ("Database", "RDBMS", "Oracle", "Oracle Database"),
    ("Storage", "All-Flash", "Dell EMC", "PowerStore"),
    ("Analytics", "Data Warehouse", "Snowflake", "Snowflake"),
    ("Compute", "Servers", "Hewlett Packard Enterprise", "ProLiant"),
]


def competitive_install(accounts, covids):
    rows = []
    for a in _subset(accounts, "competitive", 0.22):
        r = _rng("comprow", a["name"])
        for _ in range(r.randint(1, 3)):
            parent, cat, vendor, product = r.choice(_COMPETITORS)
            fv = date(2023, 1, 1) + timedelta(days=r.randint(0, 900))
            rows.append([
                a["account_name"], a["account_name"], a["account_name"].upper(),
                "United States", parent, cat, vendor, product, f"{r.randint(1, 40)} units",
                fv.isoformat(), (fv + timedelta(days=r.randint(30, 300))).isoformat(),
            ])
    return COMPETITIVE_HEADERS, rows


# ── ZoomInfo enrichment mock (Account_Tiering gap-fill) ───────────────────────

def zoominfo_enrichment(account_name):
    """Stand-in for a ZoomInfo company-profile lookup. Deterministic per name."""
    r = _rng("zi", account_name)
    x = r.random()
    if x < 0.07:
        return {"ZI_Match_Status": "Unmatched", "ZI_Match_Method": "name", "ZI_Domain": "",
                "ZI_Revenue_USD": None, "ZI_Employee_Count": None,
                "ZI_Lookup_Timestamp": datetime.now().isoformat()}
    if x < 0.14:
        return {"ZI_Match_Status": "Ambiguous", "ZI_Match_Method": "name", "ZI_Domain": "",
                "ZI_Revenue_USD": None, "ZI_Employee_Count": None,
                "ZI_Lookup_Timestamp": datetime.now().isoformat()}
    employees = _employees(r)
    return {
        "ZI_Match_Status": "Matched", "ZI_Match_Method": "name",
        "ZI_Domain": _domain(account_name),
        "ZI_Revenue_USD": _revenue(r, employees),
        "ZI_Employee_Count": employees,
        "ZI_Lookup_Timestamp": datetime.now().isoformat(),
    }


# ── Buying-signal mock (Account_Tiering) ──────────────────────────────────────
_SIGNAL_TEMPLATES = [
    ("M&A", "{name} to acquire regional competitor in ${n}M deal"),
    ("Funding", "{name} secures ${n}M growth investment"),
    ("Leadership_Change", "{name} names new Chief Information Officer"),
    ("Expansion", "{name} opens new operations center, adding {n} roles"),
    ("Earnings_Financial", "{name} reports Q{q} revenue up {n}% year over year"),
    ("Partnership", "{name} announces strategic partnership on cloud modernization"),
    ("Product_Launch", "{name} unveils new digital customer platform"),
    ("Layoffs_Restructuring", "{name} restructuring to cut {n}00 jobs"),
    ("Security_Incident", "{name} discloses data security incident under review"),
    ("Regulatory_Compliance", "{name} reaches ${n}M regulatory settlement"),
    ("ESG_Commitment", "{name} commits to carbon-neutral operations by 2030"),
]


def signals_for(account_name):
    """0-3 recent buying signals per account, deterministic. Shape matches the real
    scraper: {Type, Date, Summary, Source_URL}."""
    r = _rng("signals", account_name)
    k = _weighted_choice(r, [(0, 34), (1, 30), (2, 22), (3, 14)])
    if not k:
        return []
    picks = r.sample(_SIGNAL_TEMPLATES, k)
    core = account_name.split(" Holdings")[0].split(" Group")[0].strip()
    out = []
    for sig_type, tmpl in picks:
        d = date.today() - timedelta(days=r.randint(3, 175))
        summary = tmpl.format(name=core, n=r.randint(2, 850), q=r.randint(1, 4))
        slug = "".join(ch for ch in core.lower() if ch.isalnum())[:20]
        out.append({
            "Type": sig_type,
            "Date": d.isoformat(),
            "Summary": summary[:300],
            "Source_URL": f"https://news.example.com/{slug}/{d.isoformat()}",
        })
    out.sort(key=lambda s: s["Date"], reverse=True)
    return out


# ── ZoomInfo Contact Readiness mock (Step 6) ──────────────────────────────────
_TITLES = [
    "VP of Infrastructure", "Director of IT", "Chief Information Officer",
    "Head of Cloud Platform", "VP Engineering", "Director of Data & Analytics",
    "Enterprise Architect", "IT Operations Manager", "Chief Technology Officer",
    "Director of Information Security", "VP Digital Transformation",
    "Senior Infrastructure Engineer",
]
_FIRST = ["James", "Maria", "David", "Priya", "Michael", "Sarah", "Chen", "Aisha",
          "Robert", "Emily", "Carlos", "Fatima", "Daniel", "Grace", "Omar", "Laura"]
_LAST = ["Anderson", "Patel", "Nguyen", "Okafor", "Rivera", "Kim", "Johnson", "Silva",
         "Cohen", "Murphy", "Zhang", "Ali", "Brooks", "Ferrari", "Novak", "Reed"]


def contacts_for_accounts(account_names, per_account=(2, 5)):
    """Buyer-group contacts a ZoomInfo 'Infra Outbound' filter would surface. Returns
    dicts with a 'raw' text blob (name / title / company), mirroring the real step
    which captures row innerText, not parsed emails (those cost ZoomInfo credits)."""
    out = []
    for name in account_names:
        r = _rng("contacts", name)
        for _ in range(r.randint(*per_account)):
            fn, ln, title = r.choice(_FIRST), r.choice(_LAST), r.choice(_TITLES)
            out.append({
                "first_name": fn, "last_name": ln, "title": title, "company": name,
                "raw": f"{fn} {ln}  {title}  {name}",
            })
    return out


# ── Salesloft mock (Fill Contacts advance + Bobby) ────────────────────────────
SALESLOFT_CADENCES = [
    "Targeted Outreach Cadence 3", "Targeted Outreach Cadence 4",
    "Enterprise Expansion Cadence", "Whitespace Nurture Cadence",
]
# Step layout per cadence: (day, type, name). Real Salesloft cadences interleave
# email / phone / other steps; Bobby drafts for the email steps.
_CADENCE_STEP_LAYOUT = [
    (1, "email", "Intro email"),
    (1, "phone", "Call — first attempt"),
    (3, "email", "Value follow-up"),
    (5, "phone", "Call — second attempt"),
    (5, "other", "LinkedIn touch"),
    (8, "email", "Case-study share"),
    (12, "email", "Break-up email"),
]


def salesloft_cadence_steps(cadence):
    """List of step dicts: {id, step_number, day, type, name}. Deterministic."""
    r = _rng("cadence", cadence)
    steps = []
    for i, (day, stype, sname) in enumerate(_CADENCE_STEP_LAYOUT, start=1):
        steps.append({"id": r.randint(100000, 999999) + i, "step_number": i,
                      "day": day, "type": stype, "name": sname})
    return steps


def salesloft_people_for_cadence(cadence, seed_accounts=None):
    """People enrolled on a cadence, spread across its email steps. Companies come
    from the fake account pool when available, else a generic set. Deterministic."""
    r = _rng("slpeople", cadence)
    companies = list(seed_accounts) if seed_accounts else [
        "Meridian Health", "Cascade Financial Group", "Summit Insurance",
        "Vanguard Capital", "Keystone Manufacturing", "Horizon Networks",
    ]
    r.shuffle(companies)
    companies = companies[:r.randint(6, 12)]
    email_steps = [s for s in salesloft_cadence_steps(cadence) if s["type"] == "email"]
    people = []
    pid = 500000
    for step in email_steps:
        for company in companies:
            if r.random() < 0.55:
                continue
            pid += 1
            fn, ln, title = r.choice(_FIRST), r.choice(_LAST), r.choice(_TITLES)
            people.append({
                "id": pid,
                "membership_id": pid + 90000,
                "first_name": fn, "last_name": ln, "title": title,
                "company": company, "step_id": step["id"], "step_day": step["day"],
                "email": f"{fn.lower()}.{ln.lower()}@{_domain(company)}",
            })
    # Guarantee at least a couple of people per email step for a lively demo.
    for step in email_steps:
        if not any(p["step_id"] == step["id"] for p in people):
            for company in companies[:2]:
                pid += 1
                fn, ln, title = r.choice(_FIRST), r.choice(_LAST), r.choice(_TITLES)
                people.append({
                    "id": pid, "membership_id": pid + 90000,
                    "first_name": fn, "last_name": ln, "title": title,
                    "company": company, "step_id": step["id"], "step_day": step["day"],
                    "email": f"{fn.lower()}.{ln.lower()}@{_domain(company)}",
                })
    return people
