"""Resolve the set of CovIDs selected in the ISC Scraper step (Step 1).

Precedence:
  1. Explicit list passed in (CLI --covids), highest priority.
  2. ISC_Scraper_App/output/selected_covids.json  -- the authoritative selection
     the ISC launcher writes on each /run (includes CovIDs that scraped 0 rows,
     which the DEDUPED output would otherwise drop).
  3. Distinct "Coverage ID" values parsed out of DEDUPED_ACCOUNTS (latest.xlsx) --
     a best-effort fallback so the step still works if (2) is absent.

CovIDs are T-prefixed codes, e.g. "T0016329". Returns an order-preserving
de-duplicated list.
"""
import json
import re

import config

_COVID_RE = re.compile(r"T\d{4,}", re.IGNORECASE)


def _dedup(seq):
    seen, out = set(), []
    for v in seq:
        v = str(v).strip()
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return out


def _from_selected_json():
    path = config.SELECTED_COVIDS_PATH
    if not path.exists():
        return None
    data = json.loads(path.read_text())
    # Accept either a bare list or {"covids": [...]}.
    ids = data.get("covids") if isinstance(data, dict) else data
    return _dedup(ids) if ids else None


def _from_deduped_accounts():
    path = config.DEDUPED_ACCOUNTS_PATH
    if not path.exists():
        return None
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # DEDUPED_ACCOUNTS has multiple sheets (e.g. "Company Rollup" +
    # "Companies by Industry"); the Coverage ID column lives on the rollup sheet,
    # which is not guaranteed to be the active one. Scan every sheet and use the
    # first that actually has a "Coverage ID" column rather than trusting
    # wb.active (which would silently return nothing if the active sheet changed).
    for ws in wb.worksheets:
        try:
            header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        except StopIteration:
            continue  # empty sheet
        try:
            ci = [str(h).strip() if h is not None else h for h in header].index("Coverage ID")
        except ValueError:
            continue  # no Coverage ID on this sheet; try the next
        found = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            v = row[ci] if ci < len(row) else None
            if v:
                found.extend(_COVID_RE.findall(str(v)))
        deduped = _dedup(found)
        if deduped:
            return deduped
    return None


def resolve_covids(explicit=None):
    """Return (covids, source_label). Raises if none can be resolved."""
    if explicit:
        return _dedup(explicit), "explicit"
    ids = _from_selected_json()
    if ids:
        return ids, f"selected_covids.json ({config.SELECTED_COVIDS_PATH})"
    ids = _from_deduped_accounts()
    if ids:
        return ids, f"DEDUPED_ACCOUNTS Coverage ID column ({config.DEDUPED_ACCOUNTS_PATH})"
    raise RuntimeError(
        "No CovIDs found. Run the ISC Scraper step first (so selected_covids.json / "
        "DEDUPED_ACCOUNTS exist), or pass --covids T0000000,T0000001,..."
    )
