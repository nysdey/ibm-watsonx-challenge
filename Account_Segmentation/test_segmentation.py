"""Session-free tests for Account Segmentation. Run:

    ../.venv/bin/python3 test_segmentation.py
"""
import sys
import tempfile
from pathlib import Path

import openpyxl

import json

import config
import id_match
import name_match
import segment

PASS = 0
FAIL = 0


def check(name, cond):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  PASS  {name}")
    else:
        FAIL += 1
        print(f"  FAIL  {name}")


def _xlsx(path, header, rows, sheet="Sheet1"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(path)


# ---------------- name_match units ----------------
def test_name_match():
    check("normalize possessive/case", name_match.normalize("John's Hospital") == "JOHNS HOSPITAL")
    check("normalize legal suffix", name_match.normalize("Acme Corporation, Inc.") == "ACME")
    check("normalize The/&", name_match.normalize("The Globex & Co") == "GLOBEX")
    check("similarity identical", name_match.similarity("ACME", "ACME") == 1.0)
    check("similarity word-order", name_match.similarity("JOHNS HOSPITAL", "HOSPITAL JOHNS") > 0.9)
    check("classify matched", name_match.classify(0.95) == "matched")
    check("classify review", name_match.classify(0.85) == "review")
    check("classify none", name_match.classify(0.5) == "none")

    idx = name_match.NameIndex()
    idx.add("JOHNS HOSPITAL", 0)
    idx.add("Johns Hospital Inc", 1)   # normalizes to same key -> both rows
    idx.add("Cedars Sinai Medical Ctr", 2)
    nm, score, q, rows = idx.match("John's Hospital")
    check("index exact after normalize (2 rows)", q == "matched" and score == 1.0 and set(rows) == {0, 1})
    nm, score, q, rows = idx.match("Cedars-Sinai Medical Center")
    check("index fuzzy review/match", q in ("matched", "review") and rows == [2])
    nm, score, q, rows = idx.match("Totally Different LLC")
    check("index no match", q == "none" and rows == [])


# ---------------- full segmentation run ----------------
def _run_with(tmp, attach=True, overrides=None, base_name_col="Account Name"):
    base = tmp / "base.xlsx"
    _xlsx(base, [base_name_col, "Industry"], [
        ["A5 Co", "x"], ["A4 Co", "x"], ["A3 Co", "x"],
        ["CloudPower Co", "x"], ["CloudStorage Co", "x"],
        ["CloudOnly Co", "x"], ["PowerOnly Co", "x"], ["Zero Co", "x"],
    ])
    # who appears in each type
    membership = {
        "cloud":         ["A5 Co", "A4 Co", "A3 Co", "CloudPower Co", "CloudStorage Co", "CloudOnly Co"],
        "power":         ["A5 Co", "A4 Co", "A3 Co", "CloudPower Co", "PowerOnly Co"],
        "storage":       ["A5 Co", "A4 Co", "A3 Co", "CloudStorage Co"],
        "ibm_non_infra": ["A5 Co", "A4 Co"],
        "competitive":   ["A5 Co"],
    }
    # varied name-column headers to exercise auto-detect
    name_cols = {"cloud": "Customer Name", "power": "Primary Customer Name",
                 "storage": "Client Name", "ibm_non_infra": "Parent Account",
                 "competitive": "Account"}
    paths = {}
    for key, members in membership.items():
        p = tmp / f"{key}.xlsx"
        _xlsx(p, [name_cols[key], "Val"], [[m, 1] for m in members])
        paths[key] = p

    config.DEDUPED_ACCOUNTS_PATH = base
    config.BASE_NAME_COLUMN = base_name_col
    config.OUTPUT_DIR = tmp / "out"
    config.ATTACH_INSTALL_COLUMNS = attach
    config.NAME_COLUMN_OVERRIDES = overrides or {}
    config.INSTALL_TYPES = [
        ("cloud", "Cloud", paths["cloud"]),
        ("power", "Power", paths["power"]),
        ("storage", "Storage", paths["storage"]),
        ("ibm_non_infra", "NonInfra", paths["ibm_non_infra"]),
        ("competitive", "Competitive", paths["competitive"]),
    ]
    segment.run_segmentation()
    wb = openpyxl.load_workbook(config.OUTPUT_DIR / "latest.xlsx")
    ws = wb.active
    H = [c.value for c in ws[1]]
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    return H, rows


def test_full(tmp):
    H, rows = _run_with(tmp, attach=True)
    ix = {h: i for i, h in enumerate(H)}
    order = [r[ix["Account Name"]] for r in rows]
    expected = ["A5 Co", "A4 Co", "A3 Co", "CloudPower Co", "CloudStorage Co",
                "CloudOnly Co", "PowerOnly Co", "Zero Co"]
    check("sort: count desc + priority tie-break", order == expected)
    if order != expected:
        print("     got:", order)

    a5 = rows[0]
    check("A5 count==5", a5[ix["Install_Types_Count"]] == 5)
    check("A5 types listed in priority order",
          a5[ix["Install_Types"]] == "Cloud, Power, Storage, NonInfra, Competitive")
    check("Zero count==0", rows[-1][ix["Install_Types_Count"]] == 0)

    # auto-detected name columns worked (presence set via each file's own header)
    cp = next(r for r in rows if r[ix["Account Name"]] == "CloudPower Co")
    check("auto-detect: Cloud present", cp[ix["Cloud_Present"]] == "Yes")
    check("auto-detect: Power present", cp[ix["Power_Present"]] == "Yes")
    check("auto-detect: Storage absent", not cp[ix["Storage_Present"]])  # blank cell reads back as None

    # attach columns present
    check("attach: has prefixed install column", any(h.startswith("Cloud: ") for h in H))


def test_attach_toggle(tmp):
    H, rows = _run_with(tmp, attach=False)
    check("attach off: no prefixed columns", not any(": " in (h or "") and h.split(":")[0] in
          {"Cloud", "Power", "Storage", "NonInfra", "Competitive"} for h in H))
    check("attach off: still has compact presence cols", "Cloud_Present" in H)


def test_override_and_missing(tmp):
    # name-column override honored
    H, rows = _run_with(tmp, overrides={"cloud": "Customer Name"})
    ix = {h: i for i, h in enumerate(H)}
    a5 = rows[0]
    check("override: Cloud still matched", a5[ix["Cloud_Present"]] == "Yes")

    # base missing its name column -> raises
    bad = tmp / "badbase.xlsx"
    _xlsx(bad, ["Wrong Col", "Industry"], [["X", "y"]])
    config.DEDUPED_ACCOUNTS_PATH = bad
    config.BASE_NAME_COLUMN = "Account Name"
    try:
        segment.run_segmentation()
        check("missing base name col raises", False)
    except RuntimeError:
        check("missing base name col raises", True)


def test_numeric_vs_text_aggregation(tmp):
    base = tmp / "b2.xlsx"
    _xlsx(base, ["Account Name"], [["Multi Co"]])
    inst = tmp / "power2.xlsx"
    _xlsx(inst, ["Primary Customer Name", "Spend", "Product"],
          [["Multi Co", 100, "P8"], ["Multi Co", 200, "P8"], ["Multi Co", 50, "P9"]])
    config.DEDUPED_ACCOUNTS_PATH = base
    config.BASE_NAME_COLUMN = "Account Name"
    config.OUTPUT_DIR = tmp / "out2"
    config.ATTACH_INSTALL_COLUMNS = True
    config.NAME_COLUMN_OVERRIDES = {}
    config.INSTALL_TYPES = [("power", "Power", inst)]
    segment.run_segmentation()
    wb = openpyxl.load_workbook(config.OUTPUT_DIR / "latest.xlsx")
    ws = wb.active
    H = [c.value for c in ws[1]]
    r = [c.value for c in list(ws.iter_rows(min_row=2))[0]]
    ix = {h: i for i, h in enumerate(H)}
    check("aggregate: numeric summed (350)", r[ix["Power: Spend"]] == 350)
    check("aggregate: text distinct-joined", set(str(r[ix["Power: Product"]]).split("; ")) == {"P8", "P9"})
    check("aggregate: Power_Rows==3", r[ix["Power_Rows"]] == 3)


# ---------------- id_match units ----------------
def test_id_match_units():
    check("codes_in plain", id_match.codes_in("DC46JLHF") == ["DC46JLHF"])
    check("codes_in compound key (Cloud)", id_match.codes_in("GC2TT1CS-T0016156-897") == ["GC2TT1CS"])
    check("codes_in excludes UNASSIGN sentinel", id_match.codes_in("UNASSIGN") == [])
    check("codes_in covid not a code", id_match.codes_in("T0016156") == [])
    check("norm_cust leading zeros", id_match.norm_cust("0139800") == "139800")
    check("norm_cust country suffix", id_match.norm_cust("5788184-897") == "5788184")
    check("norm_cust too short", id_match.norm_cust("12") == "")

    base_h = ["Account Name", "Account Number"]
    base_d = [["Acme Corp", "DC000001"], ["Globex LLC", "DB000002"], ["Initech", "DC000003"]]
    xw = {"cust_to_key": {"5788184": "DC000003"}}
    res = id_match.build_resolver(base_h, base_d, 1, 0, xw)
    check("resolver: base keys", res.base_keys == {"DC000001", "DB000002", "DC000003"})
    # code match wins
    k, b = res.resolve(["ZZ999999", "DB000002"], [], [])
    check("resolve by code", k == "DB000002" and b == "code")
    # customer-number bridge
    k, b = res.resolve([], ["5788184-897"], [])
    check("resolve by customer number", k == "DC000003" and b == "customer_number")
    # exact same-system name
    k, b = res.resolve([], [], ["ACME CORPORATION, INC."])  # normalizes to ACME
    check("resolve by exact name", k == "DC000001" and b == "name_exact")
    # nothing matches -> genuinely out of set (no false positive)
    k, b = res.resolve(["GC123456"], ["9999999"], ["Totally Unknown Co"])
    check("resolve none (no false positive)", k is None)


def test_id_join_full(tmp):
    """End-to-end deterministic join: names DIFFER across files, but ids match;
    and a same-token-but-different account is NOT falsely joined."""
    base = tmp / "base.xlsx"
    _xlsx(base, ["Account Name", "Account Number", "Industry"], [
        ["Chevron Federal Credit Union", "DB500XCC", "x"],   # power by code
        ["Redwood Credit Union", "DC46JLHF", "x"],           # power by code, different NAME spelling
        ["Mid-Peninsula Partners", "DC777001", "x"],         # cloud by customer number
        ["Kuakini Health System", "DC888002", "x"],          # non-infra by exact name
        ["No Installs Co", "DC999003", "x"],
    ], sheet="Company Rollup")
    # Power file: rows keyed by hierarchy code; note CBC is a DIFFERENT CU that
    # fuzzy would wrongly match to "Chevron ... Credit Union".
    power = tmp / "power.xlsx"
    _xlsx(power, ["Primary Customer Name", "domestic client id", "domestic buying group id", "Val"], [
        ["REDWOOD C U", "DC46JLHF", "DB500LNP", 1],          # matches Redwood by code (name differs!)
        ["CBC FEDERAL CREDIT UNION", "DC6YG9M7", "UNASSIGN", 1],  # NOT Chevron — must not match
        ["CHEVRON FCU", "DC5555ZZ", "DB500XCC", 1],          # matches Chevron by buying-group code
    ])
    cloud = tmp / "cloud.xlsx"
    _xlsx(cloud, ["Account Name", "Unique Account Key", "IBM Customer Number with Country Code", "Val"], [
        ["MID-PENINSULA PARTNERS, LLC", "GC2TT1CS-T0016156-897", "5788184-897", 1],
    ])
    noninfra = tmp / "noninfra.xlsx"
    _xlsx(noninfra, ["L1_ACCOUNT_NAME", "CUST_NO", "Val"], [
        ["KUAKINI HEALTH SYSTEM", "4990713", 1],             # exact same-system name
    ])
    xw = tmp / "account_crosswalk.json"
    xw.write_text(json.dumps({"keys": ["DC777001"], "cust_to_key": {"5788184": "DC777001"}}))

    config.DEDUPED_ACCOUNTS_PATH = base
    config.BASE_NAME_COLUMN = "Account Name"
    config.ACCOUNT_CROSSWALK_PATH = xw
    config.OUTPUT_DIR = tmp / "out"
    config.ATTACH_INSTALL_COLUMNS = True
    config.NAME_COLUMN_OVERRIDES = {}
    config.INSTALL_TYPES = [
        ("cloud", "Cloud", cloud),
        ("power", "Power", power),
        ("ibm_non_infra", "NonInfra", noninfra),
    ]
    segment.run_segmentation()
    wb = openpyxl.load_workbook(config.OUTPUT_DIR / "latest.xlsx")
    ws = wb.active
    H = [c.value for c in ws[1]]; ix = {h: i for i, h in enumerate(H)}
    rows = {r[ix["Account Name"]]: r for r in ([c.value for c in rr] for rr in ws.iter_rows(min_row=2))}

    redwood = rows["Redwood Credit Union"]
    check("id-join: Power matched despite different name", redwood[ix["Power_Present"]] == "Yes")
    check("id-join: basis is code", redwood[ix["Power_Match_Basis"]] == "code")

    chevron = rows["Chevron Federal Credit Union"]
    check("id-join: Chevron matched by its own code", chevron[ix["Power_Present"]] == "Yes")
    check("id-join: Chevron NOT falsely joined to CBC row (1 row)", chevron[ix["Power_Rows"]] == 1)
    check("id-join: Chevron matched name is its own", "CHEVRON" in str(chevron[ix["Power_Matched_Name"]]).upper())

    mpp = rows["Mid-Peninsula Partners"]
    check("id-join: Cloud matched by customer number", mpp[ix["Cloud_Present"]] == "Yes"
          and mpp[ix["Cloud_Match_Basis"]] == "customer_number")

    kua = rows["Kuakini Health System"]
    check("id-join: NonInfra matched by exact name", kua[ix["NonInfra_Present"]] == "Yes"
          and kua[ix["NonInfra_Match_Basis"]] == "name_exact")

    nic = rows["No Installs Co"]
    check("id-join: account with no installs stays 0", nic[ix["Install_Types_Count"]] == 0)


def main():
    print("== name_match units =="); test_name_match()
    print("== id_match units =="); test_id_match_units()
    with tempfile.TemporaryDirectory() as d:
        print("== id-join full (deterministic) =="); test_id_join_full(Path(d))
    with tempfile.TemporaryDirectory() as d:
        tmp = Path(d)
        print("== full run (5 files) =="); test_full(tmp)
    with tempfile.TemporaryDirectory() as d:
        print("== attach toggle =="); test_attach_toggle(Path(d))
    with tempfile.TemporaryDirectory() as d:
        print("== override + missing base col =="); test_override_and_missing(Path(d))
    with tempfile.TemporaryDirectory() as d:
        print("== numeric vs text aggregation =="); test_numeric_vs_text_aggregation(Path(d))
    print(f"\nSegmentation tests: {PASS} passed, {FAIL} failed")
    sys.exit(1 if FAIL else 0)


if __name__ == "__main__":
    main()
