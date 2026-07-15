"""Seller Dashboard pipeline dashboard — a real web page, same pattern as
ISC_Scraper_App/launcher.py (Flask, auto-opens a browser tab).

    python3 run_pipeline.py

Opens http://127.0.0.1:5488 showing the status of all 5 steps (present even if
a step hasn't run yet), auto-skips Steps 1/2 if they're already done (redo
buttons offered regardless), and lets you trigger Steps 3-5 with auto/manual
modes, watching each step's live log output right on the page.

Converted from a CLI dashboard to a web page 2026-07-05 per explicit user
request. Steps 4/5 have no interactive confirmation gate (removed the same
day, explicit user decision — see their own run.py module docstrings) — this
dashboard's buttons trigger them directly. Manual-mode inputs (account
selection, cadence name) are collected by this page's own form and passed as
CLI args, never via an interactive terminal prompt (a subprocess launched from
Flask has no one to answer input() calls).
"""
import json
import os
import secrets
import shutil
import socket
import subprocess
import sys
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime
from pathlib import Path

from flask import Flask, jsonify, render_template_string, request

import credential_store
import fake_data
import mock_salesloft
import seller_accounts
import shared_auth
from shared_auth import guard
from ui_templates import (
    PAGE_TEMPLATE, VIEW_TEMPLATE, TIER_VIEW_TEMPLATE, CALENDAR_TEMPLATE,
    STRATEGY_TEMPLATE, BOBBY_PAGE_TEMPLATE,
)
from mock_ui_templates import (
    MOCK_LOGIN_TEMPLATE, MOCK_CONNECTED_TEMPLATE, MOCK_SALESLOFT_TEMPLATE,
    MOCK_ZOOMINFO_TEMPLATE, MOCK_ISC_TEMPLATE,
)

REPO_ROOT = Path(__file__).resolve().parent
VENV_PYTHON = REPO_ROOT / ".venv" / "bin" / "python3"
STATE_PATH = REPO_ROOT / ".orum_pipeline_state.json"
LOGIN_CONTROL_DIR = REPO_ROOT / ".orum_login_control"

# WatsonX Clone: the Meetings tab (live_transcribe_bot) and the Pipeline tab
# (Pipeline_Review) are both removed. All external connectivity is mocked — see
# fake_data.py / mock_salesloft.py and the "MOCK" notes throughout.

# Services surfaced + managed in the Details/Access panel. In the clone these are
# all mocked: the "sessions" never touch a real site, and Details ▸ Access "Log in"
# buttons open the in-app mock login pages (/mock/<service>/login). Outlook is gone
# with the Meetings tab.
_PANEL_LOGIN_SERVICES = ["isc", "zoominfo", "salesloft"]
LOGIN_SERVICES = {svc: shared_auth.state_path(svc) for svc in _PANEL_LOGIN_SERVICES}

# The Outbound "Start Full Pipeline" login gate covers only the pipeline services.
_PIPELINE_LOGIN_SERVICES = ["isc", "zoominfo", "salesloft"]

# Every dashboard-triggered run (Start Full Pipeline or a lone step button)
# is meant to behave like a fresh demo, never resuming from a prior run's
# checkpoints — logins are the one piece of state that's meant to persist,
# so auth_state.json is deliberately untouched here. Steps 1/3 have no
# checkpoints of their own.
_STEPS_WITH_CHECKPOINTS = ["step2", "step4", "step5"]

# These two checkpoints cache slow EXTERNAL lookups (ZoomInfo company enrichment
# and web buying-signals) — minutes to rebuild for a full account pool, and they
# are stable facts about companies, not run-specific state. Reusing them is what
# makes a re-run of Outbound Strategy fast (cold = many minutes, warm = seconds),
# so they are deliberately PRESERVED across dashboard-triggered runs and restarts
# while everything else still resets to a clean slate. Results are unaffected:
# tiers are recomputed every run, and a newly-added account (not yet in the file)
# is still freshly looked up. (2026-07-10, production-speed pass.)
_DURABLE_CHECKPOINTS = {"zoominfo_checkpoint.json", "signals_checkpoint.json"}

STEPS = {
    "step1": {"name": "Step 1 — ISC Scraper", "dir": "ISC_Scraper_App",
              "output": "output/latest.xlsx", "sheet": "Company Rollup"},
    # New Step 2/3 inserted before Account Tiering (2026-07-09). IBM Scraper
    # deliberately has output=None — per the spec the dashboard shows its
    # PROCESS (streamed log + last-run marker) but never the install files
    # themselves; it writes a logs/run_*.log so status reads "ran".
    "ibm": {"name": "Step 2 — IBM Scraper", "dir": "IBM_Scraper_App",
            "output": None, "sheet": None},
    "segment": {"name": "Step 3 — Account Segmentation", "dir": "Account_Segmentation",
                "output": "output/latest.xlsx", "sheet": "Segmented Accounts"},
    "step2": {"name": "Step 4 — Account Tiering", "dir": "Account_Tiering",
              "output": "output/latest.xlsx", "sheet": "Tiered Accounts"},
    "step3": {"name": "Step 5 — Call Planning", "dir": "Call_Planning",
              "output": "output/latest.xlsx", "sheet": "Call Plan"},
    "step4": {"name": "Step 6 — ZoomInfo Contact Readiness", "dir": "ZoomInfo_Contact_Readiness",
              "output": None, "sheet": None},
    "step5": {"name": "Step 7 — Salesloft Cadence Readiness", "dir": "Salesloft_Cadence_Readiness",
              "output": None, "sheet": None},
}

# step_key -> {"proc": Popen, "lines": [...], "lock": Lock(), "started_at": iso, "returncode": None|int}
_PROCESSES = {}
_ISC_PROC = None  # separate: ISC_Scraper_App/launcher.py is its own long-lived Flask app
_ISC_PORT = 5477

_LOGIN_PROCS = {}  # service -> Popen (login_capture.py, while a login is in progress)

# Cached real-session-validity per service, filled by _login_validator_loop().
# A saved auth_state.json file merely existing does NOT mean the session is
# still alive — Salesforce/ZoomInfo/Salesloft sessions expire server-side with
# no cookie-expiry signal, so the only trustworthy check is actually navigating
# to the app headless and seeing whether it bounces to a login page (see
# login_capture.py's probe mode). status is one of: valid/expired/missing/
# error/checking. `checked_at` is an epoch float; None until first probe.
_LOGIN_VALIDITY = {}   # service -> {"status": str, "checked_at": float|None}
_LOGIN_VALIDITY_LOCK = threading.Lock()
_LOGIN_REVALIDATE_NOW = set()   # services queued for an immediate re-probe
_LOGIN_PROBE_INTERVAL = 150     # seconds between routine re-probes per service
# Services we've already auto-launched a login for since they last went
# invalid — so a still-expired service isn't relaunched every loop. Cleared
# for a service once it's confirmed valid again (so a later expiry re-triggers).
_AUTO_LOGIN_ATTEMPTED = set()
# Self-healing circuit breaker (docs/SECURITY.md I15/I34): count auto-login
# attempts that still didn't reach 'valid', and stop auto-retrying a service after
# this many, so a stale saved W3ID password isn't resubmitted across six services
# into an account lockout / MFA-fatigue. Reset once the service goes valid again.
_AUTO_LOGIN_FAILS = {}          # service -> consecutive failed auto-heals
_MAX_AUTO_LOGIN_FAILS = 2
_WATCHDOG_LAST_TICK = [None]    # epoch of the auth-watchdog loop's last pass

# Ambient step-to-step auto-chain is now DISABLED (2026-07-10 Seller Dashboard
# rework): the three combined action buttons — "Get My Accounts" (ISC → IBM →
# Segmentation), "Outbound Strategy" (Tiering → Call Planning), and "Fill
# Contacts to SalesLoft" (ZoomInfo → Salesloft) — each drive their own sequence
# explicitly via the orchestrators below, so a finished step must NOT also kick
# off the next one on its own (that would double-run). _maybe_auto_chain stays
# defined (still called by _stream_reader) but with an empty map it only manages
# the vestigial _PIPELINE_STATE and launches nothing.
_AUTO_CHAIN_NEXT = {}
_PIPELINE_STATE = {"active": False, "current": None, "error": None}
_step1_autochain_seen_run_id = None


def _login_status(service):
    """Real per-service login state:
      'missing'  — no saved session file (never logged in on this machine)
      'waiting'  — login_capture.py is up, visible browser open, not saved yet
      'checking' — session file exists but hasn't been validity-probed yet
      'ready'    — probe confirmed the saved session actually works
      'expired'  — probe found the saved session bounces to a login page
      'error'    — probe couldn't reach the app (network/browser problem)
    'ready' now means *verified live*, not just 'a file exists on disk' — that
    old check reported ready forever even after the session died server-side,
    which is exactly the 'says logged in when it isn't' bug this replaces."""
    # WatsonX Clone: sessions are mocked and never expire — always 'ready'. The
    # Details ▸ Access "Log in" button opens the in-app mock login page (for show).
    return {"state": "ready", "updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "mock": True}


def _probe_login_once(service):
    """WatsonX Clone: sessions are mocked — no real probe. Always valid."""
    return "valid"


def _svc_label(service):
    return (shared_auth.get(service) or {}).get("label", service)


def _ensure_services_ready(services):
    """WatsonX Clone: every service is mocked, so there is no real session to probe —
    always ready. (In the real app this JIT-probed each saved session before a step ran.)"""
    return True, None


def _login_validator_loop():
    """Background thread: keeps _LOGIN_VALIDITY fresh by re-probing all three
    services in parallel, plus immediately when a login proc exits. Skips a
    service while its interactive login is in progress."""
    _proc_was_running = set()
    while True:
        try:
            _, scraping = _isc_progress()
        except Exception:
            scraping = False
        # Don't launch competing headless browsers on a session file that a live
        # step is using, and don't pop login windows mid-run (I4b/I8): treat any
        # active automation — not just an ISC scrape — as "busy".
        busy = scraping or _automation_active()
        now = time.time()
        _WATCHDOG_LAST_TICK[0] = now

        # Detect any login proc that just finished → force immediate re-probe.
        for service in LOGIN_SERVICES:
            proc = _LOGIN_PROCS.get(service)
            if proc and proc.poll() is None:
                _proc_was_running.add(service)
            elif service in _proc_was_running:
                _proc_was_running.discard(service)
                _LOGIN_REVALIDATE_NOW.add(service)

        # Decide which services need a probe this tick.
        to_probe = []
        for service in LOGIN_SERVICES:
            proc = _LOGIN_PROCS.get(service)
            if proc and proc.poll() is None:
                continue  # mid-login, skip
            forced = service in _LOGIN_REVALIDATE_NOW
            if busy and not forced:
                continue
            # A capture (login_capture) holds the exclusive session lock — never
            # probe it out from under an in-progress sign-in (I4b).
            if not forced and guard.is_locked_exclusive(service):
                continue
            if not LOGIN_SERVICES[service].exists():
                with _LOGIN_VALIDITY_LOCK:
                    _LOGIN_VALIDITY[service] = {"status": "missing", "checked_at": now}
                _LOGIN_REVALIDATE_NOW.discard(service)
                continue
            with _LOGIN_VALIDITY_LOCK:
                v = _LOGIN_VALIDITY.get(service)
            due = (not v) or v.get("checked_at") is None or (now - v["checked_at"]) >= _LOGIN_PROBE_INTERVAL
            if due or forced:
                to_probe.append(service)

        if to_probe:
            # Probe all due services in parallel — total time = slowest one.
            probe_lock = threading.Lock()
            def _do_probe(svc):
                status = _probe_login_once(svc)
                with probe_lock:
                    with _LOGIN_VALIDITY_LOCK:
                        _LOGIN_VALIDITY[svc] = {"status": status, "checked_at": time.time()}
                    _LOGIN_REVALIDATE_NOW.discard(svc)
            threads = [threading.Thread(target=_do_probe, args=(s,), daemon=True) for s in to_probe]
            for t in threads:
                t.start()
            for t in threads:
                t.join(timeout=35)

        _maybe_auto_login(busy)
        time.sleep(5)


def _automation_active():
    """True if any dashboard-driven automation is mid-run (Fill Contacts, Get My
    Accounts, Outbound Strategy, Bobby, Pipeline Review). The auth watchdog uses
    this to avoid probing/relogging a session out from under a live step (I4b/I8).
    The state dicts are module globals defined later; this runs only at loop time."""
    for name in ("_FILL_STATE", "_GMA_STATE", "_STRATEGY_STATE", "_BOBBY_STATE"):
        st = globals().get(name)
        if isinstance(st, dict) and st.get("active"):
            return True
    return False


def _maybe_auto_login(busy):
    """Self-healing re-login: when a service is logged-out and a W3ID credential
    is saved, open its login window (auto-fill drives the SSO; a human still
    approves MFA). Guardrails:
      * never while a scrape/automation is running (I8) — no surprise windows,
        no competing browser on a session file mid-run;
      * a circuit breaker (I15/I34) stops auto-retrying a service after
        _MAX_AUTO_LOGIN_FAILS consecutive heals that didn't reach 'valid', so a
        stale saved password isn't resubmitted into a lockout / MFA-fatigue.
    Re-armed (fail count cleared) once the service is confirmed valid again."""
    if busy:
        return
    if not credential_store.has("w3id"):
        return
    for service in LOGIN_SERVICES:
        with _LOGIN_VALIDITY_LOCK:
            v = _LOGIN_VALIDITY.get(service)
        status = v["status"] if v else None
        if status == "valid":
            _AUTO_LOGIN_ATTEMPTED.discard(service)
            _AUTO_LOGIN_FAILS.pop(service, None)   # healed — re-arm
            continue
        # Already has a login window open for this service — don't re-open.
        proc = _LOGIN_PROCS.get(service)
        if proc and proc.poll() is None:
            continue
        # A prior auto-heal for this service finished but it's STILL not valid →
        # count the failure and allow another try, up to the breaker cap.
        if service in _AUTO_LOGIN_ATTEMPTED and status in ("expired", "missing", "error"):
            _AUTO_LOGIN_FAILS[service] = _AUTO_LOGIN_FAILS.get(service, 0) + 1
            _AUTO_LOGIN_ATTEMPTED.discard(service)
        if _AUTO_LOGIN_FAILS.get(service, 0) >= _MAX_AUTO_LOGIN_FAILS:
            continue  # circuit-broken — stop auto-retrying; surface via /api/auth/health
        if status in ("expired", "missing") and service not in _AUTO_LOGIN_ATTEMPTED:
            _AUTO_LOGIN_ATTEMPTED.add(service)
            guard.audit(service, "auto_login", "launch")
            _start_login_proc(service)


def _fmt_mtime(path):
    try:
        return datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return None


def _ibm_manifest():
    """The IBM Scraper's per-run summary (output/run_manifest.json — see
    IBM_Scraper_App/run.py's _report). Says, per sub-scraper, whether its
    install file is FRESH from this run vs STALE/MISSING, plus row counts and
    the CovID selection. This is the 'progress/results of IBM' the dashboard
    surfaces — the process summary, never the raw install files. None if the
    step hasn't produced a manifest yet."""
    path = REPO_ROOT / "IBM_Scraper_App" / "output" / "run_manifest.json"
    try:
        data = json.loads(path.read_text())
    except Exception:
        return None
    outputs = data.get("outputs", {})
    fresh = sum(1 for o in outputs.values() if o.get("freshness") == "FRESH")
    return {
        "run_at": data.get("run_at"),
        "covids": data.get("covids", []),
        "outputs": outputs,
        "fresh_count": fresh,
        "total_count": len(outputs),
        "all_fresh": bool(outputs) and fresh == len(outputs),
    }


def _isc_progress():
    """Poll the ISC launcher's own /progress endpoint (see
    ISC_Scraper_App/launcher.py's _PROGRESS) for a one-line 'which CovIDs
    right now' summary — deliberately not the full scrape log. Returns
    (text, scraping) where `scraping` is True only while a scrape is
    genuinely in flight (not yet merged) — the launcher app itself stays
    running long after any given scrape finishes, so that alone isn't a
    useful 'is Step 1 running' signal. Uses `merged`, not done>=total,
    since done can reach total after the first pass while retry sweeps for
    empty-CovID cache misses (see launcher.py) are still re-running some of
    them before the final output is actually written."""
    # WatsonX Clone: there is no separate ISC launcher/endpoint. Step 1 status is
    # derived from the output file in _step_status, so there's never a live scrape
    # to report here.
    return None, False


def _load_state():
    if STATE_PATH.exists():
        return json.loads(STATE_PATH.read_text())
    return {}


def _save_state(state):
    STATE_PATH.write_text(json.dumps(state, indent=2))


# Row counts are shown on every step card and recomputed on every 2s /api/status
# poll. Opening a multi-MB, 600-column workbook and iterating all cells each time
# made a single poll take ~3s (the Segmentation output alone is ~4 MB) — the UI
# felt frozen. Cache by (path, mtime): recount only when the file actually
# changes, and read just the first column (we only need the row *count*, not the
# values), which is orders of magnitude cheaper than pulling every cell.
_ROWCOUNT_CACHE = {}  # str(path) -> (mtime, sheet, count)


def _row_count(path, sheet):
    try:
        key = str(path)
        mtime = path.stat().st_mtime
        cached = _ROWCOUNT_CACHE.get(key)
        if cached and cached[0] == mtime and cached[1] == sheet:
            return cached[2]
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active
        count = sum(1 for _ in ws.iter_rows(min_row=2, max_col=1, values_only=True))
        wb.close()
        _ROWCOUNT_CACHE[key] = (mtime, sheet, count)
        return count
    except Exception:
        return None


def _step_status(key):
    step = STEPS[key]
    running = key in _PROCESSES and _PROCESSES[key]["proc"].poll() is None
    if key == "step1":
        progress, scraping = _isc_progress()
        running = scraping
    info = {"key": key, "name": step["name"], "running": running}

    if key == "step1" and progress:
        info["progress"] = progress

    if step["output"]:
        path = REPO_ROOT / step["dir"] / step["output"]
        if path.exists():
            info["state"] = "running" if running else "done"
            info["path"] = str(path.relative_to(REPO_ROOT))
            info["updated"] = datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
            info["rows"] = _row_count(path, step["sheet"])
        else:
            info["state"] = "running" if running else "pending"
    else:
        log_dir = REPO_ROOT / step["dir"] / "logs"
        logs = sorted(log_dir.glob("run_*.log")) if log_dir.exists() else []
        if running:
            info["state"] = "running"
        elif logs:
            info["state"] = "ran"
            info["updated"] = datetime.fromtimestamp(logs[-1].stat().st_mtime).strftime("%Y-%m-%d %H:%M")
            info["run_count"] = len(logs)
        else:
            info["state"] = "pending"

    if key == "ibm":
        manifest = _ibm_manifest()
        if manifest:
            info["manifest"] = manifest

    if key in _PROCESSES:
        with _PROCESSES[key]["lock"]:
            info["log_tail"] = _PROCESSES[key]["lines"][-200:]
        info["returncode"] = _PROCESSES[key]["proc"].poll()
    return info


def _stream_reader(step_key, proc):
    for raw_line in iter(proc.stdout.readline, b""):
        line = raw_line.decode(errors="replace").rstrip()
        with _PROCESSES[step_key]["lock"]:
            _PROCESSES[step_key]["lines"].append(line)
    proc.stdout.close()
    rc = proc.wait()
    _maybe_auto_chain(step_key, rc)


def _maybe_auto_chain(step_key, returncode):
    """Ambient pipeline behavior (2026-07-05): whenever a step finishes
    successfully, automatically kick off the next one — except Step 3, which
    deliberately does NOT auto-advance to Step 4 (the one step that stays a
    manual click, since that's where the real ZoomInfo->Salesloft write
    happens). Applies no matter how the finished step was triggered (a lone
    "Redo Step 2" click auto-advances to Step 3 too, not just full-pipeline
    runs) — this is a standing rule, not a special mode."""
    nxt = _AUTO_CHAIN_NEXT.get(step_key)
    if not nxt or returncode != 0:
        if _PIPELINE_STATE["active"]:
            _PIPELINE_STATE["current"] = "waiting_step4" if step_key == "step3" else "done"
            # "waiting_step4" is the only non-terminal stop — the chain is
            # still logically in progress (just paused for a manual click).
            # Every other outcome here (step5 finished, or any error) is
            # terminal: clear 'active' so a later "Start Full Pipeline"
            # click isn't permanently refused as "already running".
            if step_key != "step3" or returncode != 0:
                _PIPELINE_STATE["active"] = False
            if returncode != 0:
                _PIPELINE_STATE["error"] = f"{step_key} exited with code {returncode}"
        return
    if nxt == "segment":
        _launch("segment", [])
    elif nxt == "step2":
        _launch("step2", [])
    elif nxt == "step3":
        _launch("step3", [])
    elif nxt == "step5":
        cadence = _load_state().get("default_salesloft_cadence")
        if not cadence:
            if _PIPELINE_STATE["active"]:
                _PIPELINE_STATE["error"] = "step5: no saved default cadence yet — run Step 5 manually once to set one."
                _PIPELINE_STATE["active"] = False
            return
        _launch("step5", ["--cadence", cadence])
    if _PIPELINE_STATE["active"]:
        _PIPELINE_STATE["current"] = nxt


def _clear_checkpoints(step_key):
    """Wipe a step's checkpoint files so it always behaves like a first run
    (2026-07-06, explicit user request — every dashboard-triggered run should
    be a fresh demo of the whole workflow, never silently skipping
    accounts/lists/contacts because a prior run already checkpointed them).
    Login state (auth_state.json) lives elsewhere and is untouched."""
    if step_key not in _STEPS_WITH_CHECKPOINTS:
        return
    ckpt_dir = REPO_ROOT / STEPS[step_key]["dir"] / "checkpoints"
    if not ckpt_dir.exists():
        return
    for f in ckpt_dir.glob("*.json"):
        if f.name in _DURABLE_CHECKPOINTS:
            continue  # keep the expensive external-lookup cache — see above
        f.unlink()


def _launch(step_key, args):
    if step_key in _PROCESSES and _PROCESSES[step_key]["proc"].poll() is None:
        return False, "already running"
    _clear_checkpoints(step_key)
    step = STEPS[step_key]
    cmd = [str(VENV_PYTHON), "run.py"] + args
    # stdin=DEVNULL is critical: several steps have interactive input() prompts
    # (ZoomInfo lookup-pause, ZoomInfo/Salesloft login gates). Launched from the
    # dashboard there's no one to answer them — with an inherited terminal stdin
    # they'd block the step FOREVER (this was the "Outbound Strategy does nothing"
    # bug). DEVNULL turns any input() into an immediate EOFError, which each
    # step's unattended guard / try-except turns into fail-soft-and-continue.
    proc = subprocess.Popen(
        cmd, cwd=str(REPO_ROOT / step["dir"]),
        stdin=subprocess.DEVNULL,
        stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
    )
    _PROCESSES[step_key] = {"proc": proc, "lines": [], "lock": threading.Lock(),
                             "started_at": datetime.now().isoformat()}
    threading.Thread(target=_stream_reader, args=(step_key, proc), daemon=True).start()
    return True, None


def _step1_watcher():
    """WatsonX Clone: there is no separate long-lived ISC launcher to watch — Get My
    Accounts drives the mock ISC generation in-process and chains the next steps
    itself. Kept as a no-op so main()'s (now removed) thread start would be harmless."""
    return


app = Flask(__name__)

# ── Local-auth guard (docs/SECURITY.md I46/I47/I48) ──────────────────────
# The dashboard binds loopback, but loopback is not a security boundary: any web
# page the seller visits can fetch()/POST to 127.0.0.1, and a DNS-rebinding page
# can point its own hostname at 127.0.0.1. This guard closes both WITHOUT any UI
# change: browsers set Origin/Host themselves and can't forge them cross-origin,
# and every same-origin UI POST carries a loopback Origin (so the UI just works).
# The token is a shared secret for a trusted *local* tool that wants to drive the
# API headlessly — set DASHBOARD_AUTH_TOKEN in the environment to use it; there is
# deliberately no endpoint that hands it out (that would defeat the point).
_DASHBOARD_AUTH_TOKEN = os.environ.get("DASHBOARD_AUTH_TOKEN") or secrets.token_urlsafe(32)
# Strict mode additionally blocks tokenless, origin-less mutating requests (a local
# curl/script). Off by default so local tooling + the test client keep working; the
# cross-site browser drive-by + rebinding are blocked either way. The UI keeps
# working in strict mode too (its POSTs carry a same-origin loopback Origin).
_STRICT_LOCAL_AUTH = os.environ.get("DASHBOARD_STRICT_AUTH") == "1"
_LOOPBACK_HOSTS = {"127.0.0.1", "localhost", "::1"}


def _host_is_loopback(host_header):
    h = (host_header or "").rsplit(":", 1)[0].strip().strip("[]").lower()
    return h in _LOOPBACK_HOSTS


def _url_host_is_loopback(url):
    try:
        return (urllib.parse.urlparse(url).hostname or "").lower() in _LOOPBACK_HOSTS
    except Exception:
        return False


@app.before_request
def _dashboard_auth_guard():
    # 1. Anti DNS-rebinding: the Host header must be loopback.
    if not _host_is_loopback(request.host):
        return jsonify({"error": "forbidden host"}), 403
    # 2. A valid per-launch token always passes (stops another local process).
    if request.headers.get("X-Auth-Token") == _DASHBOARD_AUTH_TOKEN:
        return None
    # 3. CSRF: a browser always sends Origin on a cross-origin POST, so a
    #    non-loopback Origin/Referer on a state-changing request is a drive-by.
    if request.method in ("POST", "PUT", "DELETE", "PATCH"):
        for hdr in ("Origin", "Referer"):
            val = request.headers.get(hdr)
            if val:
                if _url_host_is_loopback(val):
                    return None
                return jsonify({"error": "cross-origin request blocked"}), 403
        # Neither Origin nor Referer present → not a browser cross-site request.
        if _STRICT_LOCAL_AUTH:
            return jsonify({"error": "auth token required"}), 403
    return None


@app.route("/")
def index():
    return render_template_string(PAGE_TEMPLATE, today=date.today().isoformat(),
                                  cadences=FILL_CADENCES, bobby_cadences=BOBBY_CADENCES)


@app.route("/api/status")
def api_status():
    out = {key: _step_status(key) for key in STEPS}
    out["_pipeline"] = dict(_PIPELINE_STATE)
    out["_actions"] = {
        "get_my_accounts": dict(_GMA_STATE),
        "outbound_strategy": dict(_STRATEGY_STATE),
        "fill_contacts": dict(_FILL_STATE),
        "bobby": dict(_BOBBY_STATE),
    }
    return jsonify(out)


@app.route("/api/step2/accounts")
def api_step2_accounts():
    """For Step 4's manual mode — the full tiered account list to render as a
    picker in the page, instead of opening a native spreadsheet app."""
    path = REPO_ROOT / "Account_Tiering" / "output" / "latest.xlsx"
    if not path.exists():
        return jsonify({"error": f"{path} doesn't exist yet — run Step 2 first."}), 404
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Tiered Accounts"]
    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))
    idx = {h: i for i, h in enumerate(header)}
    out = []
    for r in rows_iter:
        if not r or not r[idx["Account Name"]]:
            continue
        out.append({
            "name": r[idx["Account Name"]],
            "tier": r[idx["Tier"]] if "Tier" in idx else None,
            "score": r[idx["Tier_Score"]] if "Tier_Score" in idx else None,
            "industry": r[idx["Industry"]] if "Industry" in idx else None,
        })
    wb.close()
    return jsonify(out)


_ISC_TOKEN = None  # per-launch shared secret for the launcher's /run (I53/I54)
_ISC_HANDSHAKE = Path("~/.isc_scraper/launcher_port.json").expanduser()


def _ensure_isc_launcher_running():
    """WatsonX Clone: the separate ISC scraper app is gone — Get My Accounts
    generates the fake ISC output in-process (see _isc_scrape). No-op."""
    return None


@app.route("/api/step1/launch", methods=["POST"])
def api_step1_launch():
    """WatsonX Clone: there is no separate ISC app to open — the ISC scrape is mocked
    and driven headlessly by Get My Accounts."""
    return jsonify({"ok": False, "port": None,
                    "message": "The ISC scraper is mocked in this clone — use “Get My Accounts.”"})


@app.route("/api/ibm/run", methods=["POST"])
def api_ibm_run():
    # Runs all five IBM sub-scrapers (run.py, full ORDER). Power (local) always
    # works; the four browser sub-scrapers each fail-soft and continue if their
    # portal session isn't logged in / calibrated yet. run.py writes
    # output/run_manifest.json flagging which install files are fresh vs stale.
    ok, err = _launch("ibm", [])
    return jsonify({"ok": ok, "error": err})


@app.route("/api/segment/run", methods=["POST"])
def api_segment_run():
    ok, err = _launch("segment", [])
    return jsonify({"ok": ok, "error": err})


@app.route("/api/step2/run", methods=["POST"])
def api_step2_run():
    ok, err = _launch("step2", [])
    return jsonify({"ok": ok, "error": err})


@app.route("/api/step3/run", methods=["POST"])
def api_step3_run():
    ok, err = _launch("step3", [])
    return jsonify({"ok": ok, "error": err})


def _load_sheet_dicts(path, sheet):
    """(header, [row-dict, ...]) for a workbook sheet — dict rows keyed by header."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))
    rows = [dict(zip(header, r)) for r in rows_iter if r and any(c is not None for c in r)]
    wb.close()
    return header, rows


def _first(row, *keys):
    for k in keys:
        v = row.get(k)
        if v not in (None, ""):
            return v
    return None


def _fmt_money(v):
    try:
        v = float(v)
    except (TypeError, ValueError):
        return "—"
    a = abs(v)
    if a >= 1e9:
        return f"${v/1e9:.1f}B"
    if a >= 1e6:
        return f"${v/1e6:.1f}M"
    if a >= 1e3:
        return f"${v/1e3:.0f}K"
    return f"${v:,.0f}"


def _fmt_int(v):
    try:
        return f"{int(round(float(v))):,}"
    except (TypeError, ValueError):
        return "—"


# Only columns a phone seller acts on — the 600+ raw install/code columns from
# Segmentation are deliberately hidden here (kept in the workbook for downstream
# joins, never shown). label -> how to pull it off a tiered row.
def _seller_tier_rows(rows):
    ordered = sorted(rows, key=lambda r: (r.get("Tier", 9) or 9, -(r.get("Tier_Score") or 0)))
    out = []
    for i, r in enumerate(ordered, 1):
        out.append({
            "rank": i,
            "tier": r.get("Tier"),
            "account": r.get("Account Name") or "",
            "industry": r.get("Industry") or "",
            "play": r.get("Primary_Play") or "",
            "score": r.get("Tier_Score"),
            "relationship": r.get("Technology Client Status") or "—",
            "trend": r.get("Spend_Trend") or "Unknown",
            "spend": _fmt_money(r.get("IBM Spend Current Year")),
            "revenue": _fmt_money(_first(r, "ZI_Revenue_USD", "Location Annual Revenue")),
            "employees": _fmt_int(_first(r, "ZI_Employee_Count", "Employee Count")),
            "install": r.get("Install_Summary") or "No IBM installs",
            "competitor": r.get("Competitive_Displacement") or "No",
            "contacts": _fmt_int(r.get("Contact Count")) if r.get("Contact Count") not in (None, "") else "0",
            "angle": r.get("Sales_Angle") or "",
        })
    counts = {t: sum(1 for r in rows if r.get("Tier") == t) for t in (1, 2, 3)}
    return out, counts


def _calendar_view(path):
    """Month-grid calendar (today → END_OF_YEAR) with each working day's account
    count, color-coded by top tier; clicking a day lists its accounts."""
    import calendar as _calmod
    _, rows = _load_sheet_dicts(path, "Call Plan")

    accounts_by_date = {}
    for r in rows:
        iso = r.get("Planned_Call_Date")
        if not iso:
            continue
        accounts_by_date.setdefault(iso, []).append({
            "name": r.get("Account Name") or "",
            "tier": r.get("Planned_Tier") or r.get("Tier") or 3,
            "score": r.get("Tier_Score"),
            "play": r.get("Primary_Play") or "",
            "industry": r.get("Industry") or "",
            "install": r.get("Install_Summary") or "",
            "angle": r.get("Sales_Angle") or "",
            "seq": r.get("Day_Sequence_Number") or 0,
        })
    for lst in accounts_by_date.values():
        lst.sort(key=lambda a: (a["tier"], a["seq"]))

    meta = {}
    meta_path = path.parent / "plan_meta.json"
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text())
        except Exception:
            meta = {}

    today = date.today()
    end = date(2026, 12, 31)
    try:
        end = datetime.strptime(meta.get("window_end", ""), "%Y-%m-%d").date()
    except Exception:
        pass
    start = today

    # Holiday set (grey out) — reuse Call_Planning's rule-based holidays.
    holidays = set()
    try:
        cp_dir = str(REPO_ROOT / "Call_Planning")
        if cp_dir not in sys.path:
            sys.path.insert(0, cp_dir)
        import us_holidays
        for yr in range(start.year, end.year + 1):
            holidays |= us_holidays.federal_holidays(yr)
    except Exception:
        pass

    cal = _calmod.Calendar(firstweekday=6)  # Sunday-first, US convention
    months, y, m = [], start.year, start.month
    while (y, m) <= (end.year, end.month):
        weeks = []
        for week in cal.monthdatescalendar(y, m):
            cells = []
            for d in week:
                if d.month != m:
                    cells.append({"blank": True})
                    continue
                accts = accounts_by_date.get(d.isoformat(), [])
                cells.append({
                    "blank": False, "day": d.day, "iso": d.isoformat(),
                    "in_range": start <= d <= end,
                    "weekend": d.weekday() >= 5,
                    "holiday": d in holidays,
                    "count": len(accts),
                    "tier": min((a["tier"] for a in accts), default=0),
                    "today": d == today,
                })
            weeks.append(cells)
        months.append({"label": date(y, m, 1).strftime("%B %Y"), "weeks": weeks})
        m, y = (1, y + 1) if m == 12 else (m + 1, y)

    return render_template_string(
        CALENDAR_TEMPLATE, months=months,
        accounts_json=json.dumps(accounts_by_date),
        meta=meta, total=len(rows),
        path=str(path.relative_to(REPO_ROOT)),
    )


@app.route("/view/<step_key>")
def view_results(step_key):
    """Renders a step's output in its own tab — a real data view, not the raw
    .xlsx. Tiering (step2) and Call Planning (step3) get purpose-built, seller-
    facing views (color-coded tier table / clickable calendar); other steps get
    the generic sortable table."""
    if step_key not in STEPS or not STEPS[step_key]["output"]:
        return f"No viewable output for '{step_key}'.", 404
    step = STEPS[step_key]
    path = REPO_ROOT / step["dir"] / step["output"]
    if not path.exists():
        return f"{path} doesn't exist yet — run {step['name']} first.", 404

    if step_key == "step2":
        _, rows = _load_sheet_dicts(path, step["sheet"])
        seller_rows, counts = _seller_tier_rows(rows)
        return render_template_string(
            TIER_VIEW_TEMPLATE, rows=seller_rows, counts=counts,
            total=len(seller_rows), path=str(path.relative_to(REPO_ROOT)),
        )
    if step_key == "step3":
        return _calendar_view(path)

    if step_key == "segment":
        header, rows = _segment_view_rows(path, step["sheet"])
        return render_template_string(
            VIEW_TEMPLATE, title="Get My Accounts — Your Accounts", header=header, rows=rows,
            row_count=len(rows), path=str(path.relative_to(REPO_ROOT)),
        )

    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[step["sheet"]] if step["sheet"] in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))
    rows = [r for r in rows_iter if r and any(c is not None for c in r)]
    wb.close()

    return render_template_string(
        VIEW_TEMPLATE, title=step["name"], header=header, rows=rows,
        row_count=len(rows), path=str(path.relative_to(REPO_ROOT)),
    )


# The Segmentation output carries 600+ raw IBM install/code columns that are
# kept in the workbook for downstream joins but are useless (and ~14 MB of HTML,
# multi-second render) in a seller's results view. Show only the columns a seller
# reads, and read just the leading columns of each row (max_col) instead of
# parsing all 600+ — turns a ~4s / 14 MB page into a fast, compact one.
_SEGMENT_VIEW_COLUMNS = [
    "Account Name", "Coverage ID", "Industry", "Technology Client Status",
    "IBM Spend Current Year", "IBM Spend Prior Year", "Location Annual Revenue",
    "Employee Count", "Contact Count", "Install_Types_Count", "Install_Types",
]
_SEGMENT_VIEW_CACHE = {}  # str(path) -> (mtime, (cols, rows))


def _segment_view_rows(path, sheet):
    """(curated_header, rows) for the Segmentation results view — only the
    seller-relevant columns. Parsing the 4 MB / 600-column workbook takes ~1.3s,
    so cache by mtime: the file only changes when Get My Accounts re-runs, and a
    seller may open the results tab repeatedly."""
    key = str(path)
    mtime = path.stat().st_mtime
    cached = _SEGMENT_VIEW_CACHE.get(key)
    if cached and cached[0] == mtime:
        return cached[1]
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    full_header = [str(h).strip() if h is not None else h for h in next(rows_iter)]
    idx = {h: i for i, h in enumerate(full_header)}
    cols = [c for c in _SEGMENT_VIEW_COLUMNS if c in idx]
    picks = [idx[c] for c in cols]
    name_i = idx.get("Account Name", 0)
    out = []
    for r in rows_iter:
        if not r or (name_i < len(r) and not r[name_i]):
            continue
        out.append(tuple(r[i] if i < len(r) else None for i in picks))
    wb.close()
    _SEGMENT_VIEW_CACHE[key] = (mtime, (cols, out))
    return cols, out


@app.route("/api/step4/run", methods=["POST"])
def api_step4_run():
    body = request.json or {}
    mode = body.get("mode", "auto")
    target_date = body.get("date") or date.today().isoformat()
    args = ["--date", target_date, "--mode", mode]
    if mode == "manual":
        accounts = body.get("accounts", "")
        if not accounts:
            return jsonify({"ok": False, "error": "manual mode needs at least one account (or 'all')."})
        args += ["--accounts", accounts]
    ok, err = _launch("step4", args)
    return jsonify({"ok": ok, "error": err})


@app.route("/api/step5/run", methods=["POST"])
def api_step5_run():
    body = request.json or {}
    mode = body.get("mode", "auto")
    state = _load_state()
    if mode == "manual":
        cadence = body.get("cadence", "").strip()
        if not cadence:
            return jsonify({"ok": False, "error": "manual mode needs a cadence name."})
    else:
        cadence = body.get("cadence", "").strip() or state.get("default_salesloft_cadence")
        if not cadence:
            return jsonify({"ok": False, "error": "no saved default cadence yet — provide one to set it."})
        if not state.get("default_salesloft_cadence"):
            state["default_salesloft_cadence"] = cadence
            _save_state(state)
    args = ["--cadence", cadence]
    ok, err = _launch("step5", args)
    return jsonify({"ok": ok, "error": err, "cadence": cadence})


@app.route("/api/login/status")
def api_login_status():
    return jsonify({service: _login_status(service) for service in LOGIN_SERVICES})


@app.route("/api/auth/health")
def api_auth_health():
    """One-glance auth health across every managed service + the self-healing
    watchdog's state. Lets the UI (and a human) see at any moment whether all
    sessions are live, whether auto-heal is armed or circuit-broken for a service,
    and when the watchdog last ran. Backs the 'constantly ensure auth is good'
    guarantee — if a service isn't ready, the watchdog is already re-healing it
    (unless its breaker tripped, which shows here)."""
    services = {}
    for svc in LOGIN_SERVICES:
        s = _login_status(svc)
        fails = _AUTO_LOGIN_FAILS.get(svc, 0)
        services[svc] = {
            "state": s.get("state"),
            "checked_at": s.get("checked_at"),
            "heal_fails": fails,
            "auto_heal_armed": fails < _MAX_AUTO_LOGIN_FAILS,
            "login_in_progress": bool(_LOGIN_PROCS.get(svc) and _LOGIN_PROCS[svc].poll() is None),
        }
    broken = [s for s, v in services.items() if not v["auto_heal_armed"]]
    return jsonify({
        "all_ready": all(v["state"] == "ready" for v in services.values()),
        "credential_saved": credential_store.has("w3id"),
        "automation_active": _automation_active(),
        "watchdog_running": _WATCHDOG_LAST_TICK[0] is not None,
        "watchdog_age_s": (round(time.time() - _WATCHDOG_LAST_TICK[0], 1)
                           if _WATCHDOG_LAST_TICK[0] else None),
        "probe_interval_s": _LOGIN_PROBE_INTERVAL,
        "needs_manual_attention": broken,   # breaker tripped → likely stale password
        "services": services,
    })


def _start_login_proc(service):
    """WatsonX Clone: no real browser login — the Details ▸ Access "Log in" button
    opens the in-app mock login page (/mock/<service>/login). Returns (ok, message)."""
    return True, "mock"


@app.route("/api/login/<service>/start", methods=["POST"])
def api_login_start(service):
    if service not in LOGIN_SERVICES:
        return jsonify({"ok": False, "error": f"unknown service '{service}'"}), 404
    # Point the UI at the in-app mock sign-in page for this tool instead of spawning a
    # real browser. (The session is already 'ready' in the clone; this is for show.)
    return jsonify({"ok": True, "mock_url": f"/mock/{service}/login",
                    "message": f"Opening the mock {_svc_label(service)} sign-in…"})


@app.route("/api/login/<service>/confirm", methods=["POST"])
def api_login_confirm(service):
    if service not in LOGIN_SERVICES:
        return jsonify({"ok": False, "error": f"unknown service '{service}'"}), 404
    return jsonify({"ok": True})


# Which credential keys the dashboard exposes a Save-Passwords form for.
# One key — 'w3id' — covers all three services: ISC, ZoomInfo, and Salesloft
# all authenticate through IBM W3ID SSO (Salesloft redirects to it on email
# entry, no separate password).
_CREDENTIAL_KEYS = ["w3id"]


@app.route("/api/credentials/status")
def api_credentials_status():
    """Which credential keys have a stored password — booleans only, never the
    values themselves (the dashboard shows 'saved / not saved', nothing more)."""
    return jsonify({key: credential_store.has(key) for key in _CREDENTIAL_KEYS})


@app.route("/api/credentials/<key>", methods=["POST"])
def api_credentials_save(key):
    """Store an email+password into the OS Keychain. The plaintext is used
    here only to hand to credential_store.save() and is never logged or echoed
    back — the response is a bare ok/error."""
    if key not in _CREDENTIAL_KEYS:
        return jsonify({"ok": False, "error": f"unknown credential key '{key}'"}), 404
    body = request.json or {}
    email = (body.get("email") or "").strip()
    password = body.get("password") or ""
    ok, err = credential_store.save(key, email, password)
    return jsonify({"ok": ok, "error": err})


# ── Mock tool UIs (login windows + a viewable Salesloft / ZoomInfo / ISC) ──────
# Wherever the real app would open one of these tools, the clone opens a self-
# contained mock instead. The Details ▸ Access "Log in" buttons open /mock/<svc>/login;
# Fill Contacts / Bobby link to /mock/salesloft to show the loaded contacts.
_MOCK_BRAND = {
    "isc": {"label": "ISC", "brand": "#0176d3"},
    "zoominfo": {"label": "ZoomInfo", "brand": "#e5352b"},
    "salesloft": {"label": "Salesloft", "brand": "#6b4ce6"},
}


def _demo_accounts():
    """Regenerate the current demo's accounts from the CovIDs the ISC step selected
    (deterministic), so the mock ZoomInfo/ISC pages show the same companies."""
    p = REPO_ROOT / "ISC_Scraper_App" / "output" / "selected_covids.json"
    if not p.exists():
        return []
    try:
        covids = json.loads(p.read_text()).get("covids", [])
    except Exception:
        return []
    return fake_data.accounts_for_covids(covids)


@app.route("/mock/<service>/login")
def mock_login(service):
    b = _MOCK_BRAND.get(service)
    if not b:
        return "Unknown service", 404
    return render_template_string(MOCK_LOGIN_TEMPLATE, service=service,
                                  email=_signed_in_email() or "", **b)


@app.route("/mock/<service>/signin", methods=["POST"])
def mock_signin(service):
    b = _MOCK_BRAND.get(service)
    if not b:
        return "Unknown service", 404
    return render_template_string(MOCK_CONNECTED_TEMPLATE, service=service, **b)


@app.route("/mock/salesloft")
def mock_salesloft_view():
    selected = request.args.get("cadence") or fake_data.SALESLOFT_CADENCES[0]
    if selected not in fake_data.SALESLOFT_CADENCES:
        selected = fake_data.SALESLOFT_CADENCES[0]
    state = mock_salesloft.all_state().get("cadences", {})
    cadences = [{"name": n, "count": len(state.get(n, {}).get("members", []))}
                for n in fake_data.SALESLOFT_CADENCES]
    members = state.get(selected, {}).get("members", [])
    steps = fake_data.salesloft_cadence_steps(selected)
    at_step1 = sum(1 for m in members if m.get("step") == mock_salesloft.FIRST_STEP)
    at_call = sum(1 for m in members if m.get("step") == mock_salesloft.CALL_STEP)
    email_steps = sum(1 for s in steps if s["type"] == "email")
    return render_template_string(
        MOCK_SALESLOFT_TEMPLATE, cadences=cadences, selected=selected, members=members,
        steps=steps, total=len(members), at_step1=at_step1, at_call=at_call,
        email_steps=email_steps)


@app.route("/mock/zoominfo")
def mock_zoominfo_view():
    rows = [{"name": a["account_name"], "industry": a["industry"],
             "domain": fake_data._domain(a["account_name"]),
             "revenue": _fmt_money(a["global_revenue"]),
             "employees": _fmt_int(a["employees"]) if a["employees"] else "—"}
            for a in _demo_accounts()]
    return render_template_string(MOCK_ZOOMINFO_TEMPLATE, rows=rows)


@app.route("/mock/isc")
def mock_isc_view():
    rows = [{"name": a["account_name"], "coverage": ";".join(a["coverage_ids"]),
             "industry": a["industry"], "status": a["tech_client_status"],
             "contacts": a["contact_count"]} for a in _demo_accounts()]
    return render_template_string(MOCK_ISC_TEMPLATE, rows=rows)


def _start_pipeline():
    """Kicks off the ambient chain: launches Step 1's app and opens its
    map-picker tab for you to select territories and click Run yourself —
    does NOT auto-scrape every territory (changed 2026-07-06, explicit user
    request: Start Full Pipeline was silently forcing a POST /run
    {"all": true} every time, ignoring whatever was actually selected on the
    map — only ever scrape what's actually chosen there). Once you finish
    that run in the opened tab, the standing _maybe_auto_chain/_step1_watcher
    rules below detect its completion the same way they would for a manual
    "Redo Step 1" run, and carry it on through Steps 2-3. Stops there — Step 4
    is the one deliberate manual step (2026-07-05, explicit user request)
    since that's where the real ZoomInfo->Salesloft write happens. Once Step
    4 finishes (however it's triggered), the same ambient rule carries it on
    to Step 5."""
    _PIPELINE_STATE.update({"active": True, "current": "waiting_step1_manual", "error": None})
    try:
        _ensure_isc_launcher_running()
        for _ in range(30):
            try:
                urllib.request.urlopen(f"http://127.0.0.1:{_ISC_PORT}/", timeout=2)
                break
            except Exception:
                time.sleep(1)
        else:
            _PIPELINE_STATE.update({"active": False, "error": "ISC_Scraper_App didn't come up in time."})
            return
        # No webbrowser.open() here — ISC_Scraper_App/launcher.py's own
        # main() already opens its own tab automatically on startup (see its
        # threading.Timer(0.9, ...) call). Adding a second open() here (as an
        # earlier version of this fix did) opened two tabs for every fresh
        # launch. If the app was already running from before, no new tab is
        # needed — switch to the existing one.
    except Exception as e:
        _PIPELINE_STATE.update({"active": False, "error": str(e)})


@app.route("/api/pipeline/run_all", methods=["POST"])
def api_pipeline_run_all():
    if _PIPELINE_STATE["active"]:
        return jsonify({"ok": False, "error": "already running"})
    not_ready = [s for s in _PIPELINE_LOGIN_SERVICES if _login_status(s)["state"] != "ready"]
    if not_ready:
        return jsonify({"ok": False, "error": f"log in first: {', '.join(not_ready)}"})
    threading.Thread(target=_start_pipeline, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/state")
def api_state():
    return jsonify(_load_state())


# ── Combined action orchestrators (Seller Dashboard, 2026-07-10) ──────────────
# Each of the three Outbound-page buttons runs a fixed sequence of the existing
# backend steps and reports its own natural-language progress (not the raw
# terminal log). The individual step programs are all unchanged — these just
# drive them in order and narrate what's happening.
#
#   Get My Accounts     -> ISC Scraper (CovIDs from Name Match) -> IBM -> Segment
#   Outbound Strategy   -> Account Tiering -> Call Planning
#   Fill Contacts       -> ZoomInfo Contact Readiness -> Salesloft (chosen cadence)
#
# phase is a machine tag; message is the human sentence the page shows.
_GMA_STATE = {"active": False, "phase": "idle", "message": "", "error": None,
              "seller": None, "covids": 0, "done": False}
_STRATEGY_STATE = {"active": False, "phase": "idle", "message": "", "error": None, "done": False}
_FILL_STATE = {"active": False, "phase": "idle", "message": "", "error": None,
               "cadence": None, "done": False}
# Bobby, the AI Emailer — its own async action (not one of the pipeline steps).
_BOBBY_STATE = {"active": False, "phase": "idle", "message": "", "error": None,
                "cadence": None, "done": False, "summary": None, "log": []}
_BOBBY_PROC = {"proc": None}   # the currently-running Bobby subprocess, if any


def _signed_in_email():
    """The email the seller signed in with — the single W3ID credential's
    account. None if they haven't signed in yet."""
    return credential_store.get_email("w3id")


def _launch_and_wait(step_key, args):
    """Start a step's run.py and block until it exits; returns (returncode,
    error). If the step is somehow already running, waits on that existing
    process rather than starting a second one."""
    ok, err = _launch(step_key, args)
    if not ok and err != "already running":
        return -1, err
    entry = _PROCESSES.get(step_key)
    if not entry:
        return -1, "process did not start"
    return entry["proc"].wait(), None


def _last_step_error(step_key, fallback):
    """The most useful error line from a finished step's captured output, so the
    dashboard can show the specific cause (e.g. 'your ZoomInfo session has expired')
    instead of a generic 'see the log'. Falls back to `fallback` if none found."""
    import re as _re
    entry = _PROCESSES.get(step_key)
    if not entry:
        return fallback
    with entry["lock"]:
        lines = list(entry["lines"])

    def _clean(l):
        l = _re.sub(r"^\d{4}-\d\d-\d\d[ T]\d\d:\d\d:\d\d[.,]\d+\s+", "", l)
        l = _re.sub(r"^(INFO|WARNING|WARN|ERROR|DEBUG)\s+", "", l)
        l = _re.sub(r"^\[[^\]]+\]\s+", "", l)
        return l.strip()

    for l in reversed(lines):
        low = l.lower()
        if ("error" in low or "expired" in low or "session" in low) and "traceback" not in low:
            c = _clean(l)
            if c:
                # Redact any URL before surfacing to the UI — SSO bounce URLs carry
                # SAMLResponse / auth codes in the query string (I65).
                c = _re.sub(r'https?://[^\s)\'"]+',
                            lambda m: guard.redact_url(m.group(0)), c)
                return c[:300]
    return fallback


def _isc_scrape(covids, on_progress, timeout=1200):
    """WatsonX Clone: the ISC (Salesforce) scrape is MOCKED. Instead of driving a
    headless Aura/Playwright scrape, generate the identical handoff artifacts
    (ISC_Scraper_App/output/latest.xlsx 'Company Rollup' + 'Companies by Industry',
    selected_covids.json, account_crosswalk.json) from the deterministic fake-data
    pool, narrating progress so the card feels live. Returns (ok, error)."""
    covids = list(dict.fromkeys(str(c).strip() for c in covids if str(c).strip()))
    isc_out = REPO_ROOT / "ISC_Scraper_App" / "output"
    total = len(covids) or 1
    pace = min(0.3, 4.0 / total)   # spread ~4s of "work" across the CovIDs, max 0.3s each
    for i, cov in enumerate(covids, start=1):
        on_progress(f"Pulling your territories from ISC — {i}/{total} covered (working on {cov})")
        time.sleep(pace)
    on_progress("Deduping locations and rolling up accounts by company…")
    try:
        accounts = fake_data.generate_isc_output(covids, isc_out)
    except Exception as e:
        return False, f"could not generate ISC accounts: {e}"
    on_progress(f"Pulled {len(accounts)} companies across {total} coverage ID(s).")
    return True, None


def _run_get_my_accounts():
    """ISC (Name-Match CovIDs) → IBM → Segmentation, narrated in plain language."""
    def set_state(**kw):
        _GMA_STATE.update(kw)

    set_state(active=True, phase="resolve", error=None, done=False,
              message="Looking up your territories…")
    try:
        email = _signed_in_email()
        if not email:
            raise RuntimeError("You're not signed in — sign in with your IBM email first.")
        seller = seller_accounts.resolve_seller(email)
        covids = seller.get("covids") or []
        seller_name = seller.get("seller_name")
        if not seller.get("matched") or not covids:
            # WatsonX Clone: any email works. When it isn't a real rep in Name
            # Match.xlsx, assign a stable demo territory so the flow never dead-ends.
            covids = fake_data.demo_covids_for_email(email)
            seller_name = seller_name or email.split("@")[0].replace(".", " ").title()
        set_state(seller=seller_name, covids=len(covids),
                  message=f"Found {len(covids)} coverage IDs for {seller_name}.")

        # 1) ISC Scraper — pull the accounts in those territories. JIT-verify the
        # ISC session is actually live (I1) rather than spinning up the engine only
        # to stall on an expired session — the seller logs in once via Details.
        ok, msg = _ensure_services_ready(["isc"])
        if not ok:
            raise RuntimeError(msg)
        set_state(phase="isc", message=f"Pulling your {len(covids)} territories from ISC…")
        ok, err = _isc_scrape(covids, lambda t: set_state(message=t))
        if not ok:
            raise RuntimeError(err or "ISC scrape failed")

        # 2) IBM Scraper — install base for the same CovIDs (reads selected_covids.json).
        set_state(phase="ibm", message="Scanning IBM's install base for your accounts…")
        rc, err = _launch_and_wait("ibm", [])
        if err:
            raise RuntimeError(f"IBM Scraper couldn't start: {err}")

        # 3) Account Segmentation — join it all together.
        set_state(phase="segment", message="Segmenting and organizing your accounts…")
        rc, err = _launch_and_wait("segment", [])
        if err:
            raise RuntimeError(f"Account Segmentation couldn't start: {err}")
        if rc != 0:
            raise RuntimeError("Account Segmentation finished with an error — see the log.")

        seg_path = REPO_ROOT / "Account_Segmentation" / "output" / "latest.xlsx"
        rows = _row_count(seg_path, "Segmented Accounts")
        # Pre-warm the results view's parse (the 4 MB / 600-column workbook takes
        # ~1.3s to parse) so the seller's first "Show results" click is instant.
        try:
            _segment_view_rows(seg_path, "Segmented Accounts")
        except Exception:
            pass
        done_msg = "Your accounts are ready."
        if rows:
            done_msg = f"Your accounts are ready — {rows} accounts segmented."
        set_state(phase="done", done=True, active=False, message=done_msg)
    except Exception as e:
        set_state(phase="error", active=False, done=False, error=str(e),
                  message=f"Stopped: {e}")


def _run_outbound_strategy():
    """Account Tiering → Call Planning, run back to back."""
    def set_state(**kw):
        _STRATEGY_STATE.update(kw)

    set_state(active=True, phase="tiering", error=None, done=False,
              message="Scoring and tiering your accounts…")
    try:
        if not (REPO_ROOT / "Account_Segmentation" / "output" / "latest.xlsx").exists():
            raise RuntimeError("Get My Accounts first — there are no segmented accounts to strategize yet.")
        rc, err = _launch_and_wait("step2", [])
        if err:
            raise RuntimeError(f"Account Tiering couldn't start: {err}")
        if rc != 0:
            raise RuntimeError("Account Tiering finished with an error — see the log.")

        set_state(phase="callplanning", message="Laying out your call calendar through year-end…")
        rc, err = _launch_and_wait("step3", [])
        if err:
            raise RuntimeError(f"Call Planning couldn't start: {err}")
        if rc != 0:
            raise RuntimeError("Call Planning finished with an error — see the log.")

        set_state(phase="done", done=True, active=False,
                  message="Your outbound strategy is ready — tiering and call plan.")
    except Exception as e:
        set_state(phase="error", active=False, done=False, error=str(e),
                  message=f"Stopped: {e}")


def _run_fill_contacts(cadence):
    """ZoomInfo Contact Readiness → Salesloft, loading contacts into `cadence`."""
    def set_state(**kw):
        _FILL_STATE.update(kw)

    set_state(active=True, phase="zoominfo", error=None, done=False, cadence=cadence,
              message="Preparing your contacts in ZoomInfo…")
    try:
        # JIT auth verification (I1): probe both sessions LIVE now, rather than
        # launching the whole browser automation only to bounce to a login page
        # mid-run (the recurring "ZoomInfo session expired" the screenshot showed).
        ok, msg = _ensure_services_ready(["zoominfo", "salesloft"])
        if not ok:
            raise RuntimeError(msg)

        today = date.today().isoformat()
        # Pass the chosen cadence to ZoomInfo too, so the export lands in the SAME
        # cadence Salesloft then advances (they used to diverge — ZoomInfo always
        # exported to the config default while Salesloft advanced the dropdown pick).
        rc, err = _launch_and_wait("step4", ["--date", today, "--mode", "auto", "--cadence", cadence])
        if err:
            raise RuntimeError(f"ZoomInfo step couldn't start: {err}")
        if rc != 0:
            raise RuntimeError(_last_step_error("step4",
                               "ZoomInfo Contact Readiness finished with an error — see the log."))

        set_state(phase="salesloft", message=f"Loading your contacts into “{cadence}”…")
        rc, err = _launch_and_wait("step5", ["--cadence", cadence])
        if err:
            raise RuntimeError(f"Salesloft step couldn't start: {err}")
        if rc != 0:
            raise RuntimeError(_last_step_error("step5",
                               "Salesloft Cadence Readiness finished with an error — see the log."))

        set_state(phase="done", done=True, active=False,
                  message=f"Done — your contacts are loaded into “{cadence}”.")
    except Exception as e:
        set_state(phase="error", active=False, done=False, error=str(e),
                  message=f"Stopped: {e}")


# Cadence options offered by the Fill Contacts dropdown (per current spec).
FILL_CADENCES = ["Targeted Outreach Cadence 3", "Targeted Outreach Cadence 4"]

# Bobby's cadence choices mirror the Fill Contacts options (per spec), but shown as
# multiple-choice radios rather than a dropdown.
BOBBY_CADENCES = list(FILL_CADENCES)


def _run_bobby(cadence):
    """Bobby, the AI Emailer — drive Bobby_AI_Emailer/run.py and surface its
    natural-language progress. Bobby reads the chosen Salesloft cadence's email
    steps + enrolled people via the Salesloft API and writes a personalized email
    per person (by cadence day / title / company)."""
    _BOBBY_STATE.update(active=True, phase="running", error=None, done=False,
                        cadence=cadence, summary=None, log=[],
                        message="Waking Bobby up…")
    bobby_dir = REPO_ROOT / "Bobby_AI_Emailer"
    # Fresh, real-time run every time: wipe the previous run's drafts so nothing
    # presaved is ever shown while this run is drafting (or if it errors early).
    try:
        (bobby_dir / "output" / "latest.json").unlink()
    except FileNotFoundError:
        pass
    except Exception:
        pass
    lines = []
    watchdog = None
    try:
        proc = subprocess.Popen(
            [str(VENV_PYTHON), "run.py", "--cadence", cadence], cwd=str(bobby_dir),
            stdin=subprocess.DEVNULL, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        _BOBBY_PROC["proc"] = proc
        # A hung Salesloft browser must not wedge Bobby "running" forever — kill it
        # after a generous ceiling so the run always resolves and can be re-run.
        watchdog = threading.Timer(600, lambda: proc.poll() is None and proc.kill())
        watchdog.daemon = True
        watchdog.start()
        for raw in iter(proc.stdout.readline, b""):
            line = raw.decode(errors="replace").rstrip()
            if not line:
                continue
            lines.append(line)
            clean = _clean_log_line(line)
            if clean:
                _BOBBY_STATE.update(message=clean, log=lines[-40:])
        proc.stdout.close()
        rc = proc.wait()
        if rc == 0:
            summary = None
            try:
                summary = json.loads((bobby_dir / "output" / "latest.json").read_text())
            except Exception:
                pass
            n = summary.get("drafted", 0) if summary else 0
            steps = summary.get("email_step_count", 0) if summary else 0
            _BOBBY_STATE.update(
                active=False, done=True, phase="done", log=lines[-40:], summary=summary,
                message=(f"Bobby drafted {n} email(s) across {steps} email step(s) for “{cadence}”."
                         if n else f"Bobby finished for “{cadence}”."))
        else:
            err = _pick_error_line(lines) or "Bobby finished with an error — see the log."
            _BOBBY_STATE.update(active=False, done=False, phase="error", error=err,
                                log=lines[-40:], message=f"Stopped: {err}")
    except Exception as e:
        _BOBBY_STATE.update(active=False, done=False, phase="error", error=str(e),
                            message=f"Stopped: {e}")
    finally:
        if watchdog:
            watchdog.cancel()
        _BOBBY_PROC["proc"] = None
        # Belt-and-suspenders: never leave the flag stuck on if we fell out of the
        # loop without hitting a branch above.
        if _BOBBY_STATE.get("active"):
            _BOBBY_STATE.update(active=False, phase="error",
                                message="Bobby stopped unexpectedly — try again.")


def _clean_log_line(line):
    import re as _re
    l = _re.sub(r"^\d{4}-\d\d-\d\d[ T]\d\d:\d\d:\d\d[.,]\d+\s+", "", line)
    l = _re.sub(r"^(INFO|WARNING|WARN|ERROR|DEBUG)\s+", "", l)
    l = _re.sub(r"^\[[^\]]+\]\s+", "", l)
    return l.strip()


def _pick_error_line(lines):
    for l in reversed(lines):
        low = l.lower()
        if ("error" in low or "expired" in low or "not found" in low or "no salesloft" in low) \
                and "traceback" not in low:
            c = _clean_log_line(l)
            if c:
                return c[:300]
    return None


def _bobby_running():
    """True only if a Bobby subprocess is genuinely alive — so a state flag left
    stuck 'active' by a crash/kill doesn't permanently block re-runs."""
    proc = _BOBBY_PROC.get("proc")
    return bool(_BOBBY_STATE["active"] and proc is not None and proc.poll() is None)


@app.route("/api/bobby/run", methods=["POST"])
def api_bobby_run():
    if _bobby_running():
        return jsonify({"ok": False, "error": "already running"})
    body = request.json or {}
    cadence = (body.get("cadence") or "").strip()
    if cadence not in BOBBY_CADENCES:
        return jsonify({"ok": False, "error": "pick a cadence for Bobby to write for"})
    # Clear any stale finished/errored state before a fresh real-time run.
    _BOBBY_STATE.update(active=False, done=False, summary=None, error=None, log=[])
    threading.Thread(target=_run_bobby, args=(cadence,), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/bobby/reset", methods=["POST"])
def api_bobby_reset():
    """Clear Bobby's drafts/state so leaving and returning starts clean — nothing
    presaved. Refuses while a run is genuinely in progress."""
    if _bobby_running():
        return jsonify({"ok": False, "error": "a run is in progress"})
    _BOBBY_STATE.update(active=False, phase="idle", message="", error=None,
                        cadence=None, done=False, summary=None, log=[])
    try:
        (REPO_ROOT / "Bobby_AI_Emailer" / "output" / "latest.json").unlink()
    except FileNotFoundError:
        pass
    except Exception:
        pass
    return jsonify({"ok": True})


@app.route("/bobby")
def bobby_page():
    """Bobby's own page: live progress while drafting, then the email steps grouped
    by day with each person's drafted email + a Send All button."""
    return render_template_string(BOBBY_PAGE_TEMPLATE)


@app.route("/api/bobby/state")
def api_bobby_state():
    """Bobby's LIVE run state + the current run's structured drafts (in-memory only).
    Deliberately does NOT read the last run's latest.json off disk — that produced a
    stale/presaved view; each Run Bobby is a fresh real-time run."""
    return jsonify(dict(_BOBBY_STATE))


@app.route("/api/bobby/send", methods=["POST"])
def api_bobby_send():
    """Send All — currently gated (see Bobby_AI_Emailer/bobby.send_all): Bobby drafts
    the emails but the live Salesloft send is not calibrated/enabled yet, so this
    returns a clear message rather than firing an untested bulk send."""
    if _bobby_running():
        return jsonify({"ok": False, "error": "Bobby is still drafting — wait for it to finish."})
    bobby_dir = REPO_ROOT / "Bobby_AI_Emailer"
    proc = subprocess.run(
        [str(VENV_PYTHON), "-c", "import bobby; bobby.send_all()"],
        cwd=str(bobby_dir), stdin=subprocess.DEVNULL,
        capture_output=True, text=True)
    if proc.returncode == 0:
        return jsonify({"ok": True, "message": "Sent."})
    # Surface the (gated) reason cleanly.
    err = (proc.stderr or proc.stdout or "").strip().splitlines()
    msg = next((l for l in reversed(err) if "Error" in l or "enabled" in l or "send" in l.lower()),
               err[-1] if err else "Send is not available yet.")
    msg = msg.split(": ", 1)[-1] if ": " in msg else msg
    return jsonify({"ok": False, "error": msg[:400]})


@app.route("/api/seller")
def api_seller():
    """Who's signed in and how many territories map to them — powers the Get My
    Accounts card's subtitle. Never returns the password, only the email/name."""
    email = _signed_in_email()
    if not email:
        return jsonify({"signed_in": False})
    try:
        seller = seller_accounts.resolve_seller(email)
    except Exception:
        seller = {"matched": False, "seller_name": None, "covids": [], "industries": {}}
    covids = seller.get("covids") or []
    seller_name = seller.get("seller_name")
    if not seller.get("matched") or not covids:
        # WatsonX Clone: any email works — show its stable demo territory (matching
        # what Get My Accounts will actually pull), so the card never says "no territory".
        covids = fake_data.demo_covids_for_email(email)
        seller_name = seller_name or email.split("@")[0].replace(".", " ").title()
    return jsonify({
        "signed_in": True, "email": email,
        "seller_name": seller_name,
        "matched": True,
        "covids": len(covids),
        "industries": seller.get("industries") or {},
    })


@app.route("/api/get_my_accounts/run", methods=["POST"])
def api_get_my_accounts_run():
    if _GMA_STATE["active"]:
        return jsonify({"ok": False, "error": "already running"})
    if not _signed_in_email():
        return jsonify({"ok": False, "error": "sign in with your IBM email first"})
    threading.Thread(target=_run_get_my_accounts, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/outbound_strategy/run", methods=["POST"])
def api_outbound_strategy_run():
    if _STRATEGY_STATE["active"]:
        return jsonify({"ok": False, "error": "already running"})
    threading.Thread(target=_run_outbound_strategy, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/fill_contacts/run", methods=["POST"])
def api_fill_contacts_run():
    if _FILL_STATE["active"]:
        return jsonify({"ok": False, "error": "already running"})
    body = request.json or {}
    cadence = (body.get("cadence") or "").strip()
    if cadence not in FILL_CADENCES:
        return jsonify({"ok": False, "error": "pick a cadence to load contacts into"})
    threading.Thread(target=_run_fill_contacts, args=(cadence,), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/view/strategy")
def view_strategy():
    """Combined Outbound Strategy result — two tabs (Account Tiering + Call
    Planning), each the existing purpose-built view, embedded side by side."""
    tier_path = REPO_ROOT / "Account_Tiering" / "output" / "latest.xlsx"
    plan_path = REPO_ROOT / "Call_Planning" / "output" / "latest.xlsx"
    return render_template_string(
        STRATEGY_TEMPLATE,
        has_tier=tier_path.exists(), has_plan=plan_path.exists(),
    )


def _reset_for_fresh_demo():
    """Wipe every step's generated output/logs/checkpoints on dashboard
    startup, so the UI always opens on a clean slate instead of showing a
    prior run's stale 'done, N rows, Redo' status (2026-07-06, explicit user
    request — this was confusing during a live demo walkthrough). Only
    touches generated-artifact folders, never source code or .env/auth
    files (those live outside these three subfolder names entirely)."""
    # STEPS dirs, plus Bobby (not a pipeline step) so its drafts never persist
    # across restarts as a presaved view.
    dirs = [REPO_ROOT / step["dir"] for step in STEPS.values()] + [REPO_ROOT / "Bobby_AI_Emailer"]
    for step_dir in dirs:
        for sub in ("output", "logs", "checkpoints"):
            d = step_dir / sub
            if not d.exists():
                continue
            for entry in d.iterdir():
                if sub == "checkpoints" and entry.is_file() and entry.name in _DURABLE_CHECKPOINTS:
                    continue  # keep the expensive external-lookup cache — see _DURABLE_CHECKPOINTS
                if entry.is_file():
                    entry.unlink()
                elif entry.is_dir():
                    shutil.rmtree(entry)
    # Start each demo with an empty mock Salesloft (no leftover loaded contacts).
    mock_salesloft.reset()


def main():
    import socket
    import threading as _threading
    import webbrowser

    _reset_for_fresh_demo()
    # Re-secure any session/secret file to owner-only on every launch (I19), in
    # case one was written 0644 by older code or a service added since last run.
    try:
        guard.harden_perms()
    except Exception:
        pass
    # WatsonX Clone: no Meetings backend to launch, no separate ISC launcher to watch,
    # and no real auth watchdog (all sessions are mocked and always 'ready').

    port = 5488
    with socket.socket() as s:
        if s.connect_ex(("127.0.0.1", port)) == 0:
            s2 = socket.socket()
            s2.bind(("", 0))
            port = s2.getsockname()[1]
            s2.close()
    _threading.Timer(0.9, lambda: webbrowser.open(f"http://127.0.0.1:{port}")).start()
    app.run(port=port, debug=False, use_reloader=False, threaded=True)


if __name__ == "__main__":
    main()
