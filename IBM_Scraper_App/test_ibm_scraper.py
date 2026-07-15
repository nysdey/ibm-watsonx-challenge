"""Session-free tests for the IBM Scraper step (no browser, no live portal).

Covers the pure logic: Power local filter, CovID resolution precedence, the
Cloud Dash-callback query construction, the Storage CSV union-merge, and the
ISC-install helpers/tab config. Run:

    ../.venv/bin/python3 test_ibm_scraper.py
"""
import csv
import os
import sys
import tempfile
import time
from pathlib import Path

import openpyxl

import config
import covid_source
import run as ibm_run
import sub_power
import sub_cloud
import sub_storage
import sub_isc_install

PASS = 0
FAIL = 0


def check(name, cond):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  PASS  {name}")
    else:
        FAIL += 1
        print(f"  FAIL  {name}")


def _xlsx(path, header, rows, sheet="Data"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(path)


def _csv(path, header, rows):
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)


# ---------------- Power ----------------
def test_power(tmp):
    # 128-ish columns not needed; DQ resolved by header name here.
    src = tmp / "power_all.xlsx"
    header = ["MACHINE TYPE", "Primary Customer Name", "local coverage type id", "Spend"]
    rows = [
        ["A", "Acme", "T0001", 10],
        ["B", "Acme", "T0001", 20],
        ["C", "Globex", "T0002", 5],
        ["D", "Initech", "T0003", 7],
        ["E", "Blank", None, 1],
    ]
    _xlsx(src, header, rows)
    config.OUTPUT_DIR = tmp / "out"
    res = sub_power.run_power_install(["T0001", "T0003", "T9999"], source_path=src)
    check("power: matched row count (T0001x2 + T0003x1)", res["matched_rows"] == 3)
    check("power: per-covid T0001==2", res["per_covid"]["T0001"] == 2)
    check("power: missing covid reported", res["missing_covids"] == ["T9999"])
    check("power: exact match (T0002 excluded)", res["per_covid"].get("T0002") is None)
    # output file exists with header + 3 rows
    wb = openpyxl.load_workbook(res["latest_path"])
    ws = wb.active
    check("power: output has header+3 rows", ws.max_row == 4)
    check("power: output preserves columns", [c.value for c in ws[1]] == header)
    # missing source file raises
    try:
        sub_power.run_power_install(["T0001"], source_path=tmp / "nope.xlsx")
        check("power: missing file raises", False)
    except FileNotFoundError:
        check("power: missing file raises", True)


# ---------------- covid_source precedence ----------------
def test_covid_source(tmp):
    # explicit wins
    ids, src = covid_source.resolve_covids(["T1", "T1", "T2"])
    check("covid: explicit dedup+order", ids == ["T1", "T2"] and src == "explicit")

    # selected_covids.json
    sel = tmp / "selected_covids.json"
    sel.write_text('{"covids": ["T0A", "T0B"]}')
    config.SELECTED_COVIDS_PATH = sel
    config.DEDUPED_ACCOUNTS_PATH = tmp / "none.xlsx"
    ids, src = covid_source.resolve_covids(None)
    check("covid: from selected_covids.json", ids == ["T0A", "T0B"] and "selected_covids" in src)

    # fallback to DEDUPED Coverage ID
    config.SELECTED_COVIDS_PATH = tmp / "no.json"
    ded = tmp / "deduped.xlsx"
    _xlsx(ded, ["Account Name", "Coverage ID"], [["X", "T0100"], ["Y", "T0100;T0200"]], sheet="Company Rollup")
    config.DEDUPED_ACCOUNTS_PATH = ded
    ids, src = covid_source.resolve_covids(None)
    check("covid: fallback parses Coverage ID (dedup)", ids == ["T0100", "T0200"] and "DEDUPED" in src)

    # fallback finds Coverage ID even when it's NOT on the active/first sheet:
    # openpyxl makes the first-created sheet active, so put a decoy sheet first.
    config.SELECTED_COVIDS_PATH = tmp / "still_no.json"
    ded2 = tmp / "deduped2.xlsx"
    wb = openpyxl.Workbook()
    decoy = wb.active
    decoy.title = "Companies by Industry"       # active sheet, no Coverage ID
    decoy.append(["Industry", "Company Count"])
    decoy.append(["Tech", 5])
    roll = wb.create_sheet("Company Rollup")     # real data on a later sheet
    roll.append(["Account Name", "Coverage ID"])
    roll.append(["X", "T0300"])
    roll.append(["Y", "T0300;T0400"])
    wb.save(ded2)
    config.DEDUPED_ACCOUNTS_PATH = ded2
    ids, src = covid_source.resolve_covids(None)
    check("covid: fallback scans non-active sheets", ids == ["T0300", "T0400"])

    # none -> raises
    config.SELECTED_COVIDS_PATH = tmp / "no2.json"
    config.DEDUPED_ACCOUNTS_PATH = tmp / "gone.xlsx"
    try:
        covid_source.resolve_covids(None)
        check("covid: none raises", False)
    except RuntimeError:
        check("covid: none raises", True)


# ---------------- Cloud Dash-callback query construction ----------------
def test_cloud_query(tmp):
    # Default trailing-12-month filter is a valid pandas-query string.
    dfilter = sub_cloud._default_date_filter()
    check("cloud: date filter has metric_date", "metric_date" in dfilter)
    check("cloud: date filter has kyndryl flag", "kyndryl_account_flag in ['N']" in dfilter)
    check("cloud: date filter has two bounds", dfilter.count("metric_date") == 2)

    # The callback body targets the right table, data source and query.
    body = sub_cloud._table_body("metric_date >= '2025-05-01' & local_coverage_type_id in ['T0016329']")
    check("cloud: callback output is the deal-list table",
          body["output"] == "table-daterange-account-details.children")
    ds = next(i["value"] for i in body["inputs"] if i["id"] == "global-data-source-choice")
    check("cloud: data source = IBM Customer Number", ds == "IBM Customer Number")
    q = next(i["value"] for i in body["inputs"] if i["id"] == "global-filter-query")
    check("cloud: query carries the covid clause", "local_coverage_type_id in ['T0016329']" in q)
    check("cloud: changedPropIds includes both stores",
          set(body["changedPropIds"]) == {"global-data-source-choice.data", "global-filter-query.data"})
    # Column map exposes trailing-12M spend under a friendly header.
    labels = dict(sub_cloud.COLUMN_MAP)
    check("cloud: revenue mapped to T12M spend header",
          labels["revenue_amt"] == "Trailing 12M Cloud Spend")


# ---------------- Storage union merge ----------------
def test_storage_merge(tmp):
    a = tmp / "fs.csv"
    b = tmp / "tape.csv"
    _csv(a, ["Customer", "Capacity"], [["Acme", "10"], ["Globex", "20"]])
    _csv(b, ["Customer", "Drives"], [["Initech", "3"]])  # different 2nd column
    wb, union, total, per_cat = sub_storage._merge_csvs({"FlashSystem + SVC": a, "Tape": b})
    check("storage: leading Storage Category col", union[0] == "Storage Category")
    check("storage: unioned columns", set(union) == {"Storage Category", "Customer", "Capacity", "Drives"})
    check("storage: total rows across families", total == 3)
    check("storage: per-category counts", per_cat["FlashSystem + SVC"] == 2 and per_cat["Tape"] == 1)
    # verify a Tape row has blank Capacity and filled Drives
    ws = wb.active
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    hdr = [c.value for c in ws[1]]
    tape_row = next(r for r in rows if r[hdr.index("Storage Category")] == "Tape")
    check("storage: missing col blank for family", tape_row[hdr.index("Capacity")] in ("", None))
    check("storage: present col filled", tape_row[hdr.index("Drives")] == "3")


# ---------------- ISC install pure helpers ----------------
def test_isc_helpers():
    check("isc: login marker detected", sub_isc_install._looks_like_login("https://login.w3.ibm.com/authsvc"))
    check("isc: valid app url not login", not sub_isc_install._looks_like_login("https://ibmsc.lightning.force.com/lightning/page/analytics"))
    # Both tabs are configured with a dataset and L1_ACCOUNT_NAME as first field.
    for key in ("ibm_non_infra", "competitive"):
        spec = sub_isc_install.TABS[key]
        check(f"isc: {key} has dataset name", bool(spec["dataset_name"]))
        check(f"isc: {key} first field is parent account", spec["fields"][0] == "L1_ACCOUNT_NAME")
    # Per-account breakdown groups rows by parent account.
    rows = [{"L1_ACCOUNT_NAME": "ACME"}, {"L1_ACCOUNT_NAME": "ACME"}, {"L1_ACCOUNT_NAME": "GLOBEX"}, {}]
    counts = sub_isc_install._account_breakdown(rows)
    check("isc: breakdown counts per parent", counts["ACME"] == 2 and counts["GLOBEX"] == 1)
    check("isc: breakdown handles blank account", counts["(blank)"] == 1)


# ---------------- w3id seed freshness ----------------
def test_w3id_seed(tmp):
    import w3id_seed
    w3id_seed._W3ID_SESSION_FILES = []  # isolate from this machine's real sessions
    a = tmp / "isc.json"; b = tmp / "gtm.json"; c = tmp / "cid.json"
    for f in (a, b, c):
        f.write_text("{}")
    # make explicit, distinct mtimes: b oldest, c middle, a newest
    os.utime(b, (1000, 1000))
    os.utime(c, (2000, 2000))
    os.utime(a, (3000, 3000))
    order = w3id_seed.w3id_seeds_by_freshness(extra=[a, b, c])
    check("w3id: freshest (highest mtime) first", order[0] == a and order[-1] == b)
    # missing files are skipped, not returned
    order2 = w3id_seed.w3id_seeds_by_freshness(extra=[tmp / "gone.json", a])
    check("w3id: missing seed skipped", (tmp / "gone.json") not in order2 and a in order2)


# ---------------- cloud auth-wall detection ----------------
def test_cloud_auth_detection():
    # An expired GTM app session bounces to its own /login landing.
    check("cloud: gtm app-login landing detected",
          sub_cloud._gtm_app_login_expired("https://w3.ibm.com/sales/gtm-navigator/login/app"))
    check("cloud: live app url is not app-login",
          not sub_cloud._gtm_app_login_expired("https://w3.ibm.com/sales/gtm-navigator/app"))
    # The two real passkey/LDAP walls we actually observed must both count.
    for wall in ("https://login.w3.ibm.com/authsvc/mtfim/sps/authsvc?PolicyId=urn:...onpremldap",
                 "https://w3id-ns.sso.ibm.com/pages/password-blocked.html",
                 "https://login.ibm.com/authsvc/mtfim/sps/authsvc"):
        check(f"cloud: auth wall detected ({wall.split('//')[1][:20]})", sub_cloud._at_auth_wall(wall))
    check("cloud: cirrus dashboard is not an auth wall",
          not sub_cloud._at_auth_wall("https://gtmnav-dashboard.next-best-workload.dal.app.cirrus.ibm.com/dashboard/"))


# ---------------- run.py freshness / ORDER ----------------
def test_run_freshness(tmp):
    # Full run must produce all five install files (no silent omission) so
    # downstream Segmentation joins a consistent same-run set.
    check("run: full ORDER covers all five sub-scrapers",
          set(ibm_run.ORDER) == set(config.OUTPUT_NAMES))

    config.OUTPUT_DIR = tmp / "fresh_out"
    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    run_started = time.time()

    # ok + file written this run -> FRESH
    _, latest = config.dated_and_latest_paths("power")
    latest.write_text("x")  # mtime = now (>= run_started)
    state, _ = ibm_run._freshness("power", {"status": "ok"}, run_started)
    check("run: ok + rewritten file -> FRESH", state == "FRESH")

    # error but an OLD _latest.xlsx still on disk -> STALE (accuracy hazard)
    _, latest = config.dated_and_latest_paths("storage")
    latest.write_text("old")
    old = run_started - 7200  # 2h before the run started
    os.utime(latest, (old, old))
    state, detail = ibm_run._freshness("storage", {"status": "error"}, run_started)
    check("run: failed but old file present -> STALE", state == "STALE")

    # error and no file at all -> MISSING
    state, _ = ibm_run._freshness("cloud", {"status": "error"}, run_started)
    check("run: failed and no file -> MISSING", state == "MISSING")

    # ok status but file not refreshed this run -> STALE (don't trust the claim)
    _, latest = config.dated_and_latest_paths("competitive")
    latest.write_text("older")
    os.utime(latest, (old, old))
    state, _ = ibm_run._freshness("competitive", {"status": "ok"}, run_started)
    check("run: ok but file not refreshed -> STALE", state == "STALE")


def main():
    with tempfile.TemporaryDirectory() as d:
        tmp = Path(d)
        print("== Power =="); test_power(tmp)
        print("== covid_source =="); test_covid_source(tmp)
        print("== Cloud query =="); test_cloud_query(tmp)
        print("== Storage merge =="); test_storage_merge(tmp)
        print("== ISC install helpers =="); test_isc_helpers()
        print("== w3id seed freshness =="); test_w3id_seed(tmp)
        print("== cloud auth detection =="); test_cloud_auth_detection()
        print("== run freshness =="); test_run_freshness(tmp)
    print(f"\nIBM Scraper tests: {PASS} passed, {FAIL} failed")
    sys.exit(1 if FAIL else 0)


if __name__ == "__main__":
    main()
