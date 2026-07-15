"""
ISC Territory Prospecting Tool Scraper
Fully automated: logs in, creates prospecting list, filters by Coverage ID,
loads all rows, and exports to Excel.

Usage:
    python isc_scraper.py
    python isc_scraper.py --output my_accounts.xlsx
    python isc_scraper.py --cov-id T0018062
"""

import argparse
import os
import subprocess
import sys
import time
import re
from pathlib import Path

# Ensure Playwright finds Firefox in the system cache (needed for frozen exe)
os.environ.setdefault(
    "PLAYWRIGHT_BROWSERS_PATH",
    str(Path.home() / "Library" / "Caches" / "ms-playwright"),
)

from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


ISC_URL = "https://ibmsc.lightning.force.com/"
AUTH_STATE_FILE = Path.home() / ".isc_scraper" / "auth_state.json"
DEFAULT_COV_ID = "T0016156"


def _url_is_authenticated(url):
    """True only if `url` is the logged-in Salesforce app, not a login/SSO
    bounce. Used to refuse overwriting a good saved session with a logged-out
    one (I18)."""
    u = (url or "").lower()
    if any(x in u for x in ("login", "w3id", "authsvc", "sso")):
        return False
    return ("lightning.force.com" in u) or ("ibmsc" in u)


def _interactive_ok():
    """True only when a human can respond to a headful browser / input() prompt.
    In unattended/headless runs (ISC_NO_BROWSER=1 or no TTY) we must fail fast
    instead of hanging on wait_for_url(300s) or EOFError'ing on input() (I63)."""
    try:
        return bool(sys.stdin.isatty()) and not os.environ.get("ISC_NO_BROWSER")
    except Exception:
        return False


class TeeOutput:
    """Write to both terminal and a log file that the floating window reads."""

    def __init__(self, log_file):
        self._original = sys.stdout
        self._original_stderr = sys.stderr
        log_file.write_text("")
        self._f = open(log_file, "a", buffering=1)

    def write(self, msg):
        self._original.write(msg)
        self._original.flush()
        try:
            self._f.write(msg)
            self._f.flush()
            import os
            os.fsync(self._f.fileno())
        except Exception:
            pass

    def flush(self):
        self._original.flush()
        try:
            self._f.flush()
        except Exception:
            pass

    def close(self):
        self._f.close()
        sys.stdout = self._original


def wait_and_click(page, selector, timeout=30000, description="element"):
    print(f"  Waiting for {description}...")
    el = page.wait_for_selector(selector, timeout=timeout, state="visible")
    el.scroll_into_view_if_needed()
    time.sleep(0.5)
    el.click()
    print(f"  Clicked {description}")
    return el


def navigate_to_territory_prospecting(page):
    print("\n--- Step 1: Navigate to Territory Prospecting ---")
    page.goto(ISC_URL, wait_until="domcontentloaded", timeout=120_000)

    if "login" in page.url.lower() or "w3id" in page.url.lower() or "sso" in page.url.lower():
        if not _interactive_ok():
            raise RuntimeError(
                "ISC session expired — re-login required. Run the ISC login flow; "
                "unattended/headless mode will not wait for an interactive login."
            )
        print("  Login required. Enter your w3id credentials in the browser window.")
        print("  Waiting for login to complete...")
        page.wait_for_url("**/lightning/**", timeout=300_000)
        print("  Login successful!")

    time.sleep(5)
    wait_and_click(page, "a:has-text('Territory Prospecting'), one-app-nav-bar-item-root:has-text('Territory Prospecting')",
                   description="Territory Prospecting tab")
    time.sleep(5)


def create_prospecting_list(page, args=None):
    print("\n--- Step 2: Create Prospecting List ---")

    time.sleep(5)

    # Uncheck "Use my territory" toggle
    # It's a slds-checkbox_toggle — click the faux span to toggle it
    faux = page.query_selector("span.slds-checkbox_faux[part='indicator']")
    if faux and faux.is_visible():
        faux.click()
        print("  Toggled 'Use my territory'")
        time.sleep(3)
    else:
        # Fallback: find the toggle input near "Use my territory" text
        toggled = page.evaluate("""() => {
            const labels = document.querySelectorAll('label');
            for (const label of labels) {
                if (label.textContent.includes('Use my territory')) {
                    const input = label.querySelector('input[type="checkbox"]');
                    if (input) { input.click(); return true; }
                    const faux = label.querySelector('.slds-checkbox_faux');
                    if (faux) { faux.click(); return true; }
                }
            }
            // Try shadow DOM
            const allInputs = document.querySelectorAll('input[type="checkbox"]');
            for (const inp of allInputs) {
                const parent = inp.closest('.slds-checkbox_toggle, .slds-form-element');
                if (parent && parent.textContent.includes('Use my territory')) {
                    inp.click();
                    return true;
                }
            }
            return false;
        }""")
        if toggled:
            print("  Toggled 'Use my territory' (fallback)")
            time.sleep(3)
        else:
            print("  Could not find 'Use my territory' toggle. Please uncheck it manually.")
            if not _interactive_ok():
                raise RuntimeError("Could not toggle 'Use my territory' and no "
                                   "interactive session to fix it (unattended mode).")
            input("  Press Enter when done...")

    # Click "Create Prospecting List" button
    wait_and_click(page, "text=Create Prospecting List", description="Create Prospecting List button")
    time.sleep(3)

    # Wait for the dialog
    page.wait_for_selector("text=Create New Prospecting List", timeout=30_000)
    print("  Create dialog appeared")

    # Fill in the name
    name_input = page.wait_for_selector("input[placeholder='Prospecting List Title']", timeout=10_000)
    name_input.fill("")
    chunk_idx = getattr(args, 'chunk_idx', 0)
    list_name = f"Scrape Chunk {chunk_idx}" if getattr(args, 'cov_ids_json', None) else f"Scrape {getattr(args, 'cov_id', 'Auto')}"
    name_input.fill(list_name)
    print("  Entered name: Automated Scrape")
    time.sleep(1)

    # "Accounts with CMR" should already be selected (it's the default)
    # Click Save
    wait_and_click(page, "button:has-text('Save')", description="Save button")
    time.sleep(5)
    print("  Prospecting list created!")


def add_state_column(page):
    print("\n--- Step 2.5: Add State/Province column ---")

    wait_and_click(page, "text=Edit columns", description="Edit columns button")
    time.sleep(3)

    # Find and click the State/Province checkbox
    # The checkbox input has an ID containing "PRMRY_ST_PROV_NAME"
    # Click its label to toggle it
    state_cb = page.query_selector("label[for*='PRMRY_ST_PROV_NAME']")
    if state_cb:
        state_cb.scroll_into_view_if_needed()
        time.sleep(0.5)
        state_cb.click()
        print("  Checked State/Province")
    else:
        # Fallback: find by text
        found = False
        labels = page.query_selector_all("label")
        for label in labels:
            if "State/Province" in label.inner_text() and "Code" not in label.inner_text():
                label.scroll_into_view_if_needed()
                time.sleep(0.5)
                label.click()
                found = True
                print("  Checked State/Province")
                break
        if not found:
            print("  Could not find State/Province checkbox.")
            print("  Please check it manually in the browser.")
            if not _interactive_ok():
                raise RuntimeError("Could not find State/Province checkbox and no "
                                   "interactive session to fix it (unattended mode).")
            input("  Press Enter when done...")

    time.sleep(1)

    # Click the blue Apply button
    apply_btns = page.query_selector_all("button")
    for btn in apply_btns:
        if btn.is_visible() and btn.inner_text().strip() == "Apply":
            btn.click()
            print("  Clicked Apply")
            break

    time.sleep(5)


def filter_by_coverage_id(page, cov_id):
    print(f"\n--- Step 3: Filter by Coverage ID: {cov_id} ---")

    # Scroll to top first to ensure filter section is visible
    page.evaluate("window.scrollTo(0, 0)")
    time.sleep(2)

    # Click on "Account Details" filter section to expand it
    acct_details = page.query_selector("text=Account Details")
    if acct_details:
        acct_details.scroll_into_view_if_needed()
        time.sleep(0.5)
        acct_details.click(force=True)
        print("  Clicked Account Details dropdown")
    else:
        wait_and_click(page, "text=Account Details", description="Account Details dropdown")
    time.sleep(3)

    # The Coverage ID field is a pill/tag input
    # Find the input near the "Coverage ID" label
    cov_input = page.evaluate_handle("""() => {
        // Find all elements containing "Coverage ID" text
        const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
        while (walker.nextNode()) {
            if (walker.currentNode.textContent.trim() === 'Coverage ID') {
                // Walk up to find the parent form element
                let parent = walker.currentNode.parentElement;
                for (let i = 0; i < 10; i++) {
                    if (!parent) break;
                    const input = parent.querySelector('input');
                    if (input) return input;
                    parent = parent.parentElement;
                }
            }
        }
        return null;
    }""")

    el = cov_input.as_element()
    if el:
        el.scroll_into_view_if_needed()
        time.sleep(0.5)
        el.click(force=True)
        time.sleep(0.5)
        el.fill(cov_id)
        print(f"  Entered Coverage ID: {cov_id}")
        time.sleep(1)
        el.press("Enter")
        time.sleep(2)
    else:
        print("  Could not find Coverage ID input automatically.")
        print("  Please enter the Coverage ID manually in the browser.")
        if not _interactive_ok():
            raise RuntimeError("Could not find Coverage ID input and no interactive "
                               "session to fix it (unattended mode).")
        input("  Press Enter here when done...")

    # Click the blue Apply button in the filter popover
    apply_btns = page.query_selector_all("button")
    for btn in apply_btns:
        if btn.is_visible() and btn.inner_text().strip() == "Apply":
            btn.click()
            print("  Clicked Apply")
            break
    time.sleep(5)
    print("  Filter applied!")


def load_all_rows(page):
    print("\n--- Step 4: Loading all rows ---")

    # Get total count
    body_text = page.inner_text("body")
    total = None
    m = re.search(r"Total:\s*([\d,]+)", body_text)
    if m:
        total = int(m.group(1).replace(",", ""))
    print(f"  Total rows expected: {total}")

    iteration = 0
    while True:
        body_text = page.inner_text("body")
        available = None
        m = re.search(r"Available:\s*([\d,]+)", body_text)
        if m:
            available = int(m.group(1).replace(",", ""))

        current_row_count = len(page.query_selector_all("table tbody tr"))
        print(f"  Available: {available}/{total} | DOM rows: {current_row_count}")

        if total and available and available >= total:
            print("  All rows loaded!")
            break

        # Find and click Load more
        load_more = None
        for sel in ["a:has-text('Load more')", "a:has-text('Load More')",
                     "button:has-text('Load more')", "span:has-text('Load more')"]:
            el = page.query_selector(sel)
            if el and el.is_visible():
                load_more = el
                break

        if not load_more:
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(2)
            for sel in ["a:has-text('Load more')", "a:has-text('Load More')"]:
                el = page.query_selector(sel)
                if el and el.is_visible():
                    load_more = el
                    break

        if not load_more:
            print("  No 'Load more' button found. Stopping.")
            break

        load_more.scroll_into_view_if_needed()
        time.sleep(0.3)
        load_more.click()
        iteration += 1

        # Wait for new rows
        prev_count = current_row_count
        waited = 0
        while waited < 60:
            time.sleep(1)
            waited += 1
            new_count = len(page.query_selector_all("table tbody tr"))
            if new_count > prev_count:
                time.sleep(1)
                break

        if iteration % 3 == 0:
            print(f"  [Progress] Clicked Load more {iteration} times")


def extract_table(page):
    print("\n--- Step 5: Extracting table data ---")

    # Extract headers from title attributes (avoids "Sort by:" prefix)
    headers = []
    ths = page.query_selector_all("table thead th")
    for th in ths:
        # Skip checkbox and row number columns
        div = th.query_selector("div[title]")
        if div:
            title = div.get_attribute("title") or ""
            if title:
                headers.append(title)
                continue
        text = th.inner_text().strip()
        # Clean up "Sort by:\nName" -> "Name"
        if text.startswith("Sort by:"):
            text = text.replace("Sort by:", "").strip()
        if text and text != "#" and "Select All" not in text and "Choose" not in text:
            headers.append(text)

    print(f"  Found {len(headers)} columns: {headers[:5]}...")

    # Extract all rows in batches using eval_on_selector_all (one call per batch)
    trs = page.query_selector_all("table tbody tr")
    total_rows = len(trs)
    print(f"  Found {total_rows} rows in DOM")

    min_cols = len(headers) - 2
    rows = []
    batch_size = 100

    for batch_start in range(0, total_rows, batch_size):
        batch_end = min(batch_start + batch_size, total_rows)
        batch_trs = trs[batch_start:batch_end]

        for tr in batch_trs:
            # Use a single evaluate call per row to get all cell data at once
            row_data = tr.evaluate("""tr => {
                const tds = tr.querySelectorAll('td');
                const row = [];
                for (let j = 2; j < tds.length; j++) {
                    const td = tds[j];
                    const div = td.querySelector('div[title]');
                    if (div && div.getAttribute('title')) {
                        row.push(div.getAttribute('title'));
                    } else {
                        const a = td.querySelector('a[title]');
                        if (a && a.getAttribute('title')) {
                            row.push(a.getAttribute('title'));
                        } else {
                            row.push(td.innerText.trim());
                        }
                    }
                }
                return row;
            }""")

            if row_data and len(row_data) >= min_cols and not all(c == "" for c in row_data):
                rows.append(row_data)

        print(f"  Processed {batch_end}/{total_rows} rows...")

    print(f"  Extracted {len(rows)} rows")
    return headers, rows


_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@")


def _sanitize_cell(val):
    """Neutralize CSV/Excel formula injection (CWE-1236) — see http_scraper.py
    for details. Scraped account/company names are external data we don't control."""
    if isinstance(val, str) and val.startswith(_FORMULA_TRIGGER_CHARS):
        return "'" + val
    return val


def save_to_excel(headers, rows, output_path):
    print("\n--- Step 6: Saving to Excel ---")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Territory Prospecting"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            if col_idx <= len(headers):
                ws.cell(row=row_idx, column=col_idx, value=_sanitize_cell(value))

    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        for row_idx in range(2, min(52, len(rows) + 2)):
            cell_len = len(str(ws.cell(row=row_idx, column=col_idx).value or ""))
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"  Saved {len(rows)} rows to {output_path}")


def clear_existing_filter(page):
    """Remove the existing Coverage ID filter pill (the X button on the tag)."""
    print("  Clearing existing Coverage ID filter...")

    # Look for the X button on the coverage ID pill/tag in Active Filters
    close_btns = page.query_selector_all("button[title='Remove'], button.deleteAction, span.deleteIcon")
    for btn in close_btns:
        if btn.is_visible():
            btn.click()
            print("  Removed existing filter")
            time.sleep(3)
            return

    # Fallback: look for X inside the Coverage ID filter area
    pills = page.query_selector_all("span.slds-pill, button.slds-pill__remove, [class*='pill'] button")
    for pill in pills:
        if pill.is_visible():
            pill.click()
            print("  Removed filter pill")
            time.sleep(3)
            return

    # Fallback: click the X on the tag shown in Active Filters section
    x_buttons = page.query_selector_all("button")
    for btn in x_buttons:
        try:
            title = btn.get_attribute("title") or ""
            aria = btn.get_attribute("aria-label") or ""
            if "remove" in title.lower() or "remove" in aria.lower() or "delete" in title.lower():
                if btn.is_visible():
                    btn.click()
                    print("  Removed filter via remove button")
                    time.sleep(3)
                    return
        except Exception:
            continue

    print("  No existing filter found to clear")


def change_coverage_id(page, cov_id):
    """Change the Coverage ID filter on an already-open prospecting list."""
    print(f"\n--- Changing Coverage ID to: {cov_id} ---")

    clear_existing_filter(page)
    time.sleep(2)

    # Now apply new filter
    filter_by_coverage_id(page, cov_id)


def append_to_excel(headers, rows, output_path, cov_id):
    """Append rows to an existing Excel file, or create it if it doesn't exist."""
    output = Path(output_path)

    if output.exists():
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        start_row = ws.max_row + 1
        for row_idx, row in enumerate(rows, start_row):
            for col_idx, value in enumerate(row, 1):
                if col_idx <= len(headers):
                    ws.cell(row=row_idx, column=col_idx, value=_sanitize_cell(value))
        wb.save(output_path)
        print(f"  Appended {len(rows)} rows to {output_path} (CovID: {cov_id})")
    else:
        save_to_excel(headers, rows, output_path)


def run_scraper(args):
    AUTH_STATE_FILE.parent.mkdir(exist_ok=True)

    # Determine the app root (exe folder, not _internal)
    if getattr(sys, 'frozen', False):
        app_root = Path(sys.executable).parent
    else:
        app_root = Path(__file__).parent.parent  # _internal/ → app root

    # Save output files to an "output" subfolder
    output_dir = app_root / "output"
    output_dir.mkdir(exist_ok=True)
    args.output = str(output_dir / Path(args.output).name)

    # Determine which CovIDs to process
    if getattr(args, 'cov_ids_json', None):
        import json as _json
        cov_ids = _json.loads(args.cov_ids_json)
        print(f"  Loaded {len(cov_ids)} territories from pool args")
    else:
        cov_ids = [args.cov_id]

    print("=" * 60)
    print("  ISC Territory Prospecting Scraper")
    print(f"  Territories: {len(cov_ids)}")
    print(f"  Output: {args.output}")
    print("=" * 60)

    with sync_playwright() as p:
        browser = p.firefox.launch(headless=False)

        # Quit entire program if browser is closed by user
        scraper_done = [False]
        def on_disconnect():
            if not scraper_done[0]:
                os._exit(0)
        browser.on("disconnected", on_disconnect)

        if AUTH_STATE_FILE.exists():
            print("\nLoading saved login session...")
            context = browser.new_context(storage_state=str(AUTH_STATE_FILE))
        else:
            print("\nNo saved session. You'll need to log in on first run.")
            context = browser.new_context()

        page = context.new_page()
        page.set_viewport_size({"width": 1920, "height": 1080})

        # Step 1: Navigate and log in
        navigate_to_territory_prospecting(page)

        # Save auth state after successful login — but never overwrite a good
        # saved session with a logged-out one if the page bounced to login (I18),
        # and lock the file down to owner-only (I19).
        if _url_is_authenticated(page.url):
            print("  Saving login session...")
            context.storage_state(path=str(AUTH_STATE_FILE))
            try:
                os.chmod(AUTH_STATE_FILE, 0o600)
            except Exception:
                pass
        else:
            print(f"  Skipping session save — page is not the authenticated app "
                  f"(url={page.url[:60]}); keeping existing saved session.")

        # Step 2: Create prospecting list (only once)
        create_prospecting_list(page, args)

        # Step 2.5: Add State/Province column (only once)
        add_state_column(page)

        # Loop through each CovID
        for idx, cov_id in enumerate(cov_ids):
            print(f"\n{'=' * 60}")
            print(f"  Territory {idx + 1}/{len(cov_ids)}: {cov_id}")
            print(f"{'=' * 60}")

            if idx == 0:
                # First CovID: apply filter normally
                filter_by_coverage_id(page, cov_id)
            else:
                # Subsequent CovIDs: clear old filter and apply new
                change_coverage_id(page, cov_id)

            # Load all rows
            load_all_rows(page)

            # Extract table
            headers, rows = extract_table(page)

            if len(rows) == 0:
                print(f"  No rows found for {cov_id}. Skipping.")
                continue

            # Append to combined Excel
            if idx == 0:
                save_to_excel(headers, rows, args.output)
            else:
                append_to_excel(headers, rows, args.output, cov_id)

            print(f"  {cov_id}: {len(rows)} rows saved")

        print("\n" + "=" * 60)
        print("  All territories scraped!")
        print("=" * 60)

        scraper_done[0] = True
        browser.close()

    if getattr(args, 'no_dedup', False):
        print("All done!")
        return

    # Run dedup
    print("\n--- Removing duplicates ---")
    dedup_output = args.output.replace(".xlsx", "_deduped.xlsx")
    try:
        from dedup import run_dedup
        run_dedup(args.output, dedup_output)
    except Exception as e:
        print(f"  Dedup import failed ({e}), trying subprocess...")
        subprocess.run([
            sys.executable, str(Path(__file__).parent / "dedup.py"),
            "--input", args.output,
            "--output", dedup_output,
        ])
    print("All done!")


def main():
    parser = argparse.ArgumentParser(description="ISC Territory Prospecting Scraper")
    parser.add_argument("--output", default="territory_prospecting_export.xlsx", help="Output Excel file")
    parser.add_argument("--cov-id", default=DEFAULT_COV_ID, help="Single Coverage ID (standalone mode)")
    parser.add_argument("--cov-ids-json", default=None, help="JSON array of Coverage IDs (pool mode)")
    parser.add_argument("--chunk-idx", type=int, default=0, help="Pool chunk index, used for log file naming")
    parser.add_argument("--no-dedup", action="store_true", help="Skip dedup step")
    parser.add_argument("--no-monitor", action="store_true", help="Skip progress Terminal window")
    args = parser.parse_args()

    log_file = Path(__file__).parent / f".scraper_log_{args.chunk_idx}.txt"

    monitor_proc = None
    if not args.no_monitor:
        monitor_script = Path(__file__).parent / "progress_window.py"
        monitor_proc = subprocess.Popen([sys.executable, str(monitor_script)])

    tee = TeeOutput(log_file)
    sys.stdout = tee

    try:
        run_scraper(args)
    finally:
        tee.close()
        if monitor_proc:
            monitor_proc.terminate()


if __name__ == "__main__":
    main()
