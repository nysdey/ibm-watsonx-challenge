"""
Fast HTTP-based scraper for ISC Territory Prospecting.
Replaces browser automation with direct Aura endpoint calls.
Uses saved auth_state.json cookies — no browser needed.
"""

import argparse
import json
import os
import sys
import time
from pathlib import Path

import urllib.request
import urllib.parse
import urllib.error

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

AUTH_STATE_FILE = Path.home() / ".isc_scraper" / "auth_state.json"
ISC_HOST = "https://ibmsc.lightning.force.com"
AURA_URL = ISC_HOST + "/aura?r=1&aura.ApexAction.execute=1"
AURA_PAGE_URI = "/lightning/n/TerritoryProspecting"

ACCOUNT_COLUMNS = [
    {"hidden": False, "id": "DOM_COMP_NAME",         "mandatory": True,  "rightJustify": False, "sortable": True, "title": "Name"},
    {"hidden": False, "id": "PRMRY_CTRY",            "mandatory": False, "rightJustify": False, "sortable": True, "title": "Country"},
    {"hidden": False, "id": "PARENT_NAME",           "mandatory": False, "rightJustify": False, "sortable": True, "title": "Account Name"},
    {"hidden": False, "id": "PARENT_ID",             "mandatory": False, "rightJustify": False, "sortable": True, "title": "Account Number"},
    {"hidden": False, "id": "COV_ID",                "mandatory": False, "rightJustify": False, "sortable": True, "title": "Coverage ID"},
    {"hidden": True,  "id": "COV_NAME",              "mandatory": False, "rightJustify": False, "sortable": True, "title": "Coverage Name"},
    {"hidden": False, "id": "STATUS",                "mandatory": False, "rightJustify": False, "sortable": True, "title": "Technology Client Status"},
    {"hidden": False, "id": "CUST_NO",               "mandatory": False, "rightJustify": False, "sortable": True, "title": "CMR Number"},
    {"hidden": False, "id": "GU_DUNS_NO",            "mandatory": False, "rightJustify": False, "sortable": True, "title": "Global D&B"},
    {"hidden": False, "id": "DUNS_NO",               "mandatory": False, "rightJustify": False, "sortable": True, "title": "Domestic D&B"},
    {"hidden": False, "id": "CM_LOCATION_ID",        "mandatory": False, "rightJustify": False, "sortable": True, "title": "Location ID"},
    {"hidden": False, "id": "CM_COMPANY_ID",         "mandatory": False, "rightJustify": False, "sortable": True, "title": "Company ID"},
    {"hidden": False, "id": "PRMRY_ADR_LINE_1",      "mandatory": False, "rightJustify": False, "sortable": True, "title": "Address"},
    {"hidden": False, "id": "PRMRY_CITY_NAME",       "mandatory": False, "rightJustify": False, "sortable": True, "title": "City"},
    {"hidden": False, "id": "INDIVCOUNT",            "mandatory": False, "rightJustify": False, "sortable": True, "title": "Contact Count"},
    {"hidden": False, "id": "IND_DSCR",              "mandatory": False, "rightJustify": False, "sortable": True, "title": "Industry"},
    {"hidden": False, "id": "SUB_IND_DSCR",          "mandatory": False, "rightJustify": False, "sortable": True, "title": "Sub Industry"},
    {"displayType": "USDcurrency", "hidden": False, "id": "FIRMO_DOM_EMP_CNT",       "mandatory": False, "rightJustify": True, "sortable": True, "title": "Employee Count"},
    {"displayType": "USDcurrency", "hidden": False, "id": "FIRMO_DOM_TURN_OVER_USD", "mandatory": False, "rightJustify": True, "sortable": True, "title": "Location Annual Revenue"},
    {"displayType": "USDcurrency", "hidden": False, "id": "FIRMO_GBL_TURN_OVER_USD", "mandatory": False, "rightJustify": True, "sortable": True, "title": "Global Annual Revenue "},
    {"displayType": "USDcurrency", "hidden": False, "id": "HG_IT_SPEND",             "mandatory": False, "rightJustify": True, "sortable": True, "title": "Total IT Spend"},
    {"displayType": "USDcurrency", "hidden": False, "id": "HG_CLOUD_SPEND",          "mandatory": False, "rightJustify": True, "sortable": True, "title": "Cloud Spend"},
    {"hidden": False, "id": "IS_HQ",                 "mandatory": False, "rightJustify": False, "sortable": True, "title": "Headquarters"},
    {"hidden": False, "id": "HQ_CTRY",               "mandatory": False, "rightJustify": False, "sortable": True, "title": "Headquarters Country"},
    {"hidden": True,  "id": "COMPANY_URL",           "mandatory": False, "rightJustify": False, "sortable": True, "title": "Company URL"},
    {"hidden": False, "id": "LINKEDIN_URL",          "mandatory": False, "rightJustify": False, "sortable": True, "title": "LinkedIn URL"},
    {"displayType": "USDcurrency", "hidden": False, "id": "REV_CY",   "mandatory": False, "rightJustify": True, "sortable": True, "title": "IBM Spend Current Year"},
    {"displayType": "USDcurrency", "hidden": False, "id": "REV_CY_M1","mandatory": False, "rightJustify": True, "sortable": True, "title": "IBM Spend Prior Year"},
    {"displayType": "USDcurrency", "hidden": False, "id": "REV_CY_M2","mandatory": False, "rightJustify": True, "sortable": True, "title": "IBM Spend Prior Year - 1"},
    {"displayType": "USDcurrency", "hidden": False, "id": "REV_CY_M3","mandatory": False, "rightJustify": True, "sortable": True, "title": "IBM Spend Prior Year - 2"},
    {"hidden": False, "id": "PRMRY_ST_PROV_NAME",   "mandatory": False, "rightJustify": False, "sortable": True, "title": "State/Province"},
    {"hidden": True,  "id": "PRMRY_ST_PROV_CD",     "mandatory": False, "rightJustify": False, "sortable": True, "title": "State/Province Code"},
    {"hidden": True,  "id": "PRMRY_POSTAL_CD",       "mandatory": False, "rightJustify": False, "sortable": True, "title": "Postal/Zip Code"},
    {"hidden": True,  "id": "PRMRY_COUNTY_NAME",     "mandatory": False, "rightJustify": False, "sortable": True, "title": "County"},
    {"hidden": True,  "id": "IBM_GBL_IMT_DSCR",     "mandatory": False, "rightJustify": False, "sortable": True, "title": "Market"},
    {"hidden": True,  "id": "IBM_GBL_IOT_CD",       "mandatory": False, "rightJustify": False, "sortable": True, "title": "Geography Code"},
    {"hidden": True,  "id": "IBM_GBL_IOT_DSCR",     "mandatory": False, "rightJustify": False, "sortable": True, "title": "Geography"},
    {"hidden": True,  "id": "IBM_GBL_IMT_CD",       "mandatory": False, "rightJustify": False, "sortable": True, "title": "Market Code"},
    {"hidden": True,  "id": "REGION_DESC",           "mandatory": False, "rightJustify": False, "sortable": True, "title": "Region"},
    {"hidden": True,  "id": "REGION_CD",             "mandatory": False, "rightJustify": False, "sortable": True, "title": "Region Code"},
    {"hidden": True,  "id": "CTRY",                  "mandatory": False, "rightJustify": False, "sortable": True, "title": "Country Code"},
    {"hidden": True,  "id": "ISSUING_COUNTRY_NO",    "mandatory": False, "rightJustify": False, "sortable": True, "title": "CMR Issuing Country Code"},
    {"hidden": True,  "id": "SAP_CUSTOMER_NO",       "mandatory": False, "rightJustify": False, "sortable": True, "title": "SAP Customer Number"},
    {"hidden": True,  "id": "DOM_CLIENT_ID",         "mandatory": False, "rightJustify": False, "sortable": True, "title": "Domestic Client ID"},
    {"hidden": True,  "id": "BUY_GRP_ID",            "mandatory": False, "rightJustify": False, "sortable": True, "title": "Domestic Buying Group ID"},
    {"hidden": True,  "id": "MAIN_IND_CD",           "mandatory": False, "rightJustify": False, "sortable": True, "title": "Industry Code"},
    {"hidden": True,  "id": "MAIN_SUB_IND_CD",       "mandatory": False, "rightJustify": False, "sortable": True, "title": "Sub Industry Code"},
    {"hidden": True,  "id": "SUB_SEGMENT",           "mandatory": False, "rightJustify": False, "sortable": True, "title": "Client Segment"},
    {"hidden": True,  "id": "GBL_BUY_GRP_ID",       "mandatory": False, "rightJustify": False, "sortable": True, "title": "Global Buying Group ID"},
    {"hidden": True,  "id": "GBL_CLIENT_ID",         "mandatory": False, "rightJustify": False, "sortable": True, "title": "Global Client ID"},
    {"hidden": True,  "id": "URN_IDM_COMP",          "mandatory": False, "rightJustify": False, "sortable": True, "title": "IBM Marketing ID"},
    {"hidden": True,  "id": "LOCATION_ID",           "mandatory": False, "rightJustify": False, "sortable": True, "title": "ZoomInfo Location ID"},
    {"hidden": True,  "id": "COMPANY_ID",            "mandatory": False, "rightJustify": False, "sortable": True, "title": "ZoomInfo Company ID"},
    {"hidden": True,  "id": "COMP_PH_NUM",           "mandatory": False, "rightJustify": False, "sortable": True, "title": "Company Phone Number"},
    {"hidden": True,  "id": "LOCAL_LANG_NAME_1",     "mandatory": False, "rightJustify": False, "sortable": True, "title": "Local Language-1"},
    {"hidden": True,  "id": "LOCAL_LANG_NAME_2",     "mandatory": False, "rightJustify": False, "sortable": True, "title": "Local Language-2"},
    {"hidden": True,  "id": "IS_TAM_ACCOUNT",        "mandatory": False, "rightJustify": False, "sortable": True, "title": "TAM Account"},
]

HEADERS_ORDER = [col["title"] for col in ACCOUNT_COLUMNS if not col["hidden"]]


AURA_BOOTSTRAP_FILE = Path.home() / ".isc_scraper" / "aura_bootstrap.json"


def _load_session():
    """Return the raw auth_state dict (cookies list + origins)."""
    if not AUTH_STATE_FILE.exists():
        raise RuntimeError(
            f"No auth state at {AUTH_STATE_FILE}. "
            "Run the browser scraper once first to log in."
        )
    with open(AUTH_STATE_FILE) as f:
        return json.load(f)


def _create_scraper_list(page):
    """Create a new prospecting list with 'Use my territory' toggled OFF.
    Returns the list ID string, or None if creation failed.
    Called once per bootstrap so every subsequent HTTP scrape uses this list,
    which has no territory restriction and returns accounts for any CovID."""
    import urllib.parse as _up
    try:
        # The toggle is a slds-checkbox_toggle. It defaults to ON (checked).
        # We need to turn it OFF before creating the list.
        page.wait_for_timeout(1500)

        # Try to click the faux span (the visual part of the toggle)
        toggled = False
        faux = page.query_selector("span.slds-checkbox_faux[part='indicator']")
        if faux and faux.is_visible():
            faux.click()
            toggled = True
        else:
            toggled = page.evaluate("""() => {
                const labels = document.querySelectorAll('label');
                for (const label of labels) {
                    if (label.textContent.includes('Use my territory')) {
                        const faux = label.querySelector('.slds-checkbox_faux, span[part="indicator"]');
                        if (faux) { faux.click(); return true; }
                        const inp = label.querySelector('input[type="checkbox"]');
                        if (inp) { inp.click(); return true; }
                    }
                }
                const all = document.querySelectorAll('.slds-checkbox_toggle, .slds-form-element');
                for (const el of all) {
                    if (el.textContent.includes('Use my territory')) {
                        const faux = el.querySelector('.slds-checkbox_faux, span[part="indicator"]');
                        if (faux) { faux.click(); return true; }
                    }
                }
                return false;
            }""")
        if toggled:
            print("  Toggled 'Use my territory' OFF")
            page.wait_for_timeout(800)
        else:
            print("  WARNING: could not find 'Use my territory' toggle — list may still be territory-restricted")

        # Click "Create Prospecting List" button
        create_btn = None
        for sel in ("button:has-text('Create Prospecting List')",
                    "button:has-text('Create Prospecting')",
                    "[title='Create Prospecting List']"):
            try:
                el = page.wait_for_selector(sel, timeout=5000, state="visible")
                if el:
                    create_btn = el
                    break
            except Exception:
                pass
        if not create_btn:
            print("  Could not find 'Create Prospecting List' button")
            return None
        create_btn.click()
        page.wait_for_timeout(1000)

        # Dismiss any "CSS Error / Refresh" interrupt that Salesforce sometimes
        # shows when a new dialog opens — it has a Close button.
        try:
            close_btn = page.wait_for_selector("button:has-text('Refresh')", timeout=1500, state="visible")
            if close_btn:
                page.query_selector("button.slds-button__close, button:has-text('×'), button[title='Close']")
        except Exception:
            pass

        # Dialog appears — find the title input by placeholder (id changes each session)
        name_input = None
        for loc_str in ("input[placeholder='Prospecting List Title']",
                        "input[placeholder*='Prospecting List']",
                        "input[placeholder*='List Title']"):
            try:
                loc = page.locator(loc_str)
                if loc.count() > 0 and loc.first.is_visible():
                    name_input = loc.first
                    break
            except Exception:
                pass
        if not name_input:
            # Fallback: find visible input inside the dialog
            try:
                dialog = page.locator("[role='dialog']")
                inputs = dialog.locator("input:visible")
                if inputs.count() > 0:
                    name_input = inputs.first
            except Exception:
                pass
        if not name_input:
            print("  Create dialog input not found")
            return None

        import time as _time
        list_name = f"ORUM_Scraper_{_time.strftime('%Y%m%d_%H%M%S')}"
        name_input.fill(list_name)
        page.wait_for_timeout(400)

        # Intercept the Aura request BEFORE clicking Save so we don't miss it
        list_id_capture = {}
        def _capture_list_id(req):
            if list_id_capture or "/aura" not in req.url or req.method != "POST":
                return
            body = req.post_data or ""
            if "getAccountPageContents" not in body:
                return
            try:
                parsed = _up.parse_qs(body)
                msg = json.loads(parsed["message"][0])
                lid = msg["actions"][0]["params"]["params"].get("id")
                if lid:
                    list_id_capture["id"] = lid
            except Exception:
                pass
        page.on("request", _capture_list_id)

        # Click Save
        save_btn = None
        for sel in ("button:has-text('Save')", "button[title='Save']"):
            try:
                el = page.locator(sel).first
                if el.is_visible():
                    save_btn = el
                    break
            except Exception:
                pass
        if not save_btn:
            print("  Could not find Save button in dialog")
            return None
        save_btn.click()
        page.wait_for_timeout(3000)

        if list_id_capture.get("id"):
            return list_id_capture["id"]

        # Fallback: parse list ID from URL
        try:
            import re
            m = re.search(r"(a12g[A-Za-z0-9]{12,})", page.url)
            if m:
                return m.group(1)
        except Exception:
            pass
        return None

    except Exception as e:
        print(f"  _create_scraper_list failed: {e}")
        return None


def _url_is_authenticated(url):
    """True only if `url` is the logged-in Salesforce app, not a login/SSO
    bounce. Used to refuse overwriting a good saved session with a logged-out
    one (I18)."""
    u = (url or "").lower()
    if any(x in u for x in ("login", "w3id", "authsvc", "sso")):
        return False
    return ("lightning.force.com" in u) or ("ibmsc" in u)


def _interactive_ok():
    """True only when a human can respond to a headful browser prompt. In
    unattended/headless runs (ISC_NO_BROWSER=1 or no TTY) we must fail fast with
    a clear message instead of launching a headful login + wait_for_url(300s)
    that would hang forever on DEVNULL stdin (I63)."""
    try:
        return bool(sys.stdin.isatty()) and not os.environ.get("ISC_NO_BROWSER")
    except Exception:
        return False


def bootstrap_aura(auth_state):
    """
    Launch a headless Firefox, load Territory Prospecting with saved cookies,
    extract aura token + context via JS, save to aura_bootstrap.json.
    Returns (cookies_dict, aura_token, aura_context).
    """
    import os
    os.environ.setdefault(
        "PLAYWRIGHT_BROWSERS_PATH",
        str(Path.home() / "Library" / "Caches" / "ms-playwright"),
    )
    from playwright.sync_api import sync_playwright

    print("  Bootstrapping aura session (headless)...")
    with sync_playwright() as p:
        browser = p.firefox.launch(headless=True)
        ctx = browser.new_context(storage_state=auth_state)
        page = ctx.new_page()
        page.goto(ISC_HOST + AURA_PAGE_URI, wait_until="networkidle", timeout=90_000)

        if "login" in page.url.lower() or "w3id" in page.url.lower():
            if not _interactive_ok():
                browser.close()
                raise RuntimeError(
                    "ISC session expired — re-login required. Run the ISC login "
                    "flow; unattended/headless mode will not open an interactive "
                    "browser."
                )
            browser.close()
            print("  Session expired. Opening browser for login...")
            browser = p.firefox.launch(headless=False)
            ctx = browser.new_context()
            page = ctx.new_page()
            page.set_viewport_size({"width": 1280, "height": 800})
            page.goto(ISC_HOST + AURA_PAGE_URI, wait_until="domcontentloaded", timeout=120_000)
            page.wait_for_url("**/lightning/**", timeout=300_000)
            print("  Login detected. Saving session...")
            # Only persist a session that actually landed on the app, never a
            # login/SSO bounce (I18); lock it down to 0600 (I19).
            if _url_is_authenticated(page.url):
                ctx.storage_state(path=str(AUTH_STATE_FILE))
                try:
                    os.chmod(AUTH_STATE_FILE, 0o600)
                except Exception:
                    pass
            page.wait_for_load_state("networkidle", timeout=30_000)

        # Intercept a real ApexAction Aura POST (any one — token/context/headers
        # are shared across all Aura calls on the page) to capture the exact
        # request headers Salesforce expects, not guessed ones.
        captured = {}
        def _on_response(resp):
            if captured.get("done"):
                return
            req = resp.request
            if req.method != "POST" or "/aura" not in req.url:
                return
            try:
                import urllib.parse as _up
                body = req.post_data or ""
                if not body:
                    return
                parsed = _up.parse_qs(body)
                if "aura.token" not in parsed or "aura.context" not in parsed:
                    return
                if resp.status != 200:
                    return
                headers = {k.lower(): v for k, v in req.headers.items()}
                captured["token"] = parsed["aura.token"][0]
                captured["context"] = json.loads(parsed["aura.context"][0])
                captured["headers"] = {
                    "user-agent": headers.get("user-agent", "Mozilla/5.0"),
                    "referer": headers.get("referer", ISC_HOST + AURA_PAGE_URI),
                    "origin": headers.get("origin", ISC_HOST),
                    "accept": headers.get("accept", "*/*"),
                    "accept-language": headers.get("accept-language", "en-US,en;q=0.9"),
                }
                # Prefer a real ApexAction.execute call (guarantees this exact
                # request/response round-trip actually succeeded end-to-end),
                # but keep the first-seen capture as a fallback.
                if "aura.ApexAction.execute" in req.url:
                    captured["done"] = True
            except Exception:
                pass

        page.on("response", _on_response)
        # Trigger a fresh round of Aura requests (initial goto may have been
        # served from bfcache/service-worker with no new network activity).
        page.reload(wait_until="networkidle", timeout=60_000)
        if not captured.get("done"):
            page.wait_for_timeout(3000)

        # Create a fresh prospecting list with "Use my territory" toggled OFF.
        # This is the root cause of 0-row results for CovIDs outside the user's
        # own assigned territory: the existing lists in getProspectLists were
        # created with the toggle ON, so Salesforce silently filters them to
        # only the user's own territory and returns 0 for everything else.
        # A list created with the toggle OFF has no such restriction and returns
        # all accounts for any CovID — confirmed live against T0016227 (IL FSS).
        scraper_list_id = _create_scraper_list(page)
        if scraper_list_id:
            print(f"  Scraper list (territory OFF): {scraper_list_id}")
        else:
            print("  WARNING: could not create territory-OFF list — will fall back to existing lists")
        captured["scraper_list_id"] = scraper_list_id

        final_url = page.url
        browser.close()

    if not captured.get("token"):
        raise RuntimeError("Could not capture aura token from network requests.")

    token = captured["token"]
    aura_context = captured.get("context", {"mode": "PROD", "app": "one:one", "dn": [], "globals": {}, "uad": True})
    aura_headers = captured.get("headers", {})
    scraper_list_id = captured.get("scraper_list_id")

    # Re-read cookies from saved auth state (may have been refreshed after login)
    with open(AUTH_STATE_FILE) as f:
        fresh_state = json.load(f)
    cookies_dict = {c["name"]: c["value"] for c in fresh_state.get("cookies", [])}

    # Never persist a bootstrap captured on a logged-out/login page over a good
    # one — a bounce to login would otherwise overwrite a valid saved session
    # with an expired one (I18). If it bounced, keep the existing file and just
    # use the captured token for this run.
    if _url_is_authenticated(final_url):
        AURA_BOOTSTRAP_FILE.parent.mkdir(exist_ok=True)
        tmp_path = AURA_BOOTSTRAP_FILE.with_suffix(f".tmp{os.getpid()}")
        with open(tmp_path, "w") as f:
            json.dump({"token": token, "context": aura_context, "cookies": cookies_dict,
                       "headers": aura_headers, "scraper_list_id": scraper_list_id}, f)
        tmp_path.replace(AURA_BOOTSTRAP_FILE)
        # This file holds a live SF session cookie + aura token — treat it as the
        # secret it is: owner-only (I19/I23).
        try:
            os.chmod(AURA_BOOTSTRAP_FILE, 0o600)
        except Exception:
            pass
    else:
        print(f"  Skipping aura bootstrap save — final page is not the "
              f"authenticated app (url={final_url[:60]}); keeping existing session.")

    fwuid = aura_context.get("fwuid", "?")
    print(f"  Aura bootstrap complete (fwuid={fwuid[:20]}...)")
    return cookies_dict, token, aura_context, aura_headers


def load_bootstrap():
    """Read cached bootstrap. Returns (cookies, token, context, headers) or None."""
    if not AURA_BOOTSTRAP_FILE.exists():
        return None
    try:
        with open(AURA_BOOTSTRAP_FILE) as f:
            b = json.load(f)
        return b["cookies"], b["token"], b["context"], b.get("headers", {})
    except Exception:
        return None


def load_scraper_list_id():
    """Return the territory-OFF scraper list ID saved during bootstrap, or None."""
    if not AURA_BOOTSTRAP_FILE.exists():
        return None
    try:
        with open(AURA_BOOTSTRAP_FILE) as f:
            return json.load(f).get("scraper_list_id")
    except Exception:
        return None


def _aura_post(cookies, aura_token, aura_context, method, params, aura_headers=None):
    """POST one Aura action and return the parsed response."""
    aura_headers = aura_headers or {}
    message = {
        "actions": [{
            "id": "1;a",
            "descriptor": "aura://ApexActionController/ACTION$execute",
            "callingDescriptor": "UNKNOWN",
            "params": {
                "namespace": "",
                "classname": "TerritoryProspectingController",
                "method": method,
                "params": params,
                "cacheable": False,
                "isContinuation": False,
            }
        }]
    }

    body = urllib.parse.urlencode({
        "message": json.dumps(message),
        "aura.context": json.dumps(aura_context),
        "aura.pageURI": AURA_PAGE_URI,
        "aura.token": aura_token or "",
    }).encode()

    cookie_header = "; ".join(f"{k}={v}" for k, v in cookies.items())
    req = urllib.request.Request(
        AURA_URL,
        data=body,
        headers={
            "Content-Type": "application/x-www-form-urlencoded;charset=UTF-8",
            "Cookie": cookie_header,
            "User-Agent": aura_headers.get("user-agent", "Mozilla/5.0"),
            "Referer": aura_headers.get("referer", ISC_HOST + AURA_PAGE_URI),
            "Origin": aura_headers.get("origin", ISC_HOST),
            "Accept": aura_headers.get("accept", "*/*"),
            "Accept-Language": aura_headers.get("accept-language", "en-US,en;q=0.9"),
        },
        method="POST",
    )
    # Transient network blips (TLS handshake timeouts, connection resets) are
    # common on long paginated scrapes and otherwise crash the whole worker,
    # forcing a full re-run of every other CovID in the batch too. Retry a
    # few times with backoff before giving up.
    last_err = None
    for attempt in range(4):
        try:
            with urllib.request.urlopen(req, timeout=60) as r:
                return json.loads(r.read().decode())
        except urllib.error.HTTPError:
            raise  # e.g. 401 — handled by the caller's re-bootstrap-and-retry logic
        except (urllib.error.URLError, TimeoutError, ConnectionError) as e:
            last_err = e
            if attempt < 3:
                wait = 2 ** attempt
                print(f"  Network error ({e}) — retrying in {wait}s (attempt {attempt + 2}/4)...")
                time.sleep(wait)
    raise last_err


PAGE_SIZE = 300  # matches the ProspectList's default limitValue


def _extract_rows(response):
    """Pull account rows and total count out of an Aura response.
    Real shape (verified against live traffic): actions[0].returnValue is a
    wrapper with a nested returnValue holding {numberOfAccounts, results}."""
    actions = response.get("actions", [])
    if not actions:
        raise RuntimeError(f"Empty actions in response: {str(response)[:300]}")

    outer = actions[0].get("returnValue")
    if outer is None:
        state = actions[0].get("state", "")
        error = actions[0].get("error", [])
        raise RuntimeError(f"Action state={state} error={error}")

    inner = outer.get("returnValue") or {}
    total = inner.get("numberOfAccounts", 0)
    accounts = inner.get("results", [])
    return accounts, total


def _accounts_to_rows(accounts):
    """Convert account dicts to ordered row lists matching HEADERS_ORDER."""
    col_ids = [col["id"] for col in ACCOUNT_COLUMNS if not col["hidden"]]
    rows = []
    for acct in accounts:
        row = []
        for col_id in col_ids:
            val = acct.get(col_id, "")
            if val is None:
                val = ""
            row.append(str(val) if not isinstance(val, str) else val)
        rows.append(row)
    return rows


def scrape_cov_id(cov_id, prospect_list_id, cookies, aura_token, aura_context, aura_headers=None,
                  fallback_list_ids=None):
    """Fetch all accounts for one CovID via HTTP. Returns list of row lists.

    Cold-cache behaviour (verified against live Salesforce traffic, CONTEXT.md
    Round 4): a covId+listId combination that hasn't been queried recently can
    return 0 rows on the first hit, then correct data on a later retry. The
    cache is per (covId, listId) pair, so the fastest fix is to rotate to a
    different list ID rather than waiting minutes for the same pair to warm.
    Pass `fallback_list_ids` (a list of other IDs to try in order) so that a
    0-result first hit immediately retries on a fresh list instead of sitting
    in a 3-second spin on the same stale pair."""
    print(f"  [{cov_id}] Fetching...")
    all_accounts = []
    page = 0
    active_list_id = prospect_list_id

    while True:
        page += 1
        params = {
            "id": active_list_id,
            "type": "IBM Accounts",
            "accountColumns": ACCOUNT_COLUMNS,
            "filters": [{"id": "covId", "operator": "equals", "values": [cov_id]}],
            "sortOrder": "asc",
            "sortBy": "DOM_COMP_NAME",
            "filterListRemoved": [],
        }
        if page > 1:
            params["limitValue"] = (page - 1) * PAGE_SIZE - 1
            params["preserveCount"] = False
        resp = _aura_post(cookies, aura_token, aura_context, "getAccountPageContents", params, aura_headers)
        accounts, total = _extract_rows(resp)

        # Cold-cache zero: rotate through fallback list IDs first (different
        # listId = fresh cache entry, often returns data immediately), then
        # fall back to short time-based retries on the last available list ID.
        if page == 1 and total == 0 and not accounts:
            fallbacks = list(fallback_list_ids or [])
            attempt = 1
            while total == 0 and not accounts:
                attempt += 1
                if fallbacks:
                    next_id = fallbacks.pop(0)
                    print(f"  [{cov_id}] 0 results — trying different list ID (attempt {attempt})...")
                    active_list_id = next_id
                    params["id"] = active_list_id
                    time.sleep(0.5)
                else:
                    if attempt > 5:
                        break
                    print(f"  [{cov_id}] 0 results — retrying same list (attempt {attempt}/5)...")
                    time.sleep(5)
                resp = _aura_post(cookies, aura_token, aura_context, "getAccountPageContents", params, aura_headers)
                accounts, total = _extract_rows(resp)

        all_accounts.extend(accounts)
        print(f"  [{cov_id}] Page {page}: got {len(accounts)} rows ({len(all_accounts)}/{total})")

        if len(all_accounts) >= total or not accounts:
            break
        time.sleep(0.3)

    print(f"  [{cov_id}] Done — {len(all_accounts)} rows")
    return _accounts_to_rows(all_accounts)


def get_all_prospect_list_ids(cookies, aura_token, aura_context, aura_headers=None):
    """Return every prospecting-list ID visible to this user (empty list if none)."""
    resp = _aura_post(cookies, aura_token, aura_context, "getProspectLists", {}, aura_headers)
    actions = resp.get("actions", [])
    if not actions:
        return []
    rv = actions[0].get("returnValue")
    if not rv:
        return []
    lists = rv if isinstance(rv, list) else rv.get("returnValue", [])
    return [l.get("id") or l.get("prospectListId") for l in lists if l.get("id") or l.get("prospectListId")]


def _get_prospect_list_id(cookies, aura_token, aura_context, aura_headers=None):
    """Get the ID of the first available prospecting list, or raise."""
    ids = get_all_prospect_list_ids(cookies, aura_token, aura_context, aura_headers)
    if ids:
        return ids[0]
    raise RuntimeError("No prospecting lists found. Create one in Territory Prospecting first.")


_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@")


def _sanitize_cell(val):
    """Neutralize CSV/Excel formula injection (CWE-1236). openpyxl auto-marks
    any string starting with '=' as a live formula cell that Excel executes
    on open — and Salesforce account/company names are external data we
    don't control. Prefixing with an apostrophe forces literal-text display,
    matching Excel's own convention for escaping leading special characters."""
    if isinstance(val, str) and val.startswith(_FORMULA_TRIGGER_CHARS):
        return "'" + val
    return val


def save_to_excel(headers, all_rows, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Territory Prospecting"

    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True, size=11)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for ri, row in enumerate(all_rows, 2):
        for ci, val in enumerate(row, 1):
            if ci <= len(headers):
                ws.cell(row=ri, column=ci, value=_sanitize_cell(val))

    for ci in range(1, len(headers) + 1):
        max_len = len(str(ws.cell(row=1, column=ci).value or ""))
        for ri in range(2, min(52, len(all_rows) + 2)):
            cl = len(str(ws.cell(row=ri, column=ci).value or ""))
            if cl > max_len:
                max_len = cl
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max_len + 4, 40)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"  Saved {len(all_rows)} rows → {output_path}")


def run(args):
    auth_state = _load_session()

    if args.bootstrap_only:
        # Just warm the aura token cache — no scrape, no output file. Used by
        # launcher.py as a lead-in step before dispatching parallel workers,
        # so it must NOT also fetch the same CovID's data: an identical
        # getAccountPageContents call fired again moments later (by the
        # first real worker) has been observed to come back with 0 rows,
        # apparently a server-side collision between two near-simultaneous
        # identical queries for the same list+filter.
        bootstrap_aura(auth_state)
        return

    if args.list_prospect_ids:
        # Print every prospect-list ID this user can see, one per line, so
        # launcher.py can assign a distinct list to each concurrent worker —
        # concurrent requests that share one list ID race and clobber each
        # other's filter server-side (verified against live traffic).
        cached = load_bootstrap()
        cookies, aura_token, aura_context, aura_headers = cached if cached else bootstrap_aura(auth_state)
        try:
            ids = get_all_prospect_list_ids(cookies, aura_token, aura_context, aura_headers)
        except urllib.error.HTTPError as e:
            # A cached aura_bootstrap.json can be stale (launcher.py's
            # _ensure_bootstrap only checks that the file exists, not that
            # the token inside it still works) — every other call path in
            # this file already retries once with a forced re-bootstrap on
            # 401; this one didn't, so a stale token silently crashed this
            # subprocess and launcher.py reported a misleading "No prospect
            # lists available" with the real cause hidden in stderr.
            if e.code != 401:
                raise
            print("  Got 401 with cached session — re-bootstrapping...", file=sys.stderr)
            cookies, aura_token, aura_context, aura_headers = bootstrap_aura(auth_state)
            ids = get_all_prospect_list_ids(cookies, aura_token, aura_context, aura_headers)
        for pid in ids:
            print(pid)
        return

    if getattr(sys, "frozen", False):
        app_root = Path(sys.executable).parent
    else:
        app_root = Path(__file__).parent.parent

    output_dir = app_root / "output"
    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / Path(args.output).name

    cov_ids = json.loads(args.cov_ids_json) if args.cov_ids_json else [args.cov_id]

    print("=" * 60)
    print("  ISC HTTP Scraper")
    print(f"  CovIDs: {cov_ids}")
    print(f"  Output: {output_path}")
    print("=" * 60)

    # Try cached bootstrap first; fall back to headless browser fetch
    cached = load_bootstrap()
    if cached and not args.bootstrap:
        cookies, aura_token, aura_context, aura_headers = cached
        print("  Loaded cached aura session")
    else:
        cookies, aura_token, aura_context, aura_headers = bootstrap_aura(auth_state)

    prospect_list_id = args.prospect_list_id
    if not prospect_list_id:
        print("  Fetching prospecting list ID...")
        try:
            prospect_list_id = _get_prospect_list_id(cookies, aura_token, aura_context, aura_headers)
        except urllib.error.HTTPError as e:
            if e.code != 401:
                raise
            print("  Got 401 with cached session — re-bootstrapping...")
            cookies, aura_token, aura_context, aura_headers = bootstrap_aura(auth_state)
            prospect_list_id = _get_prospect_list_id(cookies, aura_token, aura_context, aura_headers)
    print(f"  Prospecting list: {prospect_list_id}")

    all_rows = []
    for cov_id in cov_ids:
        try:
            rows = scrape_cov_id(cov_id, prospect_list_id, cookies, aura_token, aura_context, aura_headers)
        except urllib.error.HTTPError as e:
            if e.code != 401:
                raise
            print("  Got 401 with cached session — re-bootstrapping...")
            cookies, aura_token, aura_context, aura_headers = bootstrap_aura(auth_state)
            rows = scrape_cov_id(cov_id, prospect_list_id, cookies, aura_token, aura_context, aura_headers)
        all_rows.extend(rows)

    save_to_excel(HEADERS_ORDER, all_rows, str(output_path))

    if not args.no_dedup:
        dedup_out = str(output_path).replace(".xlsx", "_deduped.xlsx")
        try:
            import importlib.util
            spec = importlib.util.spec_from_file_location("dedup", Path(__file__).parent / "dedup.py")
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            mod.run_dedup(str(output_path), dedup_out)
        except Exception as e:
            print(f"  Dedup failed: {e}")


def main():
    parser = argparse.ArgumentParser(description="ISC HTTP Scraper (no browser)")
    parser.add_argument("--cov-id", default="T0016156")
    parser.add_argument("--cov-ids-json", default=None)
    parser.add_argument("--output", default="territory_prospecting_export.xlsx")
    parser.add_argument("--prospect-list-id", default=None, help="Reuse an existing prospecting list ID")
    parser.add_argument("--fallback-list-ids", default=None,
                        help="Comma-separated list IDs to try on cold-cache 0-result (rotate before time-waits)")
    parser.add_argument("--no-dedup", action="store_true")
    parser.add_argument("--bootstrap", action="store_true", help="Force re-fetch aura token via headless browser")
    parser.add_argument("--bootstrap-only", action="store_true", help="Only warm the aura token cache, skip scraping/output entirely")
    parser.add_argument("--list-prospect-ids", action="store_true", help="Print all visible prospect list IDs (one per line) and exit")
    args = parser.parse_args()
    run(args)


if __name__ == "__main__":
    main()
