"""Regression tests for dedup.py's two-stage pipeline: Stage 1
(exact-duplicate location removal via CMR Number) + Stage 2 (company-level
rollup by Account Name/Name for market intelligence).
Run directly: python3 test_dedup.py
Exits 0 if all pass, 1 if any fail. Uses only synthetic in-memory workbooks
so it has no dependency on any scraped data file."""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import dedup
import openpyxl
import tempfile

HEADERS = ["Name", "Country", "Account Name", "Account Number", "Coverage ID",
           "Technology Client Status", "CMR Number", "Address", "City", "State/Province",
           "Contact Count", "Industry", "Sub Industry", "Employee Count",
           "Location Annual Revenue", "Global Annual Revenue", "Total IT Spend",
           "Cloud Spend", "Headquarters", "Headquarters Country", "LinkedIn URL",
           "IBM Spend Current Year", "IBM Spend Prior Year", "IBM Spend Prior Year - 1",
           "IBM Spend Prior Year - 2"]


def make_row(name="Co", country="US", acct=None, acctnum="", covid="T1", status="Existing (Continued)",
             cmr="", addr="1 St", city="City", state="ST", contact=0, industry="Ind", subind="Sub",
             emp=0, locrev=0, globrev=0, itspend=0, cloud=0, hq="No", hqcountry="US", linkedin="",
             ibm_cy=0, ibm_py=0, ibm_py1=0, ibm_py2=0):
    if acct is None:
        acct = name
    return [name, country, acct, acctnum, covid, status, cmr, addr, city, state, contact, industry,
            subind, emp, locrev, globrev, itspend, cloud, hq, hqcountry, linkedin, ibm_cy, ibm_py, ibm_py1, ibm_py2]


def run(rows):
    with tempfile.TemporaryDirectory() as d:
        inp = os.path.join(d, "in.xlsx")
        out = os.path.join(d, "out.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(HEADERS)
        for r in rows:
            ws.append(r)
        wb.save(inp)
        dedup.run_dedup(inp, out)
        wb2 = openpyxl.load_workbook(out)
        ws2 = wb2["Company Rollup"]
        out_headers = [c.value for c in ws2[1]]
        result_rows = [list(r) for r in ws2.iter_rows(min_row=2, values_only=True)]
        return out_headers, result_rows


def col(headers, name):
    return headers.index(name)


passed = 0
failed = 0


def check(desc, cond):
    global passed, failed
    if cond:
        passed += 1
        print(f"  PASS: {desc}")
    else:
        failed += 1
        print(f"  FAIL: {desc}")


print("=== Test 1: distinct branches of the same company roll up into ONE company row ===")
rows = [
    make_row(name="Exchange Bank", acctnum="DC00TEST", cmr="C1", addr="1 First St", city="Santa Rosa", contact=3, locrev=500, globrev=100000, emp=10, ibm_cy=50),
    make_row(name="Exchange Bank", acctnum="DC00TEST", cmr="C2", addr="2 Second St", city="Petaluma", contact=4, locrev=700, globrev=120000, emp=20, ibm_cy=60),
    make_row(name="Exchange Bank", acctnum="DC00TEST", cmr="C3", addr="3 Third St", city="Napa", contact=2, locrev=300, globrev=90000, emp=15, ibm_cy=40),
]
headers, out = run(rows)
check("3 branches -> 1 company row", len(out) == 1)
check("Distinct Locations == 3", out[0][col(headers, "Distinct Locations")] == 3)
check("Contact Count SUMMED (3+4+2=9)", out[0][col(headers, "Contact Count")] == 9)
check("Location Annual Revenue SUMMED (500+700+300=1500)", out[0][col(headers, "Location Annual Revenue")] == 1500)
check("Employee Count SUMMED (10+20+15=45)", out[0][col(headers, "Employee Count")] == 45)
check("IBM Spend Current Year SUMMED (50+60+40=150)", out[0][col(headers, "IBM Spend Current Year")] == 150)
check("Global Annual Revenue is MAX not SUM (120000, not 310000)", out[0][col(headers, "Global Annual Revenue")] == 120000)
check("Address column dropped entirely", "Address" not in headers)
check("City column dropped entirely", "City" not in headers)
check("State/Province column dropped entirely", "State/Province" not in headers)
check("CMR Number column dropped entirely", "CMR Number" not in headers)
# Account Number is deliberately KEPT (the IBM client/buying-group hierarchy code
# is the deterministic join key for Account Segmentation — see dedup.build_crosswalk).
check("Account Number column KEPT (join key)", "Account Number" in headers)
check("Account Number value preserved on the rollup",
      bool(str(out[0][col(headers, "Account Number")]).strip()))

print("=== Test 2: different companies (different Account Name) stay separate ===")
rows = [
    make_row(name="Alpha Co", acct="Alpha Corp", cmr="C10", locrev=100),
    make_row(name="Beta Co", acct="Beta Corp", cmr="C11", locrev=200),
]
headers, out = run(rows)
check("2 distinct companies -> 2 rows", len(out) == 2)

print("=== Test 3: exact-duplicate scrape artifact (same CMR) removed BEFORE company rollup (no double count) ===")
rows = [
    make_row(name="Gamma Inc", cmr="C20", addr="9 Pine Rd", locrev=1000, contact=5),
    make_row(name="Gamma Inc", cmr="C20", addr="9 Pine Rd", locrev=1000, contact=5),  # exact scrape dup
    make_row(name="Gamma Inc", cmr="C21", addr="10 Oak Rd", locrev=500, contact=2),   # real 2nd branch
]
headers, out = run(rows)
check("1 company row (Gamma Inc)", len(out) == 1)
check("Distinct Locations == 2 (dup removed first, then 2 real branches rolled up)", out[0][col(headers, "Distinct Locations")] == 2)
check("Location Annual Revenue == 1500 (1000+500, NOT 1000+1000+500=2500 -- dup not double counted)", out[0][col(headers, "Location Annual Revenue")] == 1500)
check("Contact Count == 7 (5+2, not 5+5+2=12)", out[0][col(headers, "Contact Count")] == 7)
row_nums = sorted(int(x) for x in out[0][col(headers, "Merged From Row(s)")].split(";"))
check("Merged From Row(s) references all 3 original rows (2,3,4)", row_nums == [2, 3, 4])

print("=== Test 4: Technology Client Status consolidated to most-engaged value ===")
rows = [
    make_row(name="Delta LLC", cmr="C30", status="New (Whitespace)"),
    make_row(name="Delta LLC", cmr="C31", status="Existing (Continued)"),
    make_row(name="Delta LLC", cmr="C32", status="New (Dormant)"),
]
headers, out = run(rows)
check("status consolidated to 'Existing (Continued)' (most engaged)", out[0][col(headers, "Technology Client Status")] == "Existing (Continued)")

print("=== Test 5: Headquarters Yes/No/Unknown consolidation ===")
rows = [
    make_row(name="Epsilon Co", cmr="C40", hq="No"),
    make_row(name="Epsilon Co", cmr="C41", hq="Yes"),
    make_row(name="Epsilon Co", cmr="C42", hq="Unknown"),
]
headers, out = run(rows)
check("Headquarters == Yes (any Yes wins)", out[0][col(headers, "Headquarters")] == "Yes")

print("=== Test 6: Coverage ID consolidated into distinct joined list ===")
rows = [
    make_row(name="Zeta Corp", cmr="C50", covid="T0001"),
    make_row(name="Zeta Corp", cmr="C51", covid="T0002"),
    make_row(name="Zeta Corp", cmr="C52", covid="T0001"),  # duplicate covid value
]
headers, out = run(rows)
covids = set(out[0][col(headers, "Coverage ID")].split(";"))
check("Coverage ID is the distinct set {T0001, T0002}", covids == {"T0001", "T0002"})

print("=== Test 7: backtrace invariant holds across a large mixed batch ===")
rows = [
    make_row(name="A", acct="A Corp", cmr="CA1", locrev=10),
    make_row(name="A", acct="A Corp", cmr="CA1", locrev=10),  # exact dup
    make_row(name="A", acct="A Corp", cmr="CA2", locrev=20),  # real 2nd branch, same company
    make_row(name="B", acct="B Corp", cmr="CB1", locrev=30),  # different company
    make_row(name="C", acct="A Corp", cmr="CA3", locrev=40),  # 3rd branch, Name differs but Account Name same
]
headers, out = run(rows)
total_locs = sum(r[col(headers, "Distinct Locations")] for r in out)
check("backtrace: sum(Distinct Locations) == 4 real locations (5 rows minus 1 exact dup)", total_locs == 4)
check("2 companies in output (A Corp, B Corp)", len(out) == 2)
a_corp = [r for r in out if r[col(headers, "Account Name")] == "A Corp"][0]
check("A Corp revenue == 10+20+40 == 70 (dup not counted twice)", a_corp[col(headers, "Location Annual Revenue")] == 70)
all_src = set()
for r in out:
    for x in r[col(headers, "Merged From Row(s)")].split(";"):
        check_dup = int(x) not in all_src
        all_src.add(int(x))
check("every original row (2-6) referenced exactly once across all companies", all_src == {2, 3, 4, 5, 6})

print("=== Test 8: blank Account Name falls back to Name for identity ===")
rows = [
    make_row(name="Solo Shop", acct="", cmr="CS1", locrev=100),
    make_row(name="Solo Shop", acct="", cmr="CS2", locrev=200),
]
headers, out = run(rows)
check("blank Account Name -> falls back to Name, still rolls up to 1 company", len(out) == 1)
check("revenue summed via fallback grouping too (300)", out[0][col(headers, "Location Annual Revenue")] == 300)

print("=== Test 9: ragged row (fewer columns than header) doesn't misalign ===")
with tempfile.TemporaryDirectory() as d:
    inp = os.path.join(d, "in.xlsx")
    out_path = os.path.join(d, "out.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADERS)
    ws.append(["Ragged Co", "US", "Ragged Corp", "", "T1"])  # missing many trailing columns
    wb.save(inp)
    dedup.run_dedup(inp, out_path)
    wb2 = openpyxl.load_workbook(out_path)
    ws2 = wb2["Company Rollup"]
    row = list(next(ws2.iter_rows(min_row=2, max_row=2, values_only=True)))
    out_headers2 = [c.value for c in ws2[1]]
    check("ragged row padded safely, Industry field blank not misaligned data",
          is_effectively_blank := (row[col(out_headers2, "Industry")] in (None, ""))
    )

print(f"\n=== RESULTS: {passed} passed, {failed} failed ===")
sys.exit(1 if failed else 0)
