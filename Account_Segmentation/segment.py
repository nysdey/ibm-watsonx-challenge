"""Core Segmentation logic: fuzzy-join the five IBM install files onto
DEDUPED_ACCOUNTS, then sort by install-type coverage.

Output SEGMENTED_ACCOUNTS layout, left to right:
  [ base DEDUPED_ACCOUNTS columns ]
  Install_Types_Count, Install_Types
  per type (Cloud, Power, Storage, NonInfra, Competitive):
     <Label>_Present, <Label>_Rows, <Label>_Match_Score, <Label>_Matched_Name, <Label>_Match_Quality
  (if ATTACH_INSTALL_COLUMNS) per type: "<Label>: <every column of that file>" (aggregated)

Sort: Install_Types_Count desc, then presence in priority order
(Cloud->Power->Storage->NonInfra->Competitive) desc, then Account Name asc.
"""
import json
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import config
import id_match
import name_match

log = logging.getLogger("segmentation")

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def _load_sheet(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = list(next(it))
    except StopIteration:
        return [], []
    data = [list(r) for r in it]
    wb.close()
    return header, data


def _detect_name_column(headers, override):
    if override:
        for i, h in enumerate(headers):
            if h and str(h).strip().lower() == override.strip().lower():
                return i
        log.warning("override name column %r not found; auto-detecting", override)
    lowered = [(i, str(h).strip().lower()) for i, h in enumerate(headers) if h]
    for pref in config.NAME_COLUMN_PREFERENCE:
        for i, h in lowered:
            if pref in h:
                return i
    # Fallback: first column.
    return 0


def _num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "")
    try:
        return float(s)
    except ValueError:
        return None


def _aggregate(indices, data, n_cols):
    """Aggregate the matched install rows into one value per column (aligned to
    the install file's columns): numeric columns summed, text columns joined
    (distinct, capped)."""
    out = []
    for c in range(n_cols):
        vals = [data[i][c] for i in indices if c < len(data[i])]
        nonblank = [v for v in vals if v not in (None, "")]
        if not nonblank:
            out.append("")
            continue
        nums, all_num = [], True
        for v in nonblank:
            n = _num(v)
            if n is None:
                all_num = False
                break
            nums.append(n)
        if all_num and nums:
            total = sum(nums)
            out.append(int(total) if float(total).is_integer() else total)
        else:
            distinct = []
            for v in nonblank:
                sv = str(v).strip()
                if sv not in distinct:
                    distinct.append(sv)
            joined = "; ".join(distinct)
            if len(joined) > config.MAX_JOINED_TEXT_LEN:
                joined = joined[: config.MAX_JOINED_TEXT_LEN - 1] + "…"
            out.append(joined)
    return out


def _load_install_type(key, label, path):
    """Return a dict describing one install type, or None if its file is absent.
    The matcher (deterministic id-join, or fuzzy fallback) is attached later, once
    the base-account resolver is built."""
    if not path.exists():
        log.info("[%s] file not found (%s) — treated as no installs for every account", key, path.name)
        return None
    headers, data = _load_sheet(path)
    if not headers:
        log.info("[%s] file empty", key)
        return None
    override = config.NAME_COLUMN_OVERRIDES.get(key)
    name_idx = _detect_name_column(headers, override)
    return {
        "key": key, "label": label, "headers": headers, "data": data,
        "name_idx": name_idx, "spec": config.ID_COLUMNS.get(key),
    }


def _attach_matcher(info, resolver):
    """Choose the deterministic id-join when possible, else the fuzzy fallback."""
    key = info["key"]
    if resolver is not None and info.get("spec"):
        idx = id_match.InstallKeyIndex(info["headers"], info["data"], info["spec"], resolver)
        info["mode"] = "id"
        info["matcher"] = idx
        log.info("[%s] id-join: %d/%d rows resolved to an account "
                 "(by %s); %d unmatched (out-of-territory)",
                 key, idx.matched_rows, len(info["data"]),
                 ", ".join(f"{b}={n}" for b, n in idx.basis_counts.items() if n), idx.unmatched)
    else:
        nidx = name_match.NameIndex()
        for r, row in enumerate(info["data"]):
            nm = row[info["name_idx"]] if info["name_idx"] < len(row) else None
            nidx.add(nm, r)
        info["mode"] = "fuzzy"
        info["matcher"] = nidx
        log.info("[%s] no id key available -> fuzzy name fallback (%d distinct names)",
                 key, len(nidx.unique_norms))


def _match_account(info, base_key, base_name):
    """Return (indices, score, quality, matched_name) for one base account against
    one install type. Deterministic id-join when info['mode']=='id'."""
    if info["mode"] == "id":
        indices = info["matcher"].rows_for(base_key) if base_key else []
        if not indices:
            return [], "", "", ""
        basis = info["matcher"].row_basis.get(indices[0], "id")
        ni = info["name_idx"]
        matched_name = str(info["data"][indices[0]][ni] or "") if ni is not None and ni < len(info["data"][indices[0]]) else ""
        return indices, 1.0, basis, matched_name
    # fuzzy fallback
    _, score, quality, indices = info["matcher"].match(base_name)
    if quality not in ("matched", "review"):
        return [], score, "", ""
    ni = info["name_idx"]
    matched_name = str(info["data"][indices[0]][ni] or "") if indices and ni is not None else ""
    return indices, round(score, 3), quality, matched_name


def run_segmentation():
    # --- base ---
    base_headers, base_data = _load_sheet(config.DEDUPED_ACCOUNTS_PATH)
    if not base_headers:
        raise RuntimeError(f"DEDUPED_ACCOUNTS is empty/missing: {config.DEDUPED_ACCOUNTS_PATH}")
    try:
        base_name_idx = base_headers.index(config.BASE_NAME_COLUMN)
    except ValueError:
        raise RuntimeError(
            f"Base name column {config.BASE_NAME_COLUMN!r} not in DEDUPED_ACCOUNTS "
            f"(has: {base_headers})"
        )
    log.info("base DEDUPED_ACCOUNTS: %d accounts, %d cols", len(base_data), len(base_headers))

    # --- deterministic account-key resolver (id-join) ---
    base_key_idx = None
    for i, h in enumerate(base_headers):
        if h and str(h).strip().lower() == config.BASE_ACCOUNT_KEY_COLUMN.strip().lower():
            base_key_idx = i
            break
    resolver = None
    if base_key_idx is not None:
        crosswalk = None
        if config.ACCOUNT_CROSSWALK_PATH.exists():
            try:
                crosswalk = json.loads(config.ACCOUNT_CROSSWALK_PATH.read_text())
            except Exception as e:
                log.warning("could not read account crosswalk (%s): %s", config.ACCOUNT_CROSSWALK_PATH.name, e)
        resolver = id_match.build_resolver(base_headers, base_data, base_key_idx, base_name_idx, crosswalk)
        log.info("id-join enabled: %d account keys, %d customer-number links, %d exact-name keys",
                 len(resolver.base_keys), len(resolver.cust_to_key), len(resolver.name_to_key))
    else:
        log.warning("base has no %r column -> falling back to fuzzy name matching for all types "
                    "(re-run the ISC step so DEDUPED carries the account key)", config.BASE_ACCOUNT_KEY_COLUMN)

    # --- load the five install types (in priority order) ---
    loaded = []
    for key, label, path in config.INSTALL_TYPES:
        info = _load_install_type(key, label, path)
        if info is not None:
            _attach_matcher(info, resolver)
        loaded.append((key, label, info))

    # --- build output header ---
    out_headers = list(base_headers)
    out_headers += ["Install_Types_Count", "Install_Types"]
    for _, label, _info in loaded:
        out_headers += [f"{label}_Present", f"{label}_Rows",
                        f"{label}_Match_Score", f"{label}_Matched_Name", f"{label}_Match_Basis"]
    attach_specs = []  # (start_col, info) for aggregated attach columns
    if config.ATTACH_INSTALL_COLUMNS:
        for key, label, info in loaded:
            if info is None:
                attach_specs.append((None, key, label, None))
                continue
            attach_specs.append((len(out_headers), key, label, info))
            out_headers += [f"{label}: {h}" for h in info["headers"]]

    # --- per-account join ---
    out_rows = []
    review_flags = 0
    for brow in base_data:
        base_name = brow[base_name_idx] if base_name_idx < len(brow) else None
        base_key = (str(brow[base_key_idx]).strip()
                    if base_key_idx is not None and base_key_idx < len(brow)
                    and brow[base_key_idx] not in (None, "") else "")
        row = list(brow) + [None] * (len(base_headers) - len(brow))

        presence = []          # 1/0 per type in priority order
        compact = []           # 5 compact cells per type
        attach_cells = {}      # label -> aggregated list

        for key, label, info in loaded:
            if info is None:
                presence.append(0)
                compact += ["", 0, "", "", ""]
                attach_cells[label] = [""] * 0
                continue
            indices, score, quality, matched_name = _match_account(info, base_key, base_name)
            present = 1 if indices else 0
            presence.append(present)
            if quality == "review":
                review_flags += 1
            compact += [
                "Yes" if present else "",
                len(indices) if present else 0,
                score if present else "",
                matched_name,
                quality if present else "",
            ]
            if config.ATTACH_INSTALL_COLUMNS:
                if present and indices:
                    attach_cells[label] = _aggregate(indices, info["data"], len(info["headers"]))
                else:
                    attach_cells[label] = [""] * len(info["headers"])

        count = sum(presence)
        type_labels = [label for (key, label, info), p in zip(loaded, presence) if p]
        row += [count, ", ".join(type_labels)]
        row += compact
        if config.ATTACH_INSTALL_COLUMNS:
            for key, label, info in loaded:
                if info is not None:
                    row += attach_cells[label]

        # sort key: count desc, presence priority tuple desc, name asc
        sort_key = (-count, tuple(-p for p in presence), name_match.normalize(base_name))
        out_rows.append((sort_key, row))

    out_rows.sort(key=lambda t: t[0])
    log.info("joined %d accounts; %d borderline (review) matches flagged", len(out_rows), review_flags)

    # --- write ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Segmented Accounts"
    ws.append(out_headers)
    for _, row in out_rows:
        ws.append(row)
    for c in range(1, len(out_headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=2, column=1).coordinate
    ws.auto_filter.ref = ws.dimensions

    dated, latest = config.dated_and_latest_paths()
    wb.save(dated)
    wb.save(latest)

    # coverage summary
    dist = {}
    for (sk, _), in [(r,) for r in out_rows]:
        dist[-sk[0]] = dist.get(-sk[0], 0) + 1
    log.info("install-type coverage distribution (count -> #accounts): %s",
             {k: dist[k] for k in sorted(dist, reverse=True)})
    log.info("wrote SEGMENTED_ACCOUNTS -> %s (%d cols)", latest, len(out_headers))

    return {
        "accounts": len(out_rows),
        "columns": len(out_headers),
        "review_matches": review_flags,
        "coverage_distribution": {k: dist[k] for k in sorted(dist, reverse=True)},
        "dated_path": str(dated),
        "latest_path": str(latest),
    }
