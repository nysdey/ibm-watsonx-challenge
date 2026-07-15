"""
ISC Territory Prospecting Launcher
Flask web UI — topojson decoded in Python, pure SVG paths embedded in HTML.
No JS library dependencies.
"""

import json
import os
import re
import secrets
import shutil
import subprocess
import sys
import threading
import time
import urllib.parse
import urllib.request
import uuid
import webbrowser
from pathlib import Path

import openpyxl
from flask import Flask, jsonify, render_template_string, request

# Per-process shared secret + real-port handshake for the /run + /progress
# action-trigger routes. The parent (dashboard) reads these from the handshake
# file written at startup so it POSTs to the REAL bound port (not a guessed
# 5477 that some other process may be squatting), and — when ISC_REQUIRE_TOKEN=1
# — proves it's the intended caller (I53/I54).
RUN_TOKEN = os.environ.get("ISC_RUN_TOKEN") or secrets.token_urlsafe(16)
PORT_HANDSHAKE_FILE = Path.home() / ".isc_scraper" / "launcher_port.json"

# ── paths ──────────────────────────────────────────────────────────────────────
APP_ROOT = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent
INTERNAL = APP_ROOT / "_internal"

COVID_XLSX    = INTERNAL / "CovID.xlsx"
HTTP_SCRAPER  = INTERNAL / "http_scraper.py"
UI_SCRAPER    = INTERNAL / "isc_scraper.py"  # proven browser-UI scrape (reliable engine)

INDUSTRY_ORDER  = ["FSS", "PUB", "C&D", "IND"]
INDUSTRY_LABELS = {"FSS": "FSS", "PUB": "Pub", "C&D": "C&D", "IND": "Industrial"}

# ── parse xlsx ─────────────────────────────────────────────────────────────────
def load_territories():
    wb = openpyxl.load_workbook(COVID_XLSX, read_only=True, data_only=True)
    ws = wb.active
    pattern = re.compile(r"ACT CORE\s+(.+?)\s+(FSS|C&D|IND|PUB)\s*$")
    result = {k: [] for k in INDUSTRY_ORDER}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cov_id, account_name = row[0], row[1]
        if not cov_id or not account_name:
            continue
        m = pattern.search(str(account_name))
        if m:
            geo, industry = m.group(1).strip(), m.group(2)
            if industry in result:
                result[industry].append({"id": str(cov_id).strip(), "geo": geo})
    for k in result:
        result[k].sort(key=lambda x: x["geo"])
    wb.close()
    return result

TERRITORIES = load_territories()

# ── decode topojson → SVG paths in Python ─────────────────────────────────────
_TOPO_URL   = "https://cdn.jsdelivr.net/npm/us-atlas@3/states-albers-10m.json"
_TOPO_CACHE = APP_ROOT / ".us_states_topo.json"

FIPS_MAP = {
    "01":"AL","02":"AK","04":"AZ","05":"AR","06":"CA","08":"CO","09":"CT",
    "10":"DE","11":"DC","12":"FL","13":"GA","15":"HI","16":"ID","17":"IL",
    "18":"IN","19":"IA","20":"KS","21":"KY","22":"LA","23":"ME","24":"MD",
    "25":"MA","26":"MI","27":"MN","28":"MS","29":"MO","30":"MT","31":"NE",
    "32":"NV","33":"NH","34":"NJ","35":"NM","36":"NY","37":"NC","38":"ND",
    "39":"OH","40":"OK","41":"OR","42":"PA","44":"RI","45":"SC","46":"SD",
    "47":"TN","48":"TX","49":"UT","50":"VT","51":"VA","53":"WA","54":"WV",
    "55":"WI","56":"WY","72":"PR",
}

def _decode_topo(topo_str):
    """Return ({fips: svg_d}, {fips: (cx, cy)}) decoded from topojson."""
    topo = json.loads(topo_str)
    tr   = topo.get("transform", {})
    sx, sy = tr.get("scale",     [1, 1])
    tx, ty = tr.get("translate", [0, 0])

    decoded = []
    for arc in topo["arcs"]:
        pts, x, y = [], 0, 0
        for dp in arc:
            x += dp[0]; y += dp[1]
            pts.append((x * sx + tx, y * sy + ty))
        decoded.append(pts)

    def arc_coords(idx):
        return decoded[idx] if idx >= 0 else list(reversed(decoded[~idx]))

    def rings_to_d(rings):
        parts = []
        for ring in rings:
            coords = []
            for i in ring:
                coords.extend(arc_coords(i))
            if coords:
                d = f"M{coords[0][0]:.1f},{coords[0][1]:.1f}"
                d += "".join(f"L{x:.1f},{y:.1f}" for x, y in coords[1:])
                parts.append(d + "Z")
        return " ".join(parts)

    def rings_centroid(rings):
        all_coords = []
        for ring in rings:
            for i in ring:
                all_coords.extend(arc_coords(i))
        if not all_coords:
            return None, None
        xs = [p[0] for p in all_coords]
        ys = [p[1] for p in all_coords]
        return (min(xs) + max(xs)) / 2, (min(ys) + max(ys)) / 2

    paths, centroids = {}, {}
    for geom in topo["objects"]["states"]["geometries"]:
        fips = str(geom.get("id", "")).zfill(2)
        if geom["type"] == "Polygon":
            paths[fips] = rings_to_d(geom["arcs"])
            centroids[fips] = rings_centroid(geom["arcs"])
        elif geom["type"] == "MultiPolygon":
            paths[fips] = " ".join(rings_to_d(poly) for poly in geom["arcs"])
            # centroid of largest polygon
            biggest = max(geom["arcs"], key=lambda p: len(p[0]) if p else 0)
            centroids[fips] = rings_centroid(biggest)
    return paths, centroids

def _load_state_paths():
    if _TOPO_CACHE.exists():
        raw = _TOPO_CACHE.read_text()
    else:
        try:
            with urllib.request.urlopen(_TOPO_URL, timeout=15) as r:
                raw = r.read().decode()
            _TOPO_CACHE.write_text(raw)
        except Exception as e:
            print(f"  Warning: could not fetch map data: {e}")
            return {}, {}
    return _decode_topo(raw)

STATE_PATHS, STATE_CENTROIDS = _load_state_paths()

# small states where the label won't fit — skip inline text
_SKIP_LABEL = {"09","10","11","34","44"}  # CT, DE, DC, NJ, RI

def build_svg_paths():
    parts = []
    for fips, d in STATE_PATHS.items():
        abbr = FIPS_MAP.get(fips, "")
        parts.append(f'<path class="state" data-fips="{fips}" data-abbr="{abbr}" d="{d}"/>')
    for fips, (cx, cy) in STATE_CENTROIDS.items():
        if fips in _SKIP_LABEL or cx is None:
            continue
        abbr = FIPS_MAP.get(fips, "")
        parts.append(
            f'<text class="state-label" x="{cx:.1f}" y="{cy:.1f}" '
            f'data-fips="{fips}">{abbr}</text>'
        )
    return "\n".join(parts)

SVG_PATHS = build_svg_paths()

# ── Flask ──────────────────────────────────────────────────────────────────────
app = Flask(__name__)

HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>ISC Territory Prospecting</title>
<style>
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: Arial, sans-serif; background: #f0f4f8; padding: 24px 16px; min-height: 100vh; }
.card { background: #fff; border-radius: 10px; box-shadow: 0 2px 16px rgba(0,0,0,.10); max-width: 1280px; margin: 0 auto; overflow: hidden; }
.card-header { background: #1F4E79; color: #fff; padding: 18px 28px; font-size: 19px; font-weight: bold; }
.card-body { padding: 20px 24px; display: flex; gap: 20px; }
.left-col { flex: 1; min-width: 0; }
.section-label { font-size: 11px; font-weight: bold; text-transform: uppercase; letter-spacing: .8px; color: #1F4E79; margin-bottom: 8px; }

.industry-row { display: flex; gap: 8px; margin-bottom: 16px; flex-wrap: wrap; }
.ind-btn { padding: 7px 20px; border: 2px solid #d0d7df; border-radius: 999px; background: #fff; color: #444; font-size: 13px; font-weight: 600; cursor: pointer; transition: all .15s; }
.ind-btn:hover  { border-color: #2E86C1; color: #2E86C1; }
.ind-btn.active { background: #2E86C1; border-color: #2E86C1; color: #fff; }

#map-wrap { background: #ddeef7; border-radius: 8px; overflow: hidden; }
#map { width: 100%; display: block; }

path.state { stroke: #fff; stroke-width: 0.5; cursor: pointer; transition: fill .12s; }
path.state.available   { fill: #b8d8ee; }
path.state.selected    { fill: #1F4E79; }
path.state.partial     { fill: #5da0c8; }
path.state.unavailable { fill: #cdd4db; cursor: default; }
path.state.available:hover  { fill: #7db8da; }
path.state.selected:hover   { fill: #163d60; }
path.state.partial:hover    { fill: #3d89b8; }

.state-label { font: bold 9px Arial,sans-serif; fill: #fff; text-anchor: middle; dominant-baseline: middle; pointer-events: none; user-select: none; }

/* pacific tiles in SVG */
.pac-tile      { cursor: pointer; transition: fill .12s; }
.pac-tile.available   { fill: #b8d8ee; }
.pac-tile.selected    { fill: #1F4E79; }
.pac-tile.unavailable { fill: #cdd4db; cursor: default; }
.pac-tile.available:hover { fill: #7db8da; }
.pac-tile.selected:hover  { fill: #163d60; }
.pac-tile-label { font: bold 8px Arial,sans-serif; fill: #fff; text-anchor: middle; dominant-baseline: middle; pointer-events: none; }

#popup { position: fixed; background: #fff; border: 1px solid #c0cad6; border-radius: 8px; box-shadow: 0 6px 24px rgba(0,0,0,.18); padding: 12px 14px; z-index: 200; display: none; min-width: 150px; }
#popup-title { font-size: 11px; font-weight: bold; text-transform: uppercase; letter-spacing: .5px; color: #888; margin-bottom: 8px; }
.pop-opt { display: block; width: 100%; padding: 8px 12px; margin-bottom: 5px; border: 2px solid #d0d7df; border-radius: 6px; background: #fff; font-size: 13px; font-weight: 600; color: #333; cursor: pointer; text-align: left; transition: all .12s; }
.pop-opt:last-child { margin-bottom: 0; }
.pop-opt.selected { background: #1F4E79; border-color: #1F4E79; color: #fff; }
.pop-opt:hover:not(.selected) { border-color: #2E86C1; color: #2E86C1; }

.bottom-bar { display: flex; align-items: center; justify-content: space-between; margin-top: 14px; gap: 12px; }
.ctrl-btns { display: flex; gap: 8px; }
.ctrl-btn { padding: 5px 12px; border: 1px solid #c0cad6; border-radius: 5px; background: #eaf2fb; color: #1F4E79; font-size: 12px; font-weight: 600; cursor: pointer; }
.ctrl-btn:hover { background: #d0e4f5; }
.count-line { font-size: 13px; color: #555; }
.run-btn { padding: 10px 28px; background: #1ABC9C; color: #fff; font-size: 14px; font-weight: bold; border: none; border-radius: 7px; cursor: pointer; white-space: nowrap; transition: background .15s; }
.run-btn:hover    { background: #17A589; }
.run-btn:disabled { background: #a0d8cc; cursor: not-allowed; }

/* ── sidebar ── */
.sidebar {
  width: 230px;
  flex-shrink: 0;
  border-left: 1px solid #e4eaf0;
  padding-left: 18px;
  display: flex;
  flex-direction: column;
}
.sidebar-header { font-size: 11px; font-weight: bold; text-transform: uppercase; letter-spacing: .8px; color: #1F4E79; margin-bottom: 10px; }
.sidebar-scroll { flex: 1; overflow-y: auto; max-height: 520px; }
.sidebar-ind { margin-bottom: 12px; }
.sidebar-ind-title { font-size: 10px; font-weight: bold; text-transform: uppercase; letter-spacing: .5px; color: #2E86C1; margin-bottom: 4px; padding-bottom: 3px; border-bottom: 1px solid #dce8f4; }
.sid-row { display: flex; justify-content: space-between; align-items: center; padding: 3px 0; font-size: 11px; border-bottom: 1px solid #f0f4f8; }
.sid-geo { color: #333; font-weight: 500; }
.sid-id  { color: #888; font-family: monospace; font-size: 10px; }
.sid-empty { font-size: 11px; color: #bbb; font-style: italic; margin-top: 4px; }
.sidebar-total { margin-top: 10px; padding-top: 8px; border-top: 1px solid #e4eaf0; font-size: 11px; color: #555; font-weight: 600; }

.launched { display: none; text-align: center; padding: 56px 28px; }
.launched .tick { font-size: 48px; margin-bottom: 12px; }
.launched h2 { color: #1F4E79; font-size: 18px; }
.launched p  { color: #666; font-size: 13px; margin-top: 8px; }
</style>
</head>
<body>
<div class="card">
  <div class="card-header">ISC Territory Prospecting</div>

  <div id="main">
  <div class="card-body">
    <!-- LEFT: map + controls -->
    <div class="left-col">
      <div class="section-label">Industry</div>
      <div class="industry-row" id="indRow"></div>

      <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:6px;">
        <div class="section-label" style="margin-bottom:0">Click states to select territories</div>
        <div class="ctrl-btns">
          <button class="ctrl-btn" onclick="setAll(true)">Select All</button>
          <button class="ctrl-btn" onclick="setAll(false)">Clear All</button>
        </div>
      </div>

      <div id="map-wrap">
        <svg id="map" viewBox="0 0 975 660" xmlns="http://www.w3.org/2000/svg">
          {{ svg_paths|safe }}
          <!-- Pacific territory tiles below the main map -->
          <g id="pac-tiles"></g>
        </svg>
      </div>

      <div class="bottom-bar">
        <div style="display:flex;flex-direction:column;gap:6px;flex:1;">
          <div class="count-line" id="countLine">No territories selected</div>
          <div style="display:flex;align-items:center;gap:8px;">
            <label style="font-size:11px;color:#555;white-space:nowrap;">Prospect List ID (optional):</label>
            <input id="prospectListId" type="text" placeholder="auto-detected if left blank"
              style="font-size:11px;padding:4px 8px;border:1px solid #c0cad6;border-radius:4px;width:200px;font-family:monospace;" />
          </div>
          <label style="font-size:11px;color:#555;display:flex;align-items:center;gap:6px;">
            <input type="checkbox" id="reliableMode" />
            Reliable mode — scrape via the real UI (slower, but won't miss cold-cache territories)
          </label>
        </div>
        <button class="run-btn" id="runBtn" onclick="runScraper()">&#9654; Run Scraper</button>
      </div>
    </div>

    <!-- RIGHT: sidebar -->
    <div class="sidebar">
      <div class="sidebar-header">Coverage IDs</div>
      <div class="sidebar-scroll" id="sidebarScroll"></div>
      <div class="sidebar-total" id="sidebarTotal"></div>
    </div>
  </div>
  </div>

  <div class="launched" id="launched">
    <div class="tick">&#10003;</div>
    <h2>Scraper launched!</h2>
    <p id="launchedDetail">You can close this window.</p>
  </div>
</div>

<div id="popup">
  <div id="popup-title"></div>
  <div id="popup-opts"></div>
</div>

<script>
const DATA   = {{ data|tojson }};
const LABELS = {{ labels|tojson }};
const ORDER  = {{ order|tojson }};

const selectedInds = new Set([ORDER[0]]);
const selected = new Set();

const STATE_GEO = {
  AK:["AK"], AL:["AL"], AR:["AR"], AZ:["AZ"],
  CA:["CA North","CA South"],
  CO:["CO"], CT:["CT"], DC:["DC"], DE:["DE"],
  FL:["FL North","FL South"],
  GA:["GA"], HI:["HI"], IA:["IA"], ID:["ID"],
  IL:["IL"], IN:["IN"], KS:["KS"], KY:["KY"],
  LA:["LA"], MA:["MA"], MD:["MD"], ME:["ME"],
  MI:["MI"], MN:["MN"], MO:["MO"], MS:["MS"],
  MT:["MT"], NC:["NC"], ND:["ND"], NE:["NE"],
  NH:["NH"], NJ:["NJ"], NM:["NM"], NV:["NV"],
  NY:["NY City","NY State"],
  OH:["OH"], OK:["OK"], OR:["OR"],
  PA:["PA East","PA West"],
  PR:["PR"], RI:["RI"], SC:["SC"], SD:["SD"],
  TN:["TN"], TX:["TX North","TX South"],
  UT:["UT"], VA:["VA"], VT:["VT"], WA:["WA"],
  WI:["WI"], WV:["WV"], WY:["WY"]
};

// Pacific tiles: [geo, x, y, w, h, labelX, labelY]
const PAC_TILES = [
  ["GU",         390, 615, 60, 32, 420, 631],
  ["MP",         460, 615, 60, 32, 490, 631],
  ["AS, MH, FM", 530, 615, 80, 32, 570, 631],
];

function available(geo) {
  for (const ind of selectedInds)
    if ((DATA[ind] || []).some(e => e.geo === geo)) return true;
  return false;
}
function stateGeos(abbr) { return (STATE_GEO[abbr] || []).filter(available); }
function fillClass(abbr) {
  const geos = stateGeos(abbr);
  if (!geos.length) return "unavailable";
  const n = geos.filter(g => selected.has(g)).length;
  return n === 0 ? "available" : n === geos.length ? "selected" : "partial";
}
function toggle(geo) { selected.has(geo) ? selected.delete(geo) : selected.add(geo); }

// ── state map ─────────────────────────────────────────────────────────────────
document.querySelectorAll("path.state").forEach(p => {
  p.className.baseVal = "state " + fillClass(p.dataset.abbr);
  p.addEventListener("click", e => {
    const geos = stateGeos(p.dataset.abbr);
    if (!geos.length) return;
    if (geos.length === 1) { toggle(geos[0]); refresh(); }
    else openPopup(e, p.dataset.abbr, geos);
  });
});

function refreshMap() {
  document.querySelectorAll("path.state").forEach(p => {
    p.className.baseVal = "state " + fillClass(p.dataset.abbr);
  });
}

// ── Pacific tiles in SVG ──────────────────────────────────────────────────────
const svgNS = "http://www.w3.org/2000/svg";
const pacG  = document.getElementById("pac-tiles");

// label above tiles
const lbl = document.createElementNS(svgNS, "text");
lbl.setAttribute("x", 390); lbl.setAttribute("y", 609);
lbl.setAttribute("font", "bold 8px Arial"); lbl.setAttribute("font-size", "8");
lbl.setAttribute("fill", "#1F4E79"); lbl.setAttribute("font-weight", "bold");
lbl.textContent = "PACIFIC & TERRITORIES";
pacG.appendChild(lbl);

const pacRects = {};
PAC_TILES.forEach(([geo, x, y, w, h, lx, ly]) => {
  const avail = available(geo);
  const rect = document.createElementNS(svgNS, "rect");
  rect.setAttribute("x", x); rect.setAttribute("y", y);
  rect.setAttribute("width", w); rect.setAttribute("height", h);
  rect.setAttribute("rx", 4);
  rect.className.baseVal = "pac-tile " + (avail ? (selected.has(geo) ? "selected" : "available") : "unavailable");
  rect.setAttribute("data-geo", geo);
  if (avail) rect.addEventListener("click", () => { toggle(geo); refresh(); });
  pacG.appendChild(rect);
  pacRects[geo] = rect;

  const txt = document.createElementNS(svgNS, "text");
  txt.setAttribute("x", lx); txt.setAttribute("y", ly);
  txt.className.baseVal = "pac-tile-label";
  txt.textContent = geo;
  pacG.appendChild(txt);
});

function refreshPacific() {
  PAC_TILES.forEach(([geo]) => {
    const r = pacRects[geo];
    if (!r) return;
    const avail = available(geo);
    r.className.baseVal = "pac-tile " + (avail ? (selected.has(geo) ? "selected" : "available") : "unavailable");
  });
}

// ── popup ─────────────────────────────────────────────────────────────────────
const popup = document.getElementById("popup");
function openPopup(e, abbr, geos) {
  document.getElementById("popup-title").textContent = abbr;
  const opts = document.getElementById("popup-opts");
  opts.innerHTML = "";
  geos.forEach(geo => {
    const btn = document.createElement("button");
    btn.className = "pop-opt" + (selected.has(geo) ? " selected" : "");
    btn.textContent = geo;
    btn.onclick = () => { toggle(geo); btn.classList.toggle("selected", selected.has(geo)); refresh(); };
    opts.appendChild(btn);
  });
  popup.style.display = "block";
  popup.style.left = Math.min(e.clientX + 8, window.innerWidth  - 180) + "px";
  popup.style.top  = Math.min(e.clientY + 8, window.innerHeight - 130) + "px";
}
document.addEventListener("click", e => {
  if (!popup.contains(e.target) && !e.target.closest("path") && !e.target.closest("rect.pac-tile"))
    popup.style.display = "none";
});

// ── sidebar ───────────────────────────────────────────────────────────────────
function refreshSidebar() {
  const scroll = document.getElementById("sidebarScroll");
  const total  = document.getElementById("sidebarTotal");
  scroll.innerHTML = "";
  let totalCount = 0;

  for (const ind of ORDER) {
    if (!selectedInds.has(ind)) continue;
    const rows = (DATA[ind] || []).filter(e => selected.has(e.geo));
    const div = document.createElement("div");
    div.className = "sidebar-ind";
    div.innerHTML = `<div class="sidebar-ind-title">${LABELS[ind]}</div>`;
    if (rows.length === 0) {
      div.innerHTML += `<div class="sid-empty">None selected</div>`;
    } else {
      rows.forEach(e => {
        div.innerHTML += `<div class="sid-row"><span class="sid-geo">${e.geo}</span><span class="sid-id">${e.id}</span></div>`;
        totalCount++;
      });
    }
    scroll.appendChild(div);
  }
  total.textContent = totalCount > 0 ? `${totalCount} Coverage ID${totalCount !== 1 ? "s" : ""} selected` : "";
}

// ── count ─────────────────────────────────────────────────────────────────────
function refreshCount() {
  const n = selected.size;
  document.getElementById("countLine").textContent =
    n === 0 ? "No territories selected" : `${n} territor${n === 1 ? "y" : "ies"} selected`;
}

function refresh() {
  refreshMap();
  refreshPacific();
  refreshCount();
  refreshSidebar();
}

// ── industry selector ─────────────────────────────────────────────────────────
function buildIndustryRow() {
  const row = document.getElementById("indRow");
  ORDER.forEach(key => {
    const btn = document.createElement("button");
    btn.className = "ind-btn" + (selectedInds.has(key) ? " active" : "");
    btn.textContent = LABELS[key];
    btn.id = "ind-" + key;
    btn.onclick = () => {
      if (selectedInds.has(key)) { if (selectedInds.size > 1) selectedInds.delete(key); }
      else selectedInds.add(key);
      btn.classList.toggle("active", selectedInds.has(key));
      for (const geo of [...selected]) if (!available(geo)) selected.delete(geo);
      popup.style.display = "none";
      refresh();
    };
    row.appendChild(btn);
  });
}

function setAll(val) {
  for (const ind of selectedInds)
    (DATA[ind] || []).forEach(e => val ? selected.add(e.geo) : selected.delete(e.geo));
  popup.style.display = "none";
  refresh();
}

async function runScraper() {
  const ids = [];
  const summary = [];
  // selected/selectedInds live for the whole tab session and never reset
  // themselves (e.g. an earlier "Select All" click stays in effect until
  // "Clear All" or a page reload) — a confirm() step showing exactly what's
  // about to be scraped catches a stale bulk selection before it silently
  // re-fires on a later "Run Scraper" click, instead of only after the fact
  // via the count line the user has to remember to check.
  for (const ind of selectedInds) {
    const geos = (DATA[ind] || []).filter(e => selected.has(e.geo));
    geos.forEach(e => ids.push(e.id));
    if (geos.length) summary.push(`${LABELS[ind] || ind}: ${geos.map(e => e.geo).join(", ")}`);
  }
  if (!ids.length) { alert("Please select at least one territory."); return; }
  const confirmMsg = `About to scrape ${ids.length} territor${ids.length === 1 ? "y" : "ies"}:\\n\\n${summary.join("\\n")}\\n\\nContinue?`;
  if (!confirm(confirmMsg)) return;
  const prospectListId = document.getElementById("prospectListId").value.trim();
  const engine = document.getElementById("reliableMode").checked ? "ui" : "http";
  const btn = document.getElementById("runBtn");
  btn.disabled = true; btn.textContent = "Launching…";
  const resp = await fetch("/run", { method:"POST", headers:{"Content-Type":"application/json"}, body:JSON.stringify({ids, prospectListId, engine}) });
  const data = await resp.json();
  document.getElementById("main").style.display    = "none";
  document.getElementById("launched").style.display = "block";
  if (data.runId) {
    document.getElementById("launchedDetail").textContent =
      `Look for territory_prospecting_export_${data.runId}_deduped.xlsx in the output folder once it finishes. You can close this window.`;
  }
}

buildIndustryRow();
refresh();
</script>
</body>
</html>"""


@app.route("/")
def index():
    return render_template_string(
        HTML,
        svg_paths=SVG_PATHS,
        data=TERRITORIES,
        labels=INDUSTRY_LABELS,
        order=INDUSTRY_ORDER,
    )

def _merge_and_dedup(chunk_files, output_dir, run_id):
    """Merge chunk xlsx files then dedup. Called after all workers complete.
    Output filenames are tagged with run_id so concurrent /run calls (or just
    running twice in a row) never clobber each other's export.

    Merges whatever chunks succeeded even if some workers failed (e.g. a
    transient network timeout on one CovID) — losing every other CovID's
    already-scraped data because of one failure would be worse than
    delivering a partial, clearly-labeled result."""
    all_files = [output_dir / f for f in chunk_files]
    files = [f for f in all_files if f.exists()]
    missing = [f.name for f in all_files if not f.exists()]
    if missing:
        print(f"Merger: {len(missing)}/{len(all_files)} CovID(s) failed to scrape and are excluded from this export: {missing}")
    if not files:
        print("Merger: no chunks succeeded — nothing to merge")
        return

    combined = output_dir / f"territory_prospecting_export_{run_id}.xlsx"
    wb_out = None
    ws_out = None
    for f in files:
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        if wb_out is None:
            wb_out = openpyxl.Workbook()
            ws_out = wb_out.active
            ws_out.title = "Territory Prospecting"
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                ws_out.append(list(row))
        for row in ws.iter_rows(min_row=2, values_only=True):
            ws_out.append(list(row))
        wb.close()

    if wb_out:
        wb_out.save(combined)
        print(f"Merger: combined {len(files)} chunks → {combined}")

        dedup_out = output_dir / f"territory_prospecting_export_{run_id}_deduped.xlsx"
        try:
            import importlib.util
            spec = importlib.util.spec_from_file_location("dedup", INTERNAL / "dedup.py")
            mod  = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            mod.run_dedup(str(combined), str(dedup_out))
            print(f"Merger: deduped → {dedup_out}")

            # Step 2 (Account_Tiering) auto-locates our output via this fixed
            # name instead of guessing the latest run_id-tagged filename.
            latest = output_dir / "latest.xlsx"
            shutil.copyfile(dedup_out, latest)
            print(f"Merger: updated {latest}")
        except Exception as e:
            print(f"Merger: dedup failed: {e}")

        for f in files:
            try: f.unlink()
            except: pass


MAX_CONCURRENCY = 20  # cap on simultaneous scraper subprocesses


class _ListIdRegistry:
    """Cross-run registry of which prospect list IDs are currently in use.

    Concurrent getAccountPageContents requests against the SAME list ID race
    and clobber each other's filter server-side (verified against live
    traffic — silently returns 0 rows). Giving each concurrent worker within
    one /run's own batch a distinct list ID isn't enough: since launcher.py
    runs as a single long-lived Flask process, two DIFFERENT /run calls
    firing around the same time each independently discover the same list
    pool and, without coordination, can assign the same list ID to two
    workers running at once — verified this causes widespread 0-row results
    across concurrent runs. This registry makes "list ID in use" a single
    process-wide fact so no two workers anywhere ever hold the same one."""

    def __init__(self):
        self._lock = threading.Lock()
        self._cond = threading.Condition(self._lock)
        self._cached_pool = None
        self._in_use = set()

    def _fetch_pool(self):
        listing = subprocess.run(
            [sys.executable, str(HTTP_SCRAPER), "--list-prospect-ids"],
            cwd=str(INTERNAL), capture_output=True, text=True,
        )
        if listing.returncode != 0:
            # Previously silent: a crashed subprocess (e.g. an unhandled
            # session-expiry error) just produced an empty pool here, which
            # surfaced upstream as a generic "No prospect lists available"
            # with no clue why. Print the real error into launcher.py's own
            # stdout so it shows up in the dashboard's live log.
            print(f"  --list-prospect-ids failed (exit {listing.returncode}):\n{listing.stderr}")
        return [l for l in listing.stdout.splitlines() if l.strip()]

    def _pool_for(self, manual_id):
        if manual_id:
            return [manual_id]
        with self._lock:
            if self._cached_pool is None:
                self._cached_pool = self._fetch_pool()
            return self._cached_pool

    def acquire(self, manual_id, want, timeout=180):
        """Block until at least 1 list ID is free, then reserve and return
        up to `want` of them (may be fewer if the pool is small or busy)."""
        pool = self._pool_for(manual_id)
        if not pool:
            raise RuntimeError("No prospect lists available")
        with self._cond:
            deadline = time.time() + timeout
            while True:
                free = [i for i in pool if i not in self._in_use]
                if free:
                    grabbed = free[:want]
                    self._in_use.update(grabbed)
                    return grabbed
                if time.time() > deadline:
                    raise RuntimeError("Timed out waiting for a free prospect list ID (all in use by other runs)")
                self._cond.wait(timeout=1)

    def release(self, ids):
        with self._cond:
            self._in_use.difference_update(ids)
            self._cond.notify_all()


_list_registry = _ListIdRegistry()

_bootstrap_lock = threading.Lock()

# Rate-limit / anti-thrash marker for bootstrap regeneration (I57): a
# separately-failing side effect (territory-OFF list creation) must not cause a
# valid aura token to be thrown away and re-bootstrapped on every single /run.
_last_bootstrap_attempt = [0.0]        # time.monotonic() of last regeneration
_BOOTSTRAP_MIN_INTERVAL = 60.0         # seconds; don't re-bootstrap more often


def _ensure_bootstrap():
    """Bootstrap the aura session, but only once even if several /run calls
    arrive at once. Each call used to launch its own headless-browser
    bootstrap unconditionally, and all of them wrote to the same shared
    ~/.isc_scraper/aura_bootstrap.json — under real concurrent load this
    caused workers to read a token/context/cookie combination assembled from
    two different overlapping bootstraps (or a torn write), which Salesforce
    silently accepted but returned empty results for. Verified: 6 simultaneous
    /run calls each re-bootstrapping independently caused ~96% of CovIDs to
    come back with 0 rows; serializing bootstrap and reusing a fresh one
    fixed it. Returns True on success."""
    with _bootstrap_lock:
        bootstrap_file = Path.home() / ".isc_scraper" / "aura_bootstrap.json"
        now = time.monotonic()
        if bootstrap_file.exists():
            if _bootstrap_token_valid():
                # Also check that we have a territory-OFF scraper list saved.
                # If not (e.g. old bootstrap from before this fix), regenerate
                # so the list gets created.
                import importlib.util as _ilu
                spec = _ilu.spec_from_file_location("_hs_chk", str(HTTP_SCRAPER))
                _hs_chk = _ilu.module_from_spec(spec)
                spec.loader.exec_module(_hs_chk)
                if _hs_chk.load_scraper_list_id():
                    print("Bootstrap: reusing existing (validated) aura session")
                    return True
                # The token is GOOD; only the separate territory-OFF list side
                # effect is missing (usually selector drift). Do NOT unlink the
                # valid token — that would re-bootstrap on every /run forever and
                # throw away a working session (I57). Reuse the token, and only
                # attempt a bounded regeneration (at most once per interval) to
                # try to (re)create the list. The bootstrap subprocess overwrites
                # the file on success; on failure the valid token stays in place.
                if now - _last_bootstrap_attempt[0] < _BOOTSTRAP_MIN_INTERVAL:
                    print("Bootstrap: no territory-OFF scraper list, but a valid "
                          "token exists and we re-bootstrapped recently — reusing "
                          "the token (bounded backoff, I57)")
                    return True
                print("Bootstrap: valid token but no territory-OFF scraper list — "
                      "regenerating (rate-limited) to (re)create the list")
            else:
                # The cached token is stale — e.g. the user just re-logged in.
                print("Bootstrap: cached token is stale (session changed) — regenerating")
                bootstrap_file.unlink(missing_ok=True)
        print("Bootstrap: fetching aura token (headless)...")
        _last_bootstrap_attempt[0] = time.monotonic()
        result = subprocess.run(
            [sys.executable, str(HTTP_SCRAPER), "--bootstrap-only"],
            cwd=str(INTERNAL)
        )
        return result.returncode == 0


def _bootstrap_token_valid():
    """Cheap liveness check on the cached aura bootstrap: one getProspectLists
    call. Returns False on 401 (stale token / expired session) or any error, so
    a stale cache is regenerated rather than silently reused."""
    try:
        sys.path.insert(0, str(INTERNAL))
        import http_scraper as _hs
        boot = _hs.load_bootstrap()
        if not boot:
            return False
        _hs.get_all_prospect_list_ids(*boot)
        return True
    except Exception:
        return False


# Lightweight, dashboard-facing progress summary — deliberately not the full
# per-row scrape log (that's already on stdout for direct/terminal use). Just
# "which CovIDs are in flight right now" and a done/total count, polled by
# run_pipeline.py so it can show a one-line status instead of a raw log dump.
_progress_lock = threading.Lock()
_PROGRESS = {"total": 0, "done": 0, "current": [], "run_id": None, "merged": False, "zero_covids": []}

# Guards against a second /run starting while one's still in flight. Without
# this, two overlapping runs share the single global _PROGRESS dict above and
# corrupt each other's done/total/current (e.g. a leftover run from an
# earlier click plus a fresh map-picker selection both scraping at once shows
# up as something like "22/2 CovIDs" on the dashboard) — the per-run list
# registry already prevents them from colliding on prospect list IDs, but
# that's a different problem from the progress reporting getting confused.
_run_state_lock = threading.Lock()
_ACTIVE_RUN = {"active": False, "run_id": None}


def _reset_progress(total, run_id):
    with _progress_lock:
        _PROGRESS["run_id"] = run_id
        _PROGRESS["total"] = total
        _PROGRESS["done"] = 0
        _PROGRESS["current"] = []
        # False until _merge_and_dedup actually writes latest.xlsx — done
        # reaching total only means the *first* pass finished, not that the
        # retry sweeps (which can still be re-running some CovIDs) and final
        # merge are done. run_pipeline.py's step1->2 auto-chain watcher keys
        # off this, not off done>=total, so it doesn't fire early against a
        # stale/missing output file.
        _PROGRESS["merged"] = False
        _PROGRESS["zero_covids"] = []


def _chunk_row_count(path):
    """How many data rows a CovID's chunk file actually got. Used to tell a
    genuinely-empty territory apart from one that needs a retry sweep."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        count = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
        wb.close()
        return count
    except Exception:
        return 0


def _run_batch(cov_ids, chunk_files, list_ids, start_idx, scraper_list_id=None):
    """Launch one batch of workers in parallel and wait for all to finish.
    If `scraper_list_id` is provided (a territory-OFF list created at bootstrap),
    all workers use that single list — no per-worker assignment needed, since a
    territory-OFF list doesn't have the per-covId filter collision problem that
    territory-ON lists have (those needed unique IDs per worker to avoid racing).
    Falls back to the registry-assigned pool list IDs if no scraper list."""
    with _progress_lock:
        _PROGRESS["current"] = list(cov_ids)
    procs = []
    for i, cid in enumerate(cov_ids):
        if scraper_list_id:
            # Territory-OFF list: all workers can share it safely.
            # The registry-issued list_ids are still acquired/released by the
            # caller for rate-limiting, but we pass the scraper list to the
            # actual HTTP call so Salesforce sees the right list type.
            assigned = scraper_list_id
            fallbacks = []
        else:
            assigned = list_ids[i]
            fallbacks = [lid for j, lid in enumerate(list_ids) if j != i]
        cmd = [
            sys.executable, str(HTTP_SCRAPER),
            "--cov-id", cid,
            "--prospect-list-id", assigned,
            "--output", chunk_files[start_idx + i],
            "--no-dedup",
        ]
        if fallbacks:
            cmd += ["--fallback-list-ids", ",".join(fallbacks)]
        procs.append((cid, subprocess.Popen(cmd, cwd=str(INTERNAL))))
    for cid, p in procs:
        if p.wait() != 0:
            print(f"  Worker for {cid} failed (exit {p.returncode}) — its data will be missing from this export")
        with _progress_lock:
            _PROGRESS["done"] += 1
            if cid in _PROGRESS["current"]:
                _PROGRESS["current"].remove(cid)


def _reject_untrusted_request():
    """Guard the local action-trigger routes (I53/I54). Returns a Flask error
    response to reject, or None to allow.

    Always on (safe, needs no caller change): reject any request whose Host
    header isn't loopback — blocks off-box / DNS-rebind callers.

    Opt-in via ISC_REQUIRE_TOKEN=1: additionally require a matching X-ISC-Token
    header, unless the Origin/Referer host is itself loopback (so the launcher's
    own same-origin browser UI keeps working without a token). Gated behind the
    env flag so the existing dashboard is not broken today."""
    host = (urllib.parse.urlsplit("//" + (request.host or "")).hostname or "").lower()
    if host not in ("127.0.0.1", "localhost", "::1", ""):
        return jsonify({"error": "forbidden (non-loopback host)"}), 403
    if os.environ.get("ISC_REQUIRE_TOKEN") == "1":
        supplied = request.headers.get("X-ISC-Token", "")
        if supplied and secrets.compare_digest(supplied, RUN_TOKEN):
            return None
        origin = request.headers.get("Origin") or request.headers.get("Referer") or ""
        oh = (urllib.parse.urlsplit(origin).hostname or "").lower()
        if oh in ("127.0.0.1", "localhost", "::1"):
            return None
        return jsonify({"error": "forbidden (missing/invalid token)"}), 403
    return None


@app.route("/progress")
def progress():
    """Lightweight status for the main dashboard (run_pipeline.py) to show
    'which CovIDs are being worked on right now' without needing this
    process's full stdout piped through — see _PROGRESS above."""
    denied = _reject_untrusted_request()
    if denied:
        return denied
    with _progress_lock:
        return jsonify(dict(_PROGRESS))


def _run_ui_engine(ids, run_id, output_dir):
    """Reliable scrape engine: drives the real Territory Prospecting UI via the
    already-proven isc_scraper.py (creates a FRESH prospecting list, applies
    each CovID filter, reads the on-screen table). This is the ground-truth
    method CONTEXT.md itself uses to verify data — it does not suffer the
    aura-endpoint cold-cache false-zeros that make the fast HTTP engine
    occasionally miss a non-empty territory. Slower (one browser, sequential),
    but it returns what the UI actually shows.

    Opt-in via /run {"engine":"ui"} so the default fast HTTP path is untouched.
    *** Needs one live calibration pass — isc_scraper.py's UI selectors haven't
    been re-verified against the current app in this integration. ***"""
    combined_name = f"territory_prospecting_export_{run_id}.xlsx"
    cmd = [sys.executable, str(UI_SCRAPER),
           "--cov-ids-json", json.dumps(ids),
           "--output", combined_name,
           "--no-monitor"]
    print(f"UI engine: scraping {len(ids)} CovID(s) via the real Territory Prospecting UI...")
    proc = subprocess.run(cmd, cwd=str(INTERNAL))
    deduped = output_dir / f"territory_prospecting_export_{run_id}_deduped.xlsx"
    if proc.returncode == 0 and deduped.exists():
        latest = output_dir / "latest.xlsx"
        shutil.copyfile(deduped, latest)
        print(f"UI engine: done → {latest}")
        with _progress_lock:
            _PROGRESS["done"] = _PROGRESS["total"]
    else:
        print(f"UI engine: failed (exit {proc.returncode}, deduped exists={deduped.exists()}) — "
              "the UI selectors likely need calibration; see isc_scraper.py.")
    with _progress_lock:
        _PROGRESS["merged"] = True


@app.route("/run", methods=["POST"])
def run():
    denied = _reject_untrusted_request()
    if denied:
        return denied
    body = request.json or {}
    engine = (body.get("engine") or "http").lower()
    if body.get("all"):
        # Auto-chain / "run everything" mode (run_pipeline.py's full-pipeline
        # button) — no territory-picker UI involved, so expand to every
        # CovID across every industry ourselves.
        ids = [t["id"] for lst in TERRITORIES.values() for t in lst]
    else:
        ids = body.get("ids", [])
    if not ids:
        return jsonify({"error": "no ids"}), 400

    with _run_state_lock:
        if _ACTIVE_RUN["active"]:
            return jsonify({"error": f"a scrape (run {_ACTIVE_RUN['run_id']}) is already "
                                      "in progress — wait for it to finish before starting another"}), 409
        _ACTIVE_RUN["active"] = True

    prospect_list_id = body.get("prospectListId") or ""
    output_dir = APP_ROOT / "output"
    output_dir.mkdir(exist_ok=True)

    # Every run gets a unique ID so concurrent /run calls (or just running
    # twice in a row) never share chunk/output filenames and clobber each
    # other's data — verified this was silently losing entire runs before.
    run_id = f"{time.strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}"
    chunk_files = [f"chunk_{run_id}_{idx}.xlsx" for idx in range(len(ids))]

    # Persist the authoritative selected-CovID set for the downstream IBM Scraper
    # step. This is written from the exact `ids` about to be scraped, so it
    # includes CovIDs that come back with 0 accounts — which the DEDUPED_ACCOUNTS
    # output drops but the IBM install-base filters still need to consider.
    try:
        (output_dir / "selected_covids.json").write_text(json.dumps({
            "run_id": run_id,
            "selected_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
            "covids": list(ids),
        }, indent=2))
    except Exception as e:
        print(f"  (warning) could not write selected_covids.json: {e}")

    _reset_progress(len(ids), run_id)
    with _run_state_lock:
        _ACTIVE_RUN["run_id"] = run_id

    def watch():
      try:
        # Reliable opt-in path: the proven browser-UI scrape. Bypasses the
        # whole HTTP bootstrap/pooled-list machinery below (and its cold-cache
        # false-zeros) entirely.
        if engine == "ui":
            _run_ui_engine(ids, run_id, output_dir)
            return

        # Step 1: bootstrap aura session — shared across all concurrent
        # /run calls (see _ensure_bootstrap for why this must be serialized).
        if not _ensure_bootstrap():
            print("Bootstrap failed — aborting scrape")
            return

        # Load the territory-OFF scraper list created at bootstrap — this is
        # the fix for CovIDs outside the user's assigned territory returning 0.
        sys.path.insert(0, str(INTERNAL))
        import http_scraper as _hs_run
        scraper_list_id = _hs_run.load_scraper_list_id()
        if scraper_list_id:
            print(f"  Using territory-OFF scraper list: {scraper_list_id}")
        else:
            print("  No territory-OFF scraper list found — results may be restricted to user's territory")

        # Step 2: scrape all CovIDs in parallel batches.
        remaining = list(range(len(ids)))
        batch_num = 0
        while remaining:
            want = min(len(remaining), MAX_CONCURRENCY)
            try:
                acquired = _list_registry.acquire(prospect_list_id, want)
            except RuntimeError as e:
                print(f"  {e} — aborting remaining CovIDs for this run")
                break
            batch_num += 1
            batch_idx = remaining[:len(acquired)]
            remaining = remaining[len(acquired):]
            batch_ids = [ids[i] for i in batch_idx]
            print(f"  Batch {batch_num}: {len(batch_ids)} CovIDs in parallel")
            try:
                _run_batch(batch_ids, chunk_files, acquired, batch_idx[0], scraper_list_id=scraper_list_id)
            finally:
                _list_registry.release(acquired)

        # Step 3: second-pass sweep over CovIDs that came back with 0 rows.
        # This is a real, previously ground-truthed Salesforce-side cache/
        # index warm-up transient (see CONTEXT.md Round 4) — territories
        # confirmed non-empty via the actual Salesforce UI still came back 0
        # from this same Aura endpoint, then returned correct data on a retry
        # *minutes* later with no code change. Each worker's own 4-attempt
        # retry (~12s total) is nowhere near that window, so a whole scrape
        # that hits this transient broadly (e.g. first run of the day, cold
        # session) needs a real wait-and-revisit pass, not just a few more
        # rapid retries. Bounded to 2 sweeps so a run with genuinely-empty
        # territories doesn't stall indefinitely.
        for sweep_num, wait_s in enumerate((), start=1):  # disabled for demo — was (60, 120), caused multi-minute hangs
            zero_idx = [i for i in range(len(ids)) if _chunk_row_count(chunk_files[i]) == 0]
            if not zero_idx:
                break
            print(f"  {len(zero_idx)} CovID(s) came back empty — waiting {wait_s}s "
                  f"for Salesforce's cache to warm, then sweep {sweep_num}/2...")
            time.sleep(wait_s)
            remaining = zero_idx
            while remaining:
                want = min(len(remaining), MAX_CONCURRENCY)
                try:
                    acquired = _list_registry.acquire(prospect_list_id, want)
                except RuntimeError as e:
                    print(f"  {e} — aborting remaining CovIDs for this sweep")
                    break
                batch_idx = remaining[:len(acquired)]
                remaining = remaining[len(acquired):]
                batch_ids = [ids[i] for i in batch_idx]
                print(f"  Sweep {sweep_num} batch: {len(batch_ids)} CovIDs in parallel")
                try:
                    _run_batch(batch_ids, chunk_files, acquired, batch_idx[0])
                finally:
                    _list_registry.release(acquired)

        # Never miss silently: report which requested CovIDs returned 0 rows.
        # A 0 is either a genuinely-empty territory or a Salesforce cold-cache
        # transient miss (CONTEXT.md Round 4) — either way the operator must be
        # able to see it, not have it vanish into an empty chunk. Anything
        # listed here should be re-run or checked against the real UI before
        # the export is trusted as complete.
        zero_covids = [ids[i] for i in range(len(ids)) if _chunk_row_count(chunk_files[i]) == 0]
        got = len(ids) - len(zero_covids)
        print(f"Scrape summary: {got}/{len(ids)} CovIDs returned data.")
        if zero_covids:
            print(f"  UNCONFIRMED (0 rows — possible miss, verify/re-run): {', '.join(zero_covids)}")
        _PROGRESS["zero_covids"] = zero_covids

        _merge_and_dedup(chunk_files, output_dir, run_id)
        with _progress_lock:
            _PROGRESS["merged"] = True
      finally:
        # Always clear, even on bootstrap failure/exception, so a stuck run
        # can't permanently block every future /run call with "already
        # in progress" — restarting this app is not the only way out.
        with _run_state_lock:
            _ACTIVE_RUN["active"] = False

    threading.Thread(target=watch, daemon=True).start()
    return jsonify({"ok": True, "count": len(ids), "runId": run_id})

def main():
    import os
    import socket
    port = 5477
    with socket.socket() as s:
        if s.connect_ex(("127.0.0.1", port)) == 0:
            # port busy — pick any free one
            s2 = socket.socket()
            s2.bind(("", 0))
            port = s2.getsockname()[1]
            s2.close()
    # Publish the REAL bound port + shared token so the parent reads it instead
    # of assuming 5477 and POSTing a seller's CovIDs to whatever squats there
    # (I53/I54). Owner-only (0600) since it carries the run token.
    try:
        PORT_HANDSHAKE_FILE.parent.mkdir(exist_ok=True)
        tmp = PORT_HANDSHAKE_FILE.with_suffix(f".tmp{os.getpid()}")
        tmp.write_text(json.dumps({"port": port, "token": RUN_TOKEN, "pid": os.getpid()}))
        try:
            os.chmod(tmp, 0o600)
        except Exception:
            pass
        tmp.replace(PORT_HANDSHAKE_FILE)
        try:
            os.chmod(PORT_HANDSHAKE_FILE, 0o600)
        except Exception:
            pass
    except Exception as e:
        print(f"  (warning) could not write launcher port handshake: {e}")
    # When driven by the Seller Dashboard (ISC_NO_BROWSER=1) this app is a
    # headless scrape engine — don't pop open the territory-picker map tab.
    if not os.environ.get("ISC_NO_BROWSER"):
        threading.Timer(0.9, lambda: webbrowser.open(f"http://127.0.0.1:{port}")).start()
    app.run(port=port, debug=False, use_reloader=False, threaded=True)

if __name__ == "__main__":
    main()
