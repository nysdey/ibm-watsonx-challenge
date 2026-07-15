"""Load Step 2's output, write Step 3's output — same dated + latest.xlsx
convention as every other step, see ../SCHEMA_CONTRACT.md."""
import shutil
from datetime import date
from pathlib import Path

import openpyxl

REQUIRED_STEP2_COLUMNS = ["Account Name", "Tier", "Tier_Score"]


class SchemaError(RuntimeError):
    pass


def load_step2_accounts(path):
    path = Path(path)
    if not path.exists():
        raise SchemaError(
            f"Step 2 output not found at {path}. Run Account_Tiering/run.py first — "
            f"it writes output/latest.xlsx on every successful run."
        )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Tiered Accounts" not in wb.sheetnames:
        raise SchemaError(
            f"{path} has no 'Tiered Accounts' sheet (found: {wb.sheetnames}). "
            f"This isn't a Step 2 export — see ../SCHEMA_CONTRACT.md."
        )
    ws = wb["Tiered Accounts"]
    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))

    missing = [c for c in REQUIRED_STEP2_COLUMNS if c not in header]
    if missing:
        raise SchemaError(
            f"{path} is missing required column(s) {missing}. "
            f"Found columns: {header}. See ../SCHEMA_CONTRACT.md for the Step 2 contract."
        )

    rows = []
    for r in rows_iter:
        row = dict(zip(header, r))
        if row.get("Account Name") and row.get("Tier") is not None:
            rows.append(row)
    wb.close()
    return header, rows


def write_dated_and_latest(rows, header, output_dir, prefix, sheet_name):
    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)
    dated_name = output_dir / f"{prefix}_{date.today().strftime('%Y%m%d')}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(header)
    for row in rows:
        ws.append([_sanitize_cell(row.get(col)) for col in header])
    wb.save(dated_name)

    latest = output_dir / "latest.xlsx"
    shutil.copyfile(dated_name, latest)
    return dated_name, latest


def _sanitize_cell(value):
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value
